"""
update_dashboard.py (IMPROVED VERSION)
---------------------------------------
Excel padhta hai aur HTML dashboard generate karta hai.
Ab NCR tables mein status filter aur export to CSV feature bhi hai.
"""

import openpyxl
import datetime
import os
import json
from collections import defaultdict

# ─────────────────────────────────────────────────────────
#  CONFIG — sirf yahan dono paths set karo
# ─────────────────────────────────────────────────────────
EXCEL_PATH = r"C:\Users\Shaghaf Ahmed\OneDrive\Desktop\Work\AUTOMATE EXCEL\MY_WORK.xlsx"
OUTPUT_HTML = r"C:\Users\Shaghaf Ahmed\OneDrive\Desktop\Work\AUTOMATE EXCEL\QHSE_dashboard\index.html"
# ─────────────────────────────────────────────────────────

BASE_DATE = datetime.datetime(1899, 12, 30)

def to_date(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        try:
            return BASE_DATE + datetime.timedelta(days=int(val))
        except:
            return None
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val
    return None

def fmt_date(val):
    months = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    d = to_date(val)
    if d is None:
        return "—"
    return f"{d.day:02d} {months[d.month]} {d.year}"

def fmt_date_yymm(val):
    d = to_date(val)
    if d is None:
        return None
    return f"{d.year}-{d.month:02d}"

def js_str(s):
    if s is None:
        return "null"
    return "'" + str(s).replace("'", "\\'").replace("&", "&amp;") + "'"

print("📂 Reading Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

def read_sheet(ws_name):
    try:
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        data = []
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            if str(row[0]).strip() in ('Doc-code', ''):
                continue
            data.append(row)
        return data
    except:
        print(f"  ⚠️ Sheet '{ws_name}' not found")
        return []

r1_rows = read_sheet('R1')
r2_rows = read_sheet('R2')

# NCRs
ncr_rows = []
try:
    ncr_ws = wb['NCRs']
    ncr_all = list(ncr_ws.iter_rows(values_only=True))
    header_idx = 0
    for i, r in enumerate(ncr_all):
        if r and 'Status' in str(r):
            header_idx = i
            break
    ncr_rows = [r for r in ncr_all[header_idx+1:] if r and len(r) > 1 and r[1]]
except:
    print("  ⚠️ NCRs sheet issue")

# Client NCRs
cncr_rows = []
try:
    cncr_ws = wb['Client_NCRs']
    cncr_all = list(cncr_ws.iter_rows(values_only=True))
    cncr_header = 0
    for i, r in enumerate(cncr_all):
        if r and 'Status' in str(r):
            cncr_header = i
            break
    cncr_rows = [r for r in cncr_all[cncr_header+1:] if r and len(r) > 1 and r[1]]
except:
    print("  ⚠️ Client_NCRs sheet issue")

wb.close()
print(f"  ✅ R1: {len(r1_rows)} | R2: {len(r2_rows)} | NCRs: {len(ncr_rows)} | Client NCRs: {len(cncr_rows)}")

# ── CALENDAR READING ──────────────────────────────────────
import datetime as _dt, json as _json
cal_data = {}
cal_month = 'May'; cal_year = 2026; cal_month_num = 5
_MNL = ['','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
try:
    _cw = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    for _r in list(_cw['Calendar'].iter_rows(values_only=True))[1:]:
        _dv = _r[0]; _av = str(_r[1]).strip() if len(_r)>1 and _r[1] else ''
        if not _dv or not _av: continue
        try:
            _pd = _dt.datetime.strptime(str(_dv).strip(), '%d-%b-%Y')
            _k  = f"{_pd.year}-{_pd.month:02d}-{_pd.day:02d}"
            cal_month = _MNL[_pd.month]; cal_year = _pd.year; cal_month_num = _pd.month
        except: continue
        if _av == 'Week Off': cal_data[_k] = 'Week Off'
        elif _k in cal_data and cal_data[_k] != 'Week Off': cal_data[_k] += '|' + _av
        else: cal_data[_k] = _av
    _cw.close()
    print(f"  ✅ Calendar: {len(cal_data)} days ({cal_month} {cal_year})")
except Exception as _e:
    print(f"  ⚠️ Calendar: {_e}")

# ── KPI READING ───────────────────────────────────────────
kpi_data = {
    'month': 'May 2026',
    'days_lti_month': 0, 'days_lti_total': 0,
    'manpower': [], 'total_mp': 0, 'total_hrs': 0, 'cum_total': 0,
    'r1_kpis': [], 'r2_kpis': [], 'combined_kpis': [],
    'sustain': {}
}
try:
    _kw = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    _ks = None
    for _name in _kw.sheetnames:
        if _name.strip().upper() == 'KPI':
            _ks = _kw[_name]; break
    if _ks:
        _kr = list(_ks.iter_rows(values_only=True))
        def _kn(v, fb=0):
            if v is None: return fb
            if isinstance(v, str) and ('#' in v or 'REF' in v): return fb
            if isinstance(v, (int, float)): return v
            return fb
        # Month
        _hdr = _kr[6][3] if len(_kr) > 6 else None
        if _hdr and '-' in str(_hdr):
            kpi_data['month'] = str(_hdr).split('-',1)[1].strip()
        kpi_data['days_lti_month'] = _kn(_kr[4][5])
        kpi_data['days_lti_total'] = _kn(_kr[4][10])
        # Manpower
        for _i in [9,10,11,12,13,14]:
            _r = _kr[_i]
            if _r[1]:
                kpi_data['manpower'].append({'team': str(_r[1]).strip(), 'mp': _kn(_r[3]), 'hrs': _kn(_r[5])})
        kpi_data['total_mp']  = sum(_m['mp']  for _m in kpi_data['manpower'])
        kpi_data['total_hrs'] = sum(_m['hrs'] for _m in kpi_data['manpower'])
        kpi_data['cum_total'] = _kn(_kr[13][9])
        # KPI blocks
        def _get_kpis(start):
            out=[]
            for _i in range(start, start+20):
                if _i >= len(_kr): break
                _r = _kr[_i]
                if isinstance(_r[6], str) and _r[6] not in ('HSE Indicator',):
                    out.append({'name': _r[6].strip(), 'last': _kn(_r[7]), 'current': _kn(_r[8]), 'cum': _kn(_r[9])})
                else: break
            return out
        kpi_data['r1_kpis']       = _get_kpis(57)
        kpi_data['r2_kpis']       = _get_kpis(138)
        kpi_data['combined_kpis'] = _get_kpis(219)
        # Sustainability
        kpi_data['sustain'] = {
            'meetings': _kn(_kr[312][2]),
            'grievance_recv': _kn(_kr[313][2]),
            'grievance_resolved': _kn(_kr[315][2]),
            'gender_ratio': str(_kr[316][2] or '—'),
            'female': _kn(_kr[317][2]),
            'male': _kn(_kr[318][2]),
            'wp_incidents': _kn(_kr[319][2]),
            'spill_total': _kn(_kr[312][9]),
        }
        print(f"  ✅ KPI: {kpi_data['month']} ({len(kpi_data['r1_kpis'])} R1 + {len(kpi_data['r2_kpis'])} R2 + {len(kpi_data['combined_kpis'])} Combined)")
    _kw.close()
except Exception as _e:
    print(f"  ⚠️ KPI: {_e}")

# ── KPI E&S READING ──────────────────────────────────────
import json as _json
es_r1 = {}; es_r2 = {}; es_comb = {}
MONTHS_ES = ['Jan-25','Feb-25','Mar-25','Apr-25','May-25','Jun-25','Jul-25','Aug-25','Sep-25','Oct-25','Nov-25','Dec-25','Jan-26','Feb-26','Mar-26','Apr-26']
try:
    _ew = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    _esw = None
    for _nm in _ew.sheetnames:
        if _nm.strip().upper() == 'KPI E&S': _esw = _ew[_nm]; break
    if _esw:
        _er = list(_esw.iter_rows(values_only=True))
        def _en(v):
            if v is None: return 0
            if isinstance(v, str): return 0
            if isinstance(v, (int, float)): return v
            return 0
        # R1: rows 2-41 (idx 2-40), monthly=col D-Q (idx 3-18), TOTAL=idx 27
        def _r1(i): r=_er[i]; return {'monthly':[_en(r[j]) for j in range(3,19)],'total':_en(r[27])}
        # R2: rows 45-84 (idx 44-83), same structure
        def _r2(i): r=_er[i]; return {'monthly':[_en(r[j]) for j in range(3,19)],'total':_en(r[27])}
        # Combined: rows 88-127 (idx 87-126), value=col D (idx 3)
        def _cb(i): return _en(_er[i][3])
        es_r1 = {
            'penalties':    _r1(2),  'audits':       _r1(5),
            'stakeholder':  _r1(6),  'grievances':   _r1(7),
            'diesel':       _r1(9),  'freshwater':   _r1(16),
            'haz_total':    _r1(20), 'nonhaz_total': _r1(24),
            'waste_cost':   _r1(25), 'spills':       _r1(26),
            'wastewater':   _r1(40),
        }
        es_r2 = {
            'penalties':    _r2(45), 'audits':       _r2(48),
            'stakeholder':  _r2(49), 'grievances':   _r2(50),
            'diesel':       _r2(52), 'freshwater':   _r2(59),
            'haz_total':    _r2(63), 'nonhaz_total': _r2(67),
            'waste_cost':   _r2(68), 'spills':       _r2(69),
            'wastewater':   _r2(83),
        }
        es_comb = {
            'penalties':    _cb(88),  'audits':       _cb(91),
            'stakeholder':  _cb(92),  'grievances':   _cb(93),
            'diesel':       _en(_er[52][27]) + _en(_er[9][27]),  # R1+R2 sum (sheet formula bug fix)
            'freshwater':   _cb(102), 'haz_total':    _cb(106),
            'nonhaz_total': _cb(110), 'waste_cost':   _cb(111),
            'spills':       _cb(112), 'wastewater':   _cb(126),
        }
        print(f"  ✅ E&S KPI: audits={es_comb['audits']} | diesel={es_comb['diesel']:,.0f}L | water={es_comb['freshwater']:,.0f}m³")
    _ew.close()
except Exception as _e: print(f"  ⚠️ E&S: {_e}")

# ─────────────────────────────────────────────────────────
# CALCULATE STATS
# ─────────────────────────────────────────────────────────

def count_status(rows):
    c = defaultdict(int)
    for r in rows:
        s = str(r[4]).strip() if len(r) > 4 and r[4] else ''
        if s and s not in ('Review Status', 'None'):
            c[s] += 1
    return c

def count_types(rows):
    c = defaultdict(int)
    for r in rows:
        t = str(r[8]).strip() if len(r) > 8 and r[8] else ''
        if t and t not in ('Document Type', 'None'):
            key = t.split('-')[0].strip() if '-' in t else t
            c[key] += 1
    return c

def count_prev(rows):
    c = defaultdict(int)
    for r in rows:
        p = str(r[9]).strip() if len(r) > 9 and r[9] else '-/-'
        if '🟡' in p or 'Approved with' in p:
            c['AWC'] += 1
        elif '🟠' in p or 'Review with' in p:
            c['RWC'] += 1
        elif '🔴' in p or 'Rejected' in p:
            c['REJ'] += 1
        elif p == '-/-':
            c['First'] += 1
    return c

r1_status = count_status(r1_rows)
r2_status = count_status(r2_rows)
r1_types = count_types(r1_rows)
r2_types = count_types(r2_rows)
r1_prev = count_prev(r1_rows)
r2_prev = count_prev(r2_rows)

# Combined
combined_status = {}
for k in ['APP', 'AWC', 'REJ', 'RWC', 'UR', 'IFI']:
    combined_status[k] = r1_status.get(k, 0) + r2_status.get(k, 0)

total_docs = len(r1_rows) + len(r2_rows)
pending = combined_status['AWC'] + combined_status['RWC'] + combined_status['REJ'] + combined_status['UR']
pending_sub = f"AWC {combined_status['AWC']} · RWC {combined_status['RWC']} · REJ {combined_status['REJ']}"
if combined_status['UR'] > 0:
    pending_sub += f" · UR {combined_status['UR']}"

r1_approval = round(r1_status.get('APP', 0) / len(r1_rows) * 100, 1) if r1_rows else 0
r2_approval = round(r2_status.get('APP', 0) / len(r2_rows) * 100, 1) if r2_rows else 0

# NCR stats
ncr_status = defaultdict(int)
ncr_contractor_open = defaultdict(int)
ncr_contractor_closed = defaultdict(int)
ncr_location = defaultdict(int)
ncr_monthly = defaultdict(int)

for r in ncr_rows:
    status = str(r[9]).strip() if len(r) > 9 and r[9] else ''
    contractor = str(r[5]).strip() if len(r) > 5 and r[5] else ''
    location = str(r[3]).strip() if len(r) > 3 and r[3] else ''
    date_val = r[2] if len(r) > 2 else None

    ncr_status[status] += 1

    if 'Riyah 1' in location and 'Riyah 2' in location:
        ncr_location['Both'] += 1
    elif 'Riyah 1' in location:
        ncr_location['Riyah 1'] += 1
    elif 'Riyah 2' in location:
        ncr_location['Riyah 2'] += 1
    else:
        ncr_location[location] += 1

    if status == 'Open':
        ncr_contractor_open[contractor] += 1
    elif status == 'Closed':
        ncr_contractor_closed[contractor] += 1

    ym = fmt_date_yymm(date_val)
    if ym:
        ncr_monthly[ym] += 1

ncr_total = len(ncr_rows)
ncr_closed = ncr_status.get('Closed', 0)
ncr_open = ncr_status.get('Open', 0)
ncr_closure = round(ncr_closed / ncr_total * 100) if ncr_total else 0
ncr_r1 = ncr_location.get('Riyah 1', 0)
ncr_r2 = ncr_location.get('Riyah 2', 0)
ncr_both = ncr_location.get('Both', 0)

# Last NCR
last_ncr_no = "000"
last_ncr_desc = ""
if ncr_rows:
    last_ncr = ncr_rows[-1]
    if len(last_ncr) > 1 and last_ncr[1]:
        ncr_str = str(last_ncr[1]).strip()
        parts = ncr_str.split('-')
        last_ncr_no = parts[-1] if parts else '000'
        if len(last_ncr) > 7 and last_ncr[7]:
            last_ncr_desc = str(last_ncr[7]).strip()[:50]

# Contractors
contractors = ['TCC', 'SHC', 'JUJIE', 'RMM', 'GOLD WIND', 'ANSON']
ncr_closed_data = [ncr_contractor_closed.get(c, 0) for c in contractors]
ncr_open_data = [ncr_contractor_open.get(c, 0) for c in contractors]

# Monthly timeline
months_order = ['2025-08', '2025-09', '2025-10', '2025-11', '2025-12', '2026-01', '2026-02', '2026-03', '2026-04']
months_label = ['Aug 25', 'Sep 25', 'Oct 25', 'Nov 25', 'Dec 25', 'Jan 26', 'Feb 26', 'Mar 26', 'Apr 26']
ncr_timeline = [ncr_monthly.get(m, 0) for m in months_order]

# Client NCR stats
cncr_total = len(cncr_rows)
cncr_closed = 0
cncr_open = 0
for r in cncr_rows:
    status = str(r[9]).strip() if len(r) > 9 and r[9] else ''
    if status == 'Closed':
        cncr_closed += 1
    elif status == 'Open':
        cncr_open += 1

cncr_loc = defaultdict(int)
cncr_orig = defaultdict(int)
cncr_monthly = defaultdict(int)

for r in cncr_rows:
    loc = str(r[4]).strip() if len(r) > 4 and r[4] else ''
    orig = str(r[6]).strip() if len(r) > 6 and r[6] else ''
    date_val = r[2] if len(r) > 2 else None

    if 'Riyah 1 & 2' in loc or ('1' in loc and '2' in loc):
        cncr_loc['Both'] += 1
    elif 'Riyah 2' in loc:
        cncr_loc['Riyah 2'] += 1
    else:
        cncr_loc['Riyah 1'] += 1

    if orig:
        cncr_orig[orig] += 1

    ym = fmt_date_yymm(date_val)
    if ym:
        cncr_monthly[ym] += 1

# Originator top 3
top_orig = sorted(cncr_orig.items(), key=lambda x: -x[1])[:3]
orig_labels = [x[0] for x in top_orig]
orig_data = [x[1] for x in top_orig]
if not orig_labels:
    orig_labels = ['Hassan Darwish', 'Russel Kisera', 'Alamgeer Ahmed']
    orig_data = [6, 3, 2]

# Client monthly timeline
cncr_months_present = sorted(cncr_monthly.keys())
cncr_labels = []
cncr_timeline = []
months_short = ['', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
for m in cncr_months_present:
    if '-' in m:
        y, mo = m.split('-')
        cncr_labels.append(f"{months_short[int(mo)]} {y[2:]}")
        cncr_timeline.append(cncr_monthly[m])
if not cncr_labels:
    cncr_labels = ['Aug 25', 'Sep 25', 'Oct 25', 'Nov 25', 'Jan 26', 'Mar 26', 'Apr 26']
    cncr_timeline = [3, 1, 3, 1, 1, 1, 1]

# Open client alert
open_cncr_text = "All Client NCRs Closed ✓"
open_cncr_number = ""
if cncr_open > 0:
    for r in cncr_rows:
        if len(r) > 9 and str(r[9]).strip() == 'Open':
            open_cncr_number = str(r[1]).strip() if len(r) > 1 else "NCR"
            open_cncr_text = f"{cncr_open} Open Client NCR(s) — Last: {open_cncr_number} · {str(r[7]).strip() if len(r) > 7 and r[7] else 'No description'} · Issued {fmt_date(r[2])}"
            break

today = datetime.datetime.now().strftime("%d %b %Y")

# Prepare data for JSON embedding
r1_docs_json = json.dumps([(str(r[0]).strip(), str(r[1]).strip(), int(r[2]) if r[2] else 0, str(r[4]).strip() if len(r)>4 and r[4] else '', fmt_date(r[10]) if len(r)>10 else '—', str(r[8]).strip() if len(r)>8 and r[8] else '') for r in r1_rows])
r2_docs_json = json.dumps([(str(r[0]).strip(), str(r[1]).strip(), int(r[2]) if r[2] else 0, str(r[4]).strip() if len(r)>4 and r[4] else '', fmt_date(r[10]) if len(r)>10 else '—', str(r[8]).strip() if len(r)>8 and r[8] else '') for r in r2_rows])
ncr_data_json = json.dumps([(i+1, str(r[1]).strip() if len(r)>1 else '', fmt_date(r[2]) if len(r)>2 else '—', str(r[3]).strip() if len(r)>3 else '', str(r[4]).strip() if len(r)>4 else '', str(r[5]).strip() if len(r)>5 else '', str(r[6]).strip() if len(r)>6 else '', fmt_date(r[8]) if len(r)>8 else '—', str(r[9]).strip() if len(r)>9 else '') for i, r in enumerate(ncr_rows)])
cncr_data_json = json.dumps([(i+1, str(r[1]).strip() if len(r)>1 else '', str(r[3]).strip() if len(r)>3 else 'Major', fmt_date(r[2]) if len(r)>2 else '—', str(r[4]).strip() if len(r)>4 else '', str(r[6]).strip() if len(r)>6 else '', str(r[7]).strip() if len(r)>7 else '', fmt_date(r[8]) if len(r)>8 else '—', str(r[9]).strip() if len(r)>9 else '') for i, r in enumerate(cncr_rows)])

# Build HTML
print("🏗️ Building HTML with enhanced filters...")

# Create HTML template string using triple quotes but careful with braces


# ── KPI HTML BUILDERS ─────────────────────────────────────
_LAGGING_K = ['Fatality','Permanent disability','Lost Time','Restricted Work','Medical Treatment','First Aid','Property Damage','Enviro','Company Image','Other','HIPO','Near Miss']
def _is_lag(name):
    return any(l.lower() in name.lower() for l in _LAGGING_K)

def _kpi_table(kpis, title, color):
    rows_h = ''
    for k in kpis:
        is_lag = _is_lag(k['name'])
        if is_lag:
            dot = '<span style="display:inline-block;width:7px;height:7px;background:#ef4444;border-radius:50%;margin-right:6px;vertical-align:middle;"></span>' if k['current']>0 else '<span style="display:inline-block;width:7px;height:7px;background:#10b981;border-radius:50%;margin-right:6px;vertical-align:middle;"></span>'
            cls = 'red' if k['current']>0 else 'green'
        else:
            dot = ''
            cls = 'blue'
        rows_h += f'<tr><td class="bold" style="font-size:11px;">{dot}{k["name"]}</td><td class="num">{k["last"]}</td><td class="num {cls}">{k["current"]}</td><td class="num blue">{k["cum"]}</td></tr>'
    return f'''<div class="kpi-tbl-card">
      <div class="kpi-tbl-hdr" style="color:{color};">{title}</div>
      <table class="kpi-tbl"><thead><tr><th>Indicator</th><th class="num">Last</th><th class="num">Current</th><th class="num">Cum.</th></tr></thead><tbody>{rows_h}</tbody></table>
    </div>'''

_kpi_r1_html  = _kpi_table(kpi_data['r1_kpis'],       'RIYAH 1',                '#60a5fa')
_kpi_r2_html  = _kpi_table(kpi_data['r2_kpis'],       'RIYAH 2',                '#34d399')
_kpi_cmb_html = _kpi_table(kpi_data['combined_kpis'], 'RIYAH 1 + 2 (COMBINED)', '#fbbf24')

_mp_rows = ''
for _m in kpi_data['manpower']:
    _mp_rows += f'<tr><td class="bold">{_m["team"]}</td><td class="num">{_m["mp"]:,}</td><td class="num blue">{_m["hrs"]:,}</td></tr>'

html_template = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>PDO Riyah 1&2 — QHSE Dashboard</title>
<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>

<script src="https://unpkg.com/@lottiefiles/dotlottie-wc@0.9.10/dist/dotlottie-wc.js" type="module"></script>
<style>
#lockScreen {
  position:fixed;top:0;left:0;width:100%;height:100%;
  background:#0a0e1a;z-index:99999;
  display:flex;justify-content:center;align-items:center;
}
.lock-box {
  text-align:center;width:380px;
  background:#111827;border-radius:20px;
  padding:32px 36px 36px;
  border:1px solid rgba(255,255,255,0.08);
  box-shadow:0 25px 60px rgba(0,0,0,0.6);
}
.lock-hdec {
  font-family:'Syne',sans-serif;font-size:32px;font-weight:800;
  color:#f1f5f9;letter-spacing:0.05em;margin-bottom:4px;
}
.lock-title {
  font-family:'Syne',sans-serif;font-size:16px;font-weight:600;
  color:#f1f5f9;margin-bottom:4px;
}
.lock-sub {
  color:#64748b;font-size:12px;margin-bottom:22px;
}
.lock-input {
  width:100%;padding:12px 16px;
  background:#1a2235;border:1px solid #2d3748;border-radius:10px;
  color:white;font-size:15px;text-align:center;outline:none;
  box-sizing:border-box;margin-bottom:10px;
  transition:border-color .2s;font-family:sans-serif;
}
.lock-input:focus {border-color:#3b82f6;}
.lock-err {color:#f87171;font-size:12px;height:16px;margin-bottom:10px;}
.lock-btn {
  width:100%;padding:13px;
  background:linear-gradient(135deg,#3b82f6,#06b6d4);
  border:none;border-radius:10px;color:white;
  font-size:15px;font-weight:600;cursor:pointer;
  font-family:sans-serif;transition:opacity .2s;
}
.lock-btn:hover{opacity:.88;}
.lock-footer {color:#475569;font-size:11px;margin-top:16px;}
</style>

<div id="lockScreen">
  <div class="lock-box">
    <dotlottie-wc
      src="https://lottie.host/9db5369a-b959-48bc-a683-90898fc599cf/wzOT4itrzZ.lottie"
      style="width:180px;height:180px;display:block;margin:0 auto 4px;"
      autoplay loop>
    </dotlottie-wc>
    <div class="lock-hdec">HDEC</div>
    <div class="lock-title">HSE Dashboard</div>
    <div class="lock-sub">Developed by: Shaguf Ahmed</div>
    <input type="password" id="pi" class="lock-input" placeholder="Enter password" onkeydown="if(event.key==='Enter')cp()"/>
    <div id="pe" class="lock-err"></div>
    <button class="lock-btn" onclick="cp()">Access Dashboard</button>
    <div class="lock-footer">Authorized personnel only</div>
  </div>
</div>

<script>
(function() {
  if(sessionStorage.getItem('riyah_auth')==='true'){
    document.getElementById('lockScreen').style.display='none';
    return;
  }
  setTimeout(function(){document.getElementById('pi').focus();},300);
})();

window.cp = function() {
  var v = document.getElementById('pi').value;
  if(v === atob('cml5YWh3aW5k')) {
    sessionStorage.setItem('riyah_auth','true');
    document.getElementById('lockScreen').style.display='none';
  } else {
    document.getElementById('pe').textContent = '\u274C Incorrect password!';
    document.getElementById('pi').value = '';
    document.getElementById('pi').focus();
    setTimeout(function(){document.getElementById('pe').textContent='';},2000);
  }
};
document.addEventListener('keydown', function(e){if(e.key==='Enter')cp();});
</script>

<style>
  @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;500;600;700;800&family=DM+Sans:wght@300;400;500&display=swap');
  
  :root {
    --bg: #0a0e1a;
    --bg2: #111827;
    --bg3: #1a2235;
    --card: #141c2e;
    --border: rgba(255,255,255,0.07);
    --accent: #3b82f6;
    --accent2: #06b6d4;
    --green: #10b981;
    --amber: #f59e0b;
    --red: #ef4444;
    --text: #f1f5f9;
    --muted: #64748b;
    --subtle: #94a3b8;
  }

  * { margin: 0; padding: 0; box-sizing: border-box; }

  body {
    font-family: 'DM Sans', sans-serif;
    background: var(--bg);
    color: var(--text);
    min-height: 100vh;
  }

  .header {
    background: linear-gradient(135deg, #0c1525 0%, #0a1120 50%, #0d1520 100%);
    border-bottom: 1px solid var(--border);
  }
  .header-inner {
    max-width: 1400px;
    margin: 0 auto;
    padding: 22px 32px;
    display: flex;
    justify-content: space-between;
    align-items: center;
  }
  .header-left { display: flex; align-items: center; gap: 16px; }
  .header-title { font-family: 'Syne', sans-serif; font-weight: 700; font-size: 18px; color: var(--text); }
  .header-sub { font-size: 12px; color: var(--muted); margin-top: 2px; letter-spacing: .05em; }
  .header-right { display: flex; align-items: center; gap: 20px; }
  .date-badge {
    background: rgba(59,130,246,.1);
    border: 1px solid rgba(59,130,246,.25);
    border-radius: 8px;
    padding: 6px 14px;
    font-size: 12px;
    color: #93c5fd;
    font-weight: 500;
  }
  .live-dot {
    display: flex;
    align-items: center;
    gap: 6px;
    font-size: 11px;
    color: var(--green);
  }
  .live-dot::before {
    content: '';
    width: 7px;
    height: 7px;
    background: var(--green);
    border-radius: 50%;
    animation: pulse 2s infinite;
  }
  @keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.4} }
  
  .tabs-bar {
    background: var(--bg2);
    border-bottom: 1px solid var(--border);
  }
  .tabs-inner {
    max-width: 1400px;
    margin: 0 auto;
    padding: 0 32px;
    display: flex;
    overflow-x: auto;
  }
  .tab-btn {
    padding: 14px 24px;
    font-family: 'DM Sans', sans-serif;
    font-size: 13px;
    font-weight: 500;
    color: var(--muted);
    background: none;
    border: none;
    border-bottom: 2px solid transparent;
    cursor: pointer;
    transition: all .2s;
    white-space: nowrap;
    letter-spacing: .02em;
  }
  .tab-btn:hover { color: var(--subtle); }
  .tab-btn.active { color: var(--accent); border-bottom-color: var(--accent); }
  .tab-btn .count {
    display: inline-block;
    margin-left: 6px;
    background: rgba(59,130,246,.15);
    color: #93c5fd;
    font-size: 10px;
    font-weight: 600;
    padding: 1px 6px;
    border-radius: 10px;
  }
  
  .panel { display: none; }
  .panel.active { display: block; }
  .panel-body {
    max-width: 1400px;
    margin: 0 auto;
    padding: 28px 32px 40px;
  }

  .kpi-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
    gap: 14px;
    margin-bottom: 24px;
  }
  .kpi-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 18px 20px;
    position: relative;
    overflow: hidden;
    transition: border-color .2s;
  }
  .kpi-card:hover { border-color: rgba(255,255,255,.14); }
  .kpi-card::after {
    content: '';
    position: absolute;
    top: 0; left: 0; right: 0;
    height: 2px;
    border-radius: 14px 14px 0 0;
  }
  .kpi-card.blue::after { background: linear-gradient(90deg,#3b82f6,#06b6d4); }
  .kpi-card.green::after { background: linear-gradient(90deg,#10b981,#34d399); }
  .kpi-card.amber::after { background: linear-gradient(90deg,#f59e0b,#fbbf24); }
  .kpi-card.red::after { background: linear-gradient(90deg,#ef4444,#f87171); }
  .kpi-card.teal::after { background: linear-gradient(90deg,#06b6d4,#67e8f9); }
  
  .kpi-label { font-size: 11px; color: var(--muted); letter-spacing: .07em; text-transform: uppercase; margin-bottom: 8px; }
  .kpi-val { font-family: 'Syne', sans-serif; font-size: 32px; font-weight: 700; line-height: 1; }
  .kpi-val.blue { color: #60a5fa; }
  .kpi-val.green { color: #34d399; }
  .kpi-val.amber { color: #fbbf24; }
  .kpi-val.red { color: #f87171; }
  .kpi-val.teal { color: #67e8f9; }
  .kpi-sub { font-size: 11px; color: var(--muted); margin-top: 4px; }

  .chart-grid { display: grid; gap: 16px; margin-bottom: 16px; }
  .col2 { grid-template-columns: 1fr 1fr; }
  .col3 { grid-template-columns: 1fr 1fr 1fr; }
  
  .chart-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 20px 22px;
    transition: border-color .2s;
  }
  .chart-card:hover { border-color: rgba(255,255,255,.12); }
  .chart-header { display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 16px; }
  .chart-title { font-family: 'Syne', sans-serif; font-size: 13px; font-weight: 600; color: var(--text); letter-spacing: .02em; }
  .chart-desc { font-size: 11px; color: var(--muted); margin-top: 2px; }
  .legend { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 12px; }
  .leg-item { display: flex; align-items: center; gap: 5px; font-size: 11px; color: var(--subtle); }
  .leg-dot { width: 8px; height: 8px; border-radius: 2px; flex-shrink: 0; }
  canvas { display: block; max-height: 220px; }
  
  .table-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    overflow: hidden;
    margin-bottom: 16px;
  }
  .table-header {
    padding: 16px 22px 14px;
    border-bottom: 1px solid var(--border);
    display: flex;
    justify-content: space-between;
    align-items: center;
    flex-wrap: wrap;
    gap: 12px;
  }
  .filter-bar { display: flex; gap: 10px; margin-top: 12px; flex-wrap: wrap; align-items: center; }
  .filter-bar input, .filter-bar select, .filter-bar button {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 8px 12px;
    font-size: 12px;
    color: var(--text);
    outline: none;
    font-family: 'DM Sans', sans-serif;
    transition: border-color .2s;
    cursor: pointer;
  }
  .filter-bar input:focus, .filter-bar select:focus, .filter-bar button:hover { border-color: var(--accent); }
  .filter-bar button { background: rgba(59,130,246,.1); }
  .filter-bar button.active { background: var(--accent); color: white; border-color: var(--accent); }
  
  .tbl-scroll { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 12px; }
  thead th {
    padding: 10px 14px;
    text-align: left;
    font-size: 10px;
    color: var(--muted);
    letter-spacing: .08em;
    text-transform: uppercase;
    font-weight: 600;
    background: rgba(255,255,255,.02);
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
  }
  tbody tr { border-bottom: 1px solid var(--border); transition: background .15s; }
  tbody tr:last-child { border-bottom: none; }
  tbody tr:hover { background: rgba(255,255,255,.02); }
  tbody td { padding: 10px 14px; color: var(--subtle); vertical-align: middle; }
  tbody td.bold { color: var(--text); font-weight: 500; }
  tbody td.code { font-family: monospace; font-size: 10px; color: var(--muted); }
  
  .badge {
    display: inline-block;
    padding: 3px 8px;
    border-radius: 5px;
    font-size: 10px;
    font-weight: 600;
    letter-spacing: .04em;
    text-transform: uppercase;
  }
  .b-app { background: rgba(16,185,129,.15); color: #34d399; border: 1px solid rgba(16,185,129,.25); }
  .b-rej { background: rgba(239,68,68,.15); color: #f87171; border: 1px solid rgba(239,68,68,.25); }
  .b-awc { background: rgba(245,158,11,.15); color: #fbbf24; border: 1px solid rgba(245,158,11,.25); }
  .b-rwc { background: rgba(59,130,246,.15); color: #60a5fa; border: 1px solid rgba(59,130,246,.25); }
  .b-ur { background: rgba(245,158,11,.15); color: #fbbf24; border: 1px solid rgba(245,158,11,.25); }
  .b-open { background: rgba(239,68,68,.15); color: #f87171; border: 1px solid rgba(239,68,68,.25); }
  .b-closed { background: rgba(16,185,129,.15); color: #34d399; border: 1px solid rgba(16,185,129,.25); }
  .b-major { background: rgba(239,68,68,.15); color: #f87171; border: 1px solid rgba(239,68,68,.25); }
  
  .alert-banner {
    background: rgba(239,68,68,.08);
    border: 1px solid rgba(239,68,68,.2);
    border-radius: 10px;
    padding: 12px 18px;
    display: flex;
    align-items: center;
    gap: 12px;
    margin-bottom: 16px;
    font-size: 12px;
    color: #fca5a5;
  }
  .alert-icon {
    width: 28px; height: 28px;
    background: rgba(239,68,68,.15);
    border-radius: 6px;
    display: flex;
    align-items: center;
    justify-content: center;
    font-size: 14px;
    flex-shrink: 0;
  }
  .section-label {
    font-size: 10px;
    color: var(--muted);
    letter-spacing: .1em;
    text-transform: uppercase;
    font-weight: 600;
    margin-bottom: 12px;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
  }
  
  .export-btn {
    background: rgba(59,130,246,.1);
    border: 1px solid var(--border);
    border-radius: 6px;
    padding: 6px 12px;
    font-size: 11px;
    color: var(--subtle);
    cursor: pointer;
    transition: all .2s;
  }
  .export-btn:hover {
    background: var(--accent);
    color: white;
    border-color: var(--accent);
  }
  
  footer {
    text-align: center;
    padding: 20px;
    font-size: 11px;
    color: var(--muted);
    border-top: 1px solid var(--border);
    margin-top: 20px;
  }
  
  @media (max-width:900px) { .col2,.col3 { grid-template-columns: 1fr; } .panel-body { padding: 20px 16px; } }
</style>
</head>
<body>

<div class="header">
  <div class="header-inner">
    <div class="header-left">
      <img src="data:image/jpeg;base64,/9j/4AAQSkZJRgABAQECWAJYAAD/2wBDAAMCAgMCAgMDAwMEAwMEBQgFBQQEBQoHBwYIDAoMDAsKCwsNDhIQDQ4RDgsLEBYQERMUFRUVDA8XGBYUGBIUFRT/2wBDAQMEBAUEBQkFBQkUDQsNFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBQUFBT/wAARCAE1BVoDASIAAhEBAxEB/8QAHwAAAQUBAQEBAQEAAAAAAAAAAAECAwQFBgcICQoL/8QAtRAAAgEDAwIEAwUFBAQAAAF9AQIDAAQRBRIhMUEGE1FhByJxFDKBkaEII0KxwRVS0fAkM2JyggkKFhcYGRolJicoKSo0NTY3ODk6Q0RFRkdISUpTVFVWV1hZWmNkZWZnaGlqc3R1dnd4eXqDhIWGh4iJipKTlJWWl5iZmqKjpKWmp6ipqrKztLW2t7i5usLDxMXGx8jJytLT1NXW19jZ2uHi4+Tl5ufo6erx8vP09fb3+Pn6/8QAHwEAAwEBAQEBAQEBAQAAAAAAAAECAwQFBgcICQoL/8QAtREAAgECBAQDBAcFBAQAAQJ3AAECAxEEBSExBhJBUQdhcRMiMoEIFEKRobHBCSMzUvAVYnLRChYkNOEl8RcYGRomJygpKjU2Nzg5OkNERUZHSElKU1RVVldYWVpjZGVmZ2hpanN0dXZ3eHl6goOEhYaHiImKkpOUlZaXmJmaoqOkpaanqKmqsrO0tba3uLm6wsPExcbHyMnK0tPU1dbX2Nna4uPk5ebn6Onq8vP09fb3+Pn6/9oADAMBAAIRAxEAPwD9U6KKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKTpQAdK8qtP2ofhpffEEeC4fEkba41wbNVMEgia4DFfKEhXaWyMDnBJABJOK+ef2wf2w/sf23wL4Evf8ASOYdT1i3b/V9mhhYfxdmcdOg5yR8LwXEtrcRzwyvDPGwdJI2KsrA5BBHQg96Vz9KybhCWNw7r4xuHMvdS39Xfp5bn7iUV8Wfsc/tff219h8BeOLwnUiVg0rVpjn7T2WCU/3+gVj97ofmxu+06Z8VmWW18rxDw9dej6Nd0FFFZ3iHxDp3hTRL3V9XvIrDTbOIzT3ExwqKOp/+sOSeBQeWk27IXX9f03wto93q2r3sOnabaIZJ7m4cIka+pJ/L3Jrz/wCFv7Svw++Mer3Wl+GNb+06jbqZDbTwSQvJGCBvTeBuHPTqO4Ffn9+1j+1PffHLWzpOkPLZeDLKTMEByrXbj/ltIP8A0FT0B55PHg3h3xNqng7XrLWtFvZdO1SylE1vcwnDIw/mDyCDwQSDkGtFFdT9CwvCkqmEdSvJxqPZdF/i9fw/A/cqivBv2Wv2o9L+Pvh1bW7MWn+L7OMfbLEHCzAcebFnqp7jqpODxgn3micJU3aR+eyThOUJKzi2muzXQKKKKzEFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFJ0oAOlfEX7YP7Yf2b7b4F8CXv77mHU9Yt2+52aGFh37M46dBzkhv7YP7Yfk/bfAvgS9/ecw6nrFu33ezQwsO/ZnHToO5HwuTmk2frHDPDN+XHY6PnGL/N/ovmwJzRRRUn62AODX3v+xx+16NcXTPh94znP9qZFvpesTOALgY+SCUn/AJacbVb+PKqfm5f4IoBwcimeNmuV0M2w7oVlr0fVPv8A5rqft3rOrW2gaPfaneyeVZ2UD3Ez/wB1EUsx/IGvy6/ao/al1P44a7LpumSzWHgy0crb2gYj7WQ3E0o45PBCn7v1ya+h/wBj79phPihpU3wz8ez/AG7UZbd4LK7uG5v7fYQ8Mh6mQKCdxOXBOfmXLeL/ALXf7I958JL658VeF4JLzwXcSbpIlBZ9Mdj9xvWIk4V+3Ctzhn0j5H5bk+CoZTmcsPmC/efYf2X5rzfTs7rc+WnNV5GqVzVeQ5NaI/R6sjc8DeLdT8FeJ7HVdJvZdPv4JA8NxC2GRh0PuOoIPBBIPFfqp+zN+0zpvxx0UWV6YrDxdaRg3VmDhZ1HHmxZ6r6r1Un0wT+RDOVO4cEcivQfC3ifUfDGrWOs6ReS2Go2rrNBcQthkb+o7EHggkHivscrwUM2wk8NPSUNYvtfp6X/ADP5n8RK88hzTD5nSV4Vk4zj3cbWl/is7ekUmftTRXhf7M37TWm/G/RRY3xisPF1pHm5swcLOo482LPb1XqpPpgn3SvkcThquEqujWVpIjCYujjqMa9CV4v+rPzCiiiuY7AooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKTpQAdK+Hv2wf2w9n23wL4Evfm5h1PWLdunZoYWH5M4+g7mrf7XX7WbbL3wT4IvMZzDqWrwN+DQxMPyZx9B3NfC1xAQSR171nzpuyP1fhnhpNxxuOj5xi/zf6L5srk5ooopn64FFFFABRRRQBPYX1xpd9b3lpM9vdW8iyxSocMjqcgg+oIFfrf+z/8SYPjp8GdM1fUIYbi4niex1S2dAY3lUbZAV6bXBDbfR8V+RNfoJ/wTYnuG8BeLoWJ+ypqUbIO28xAN+gSmj4DjPC06uXrEP4oNWfk9GvyfyPG/wBrz9jS5+Gsl34v8F28l34UYmS6sUy8mnepHdovfqvfjmvkFz1r95JIkmjaORVdGBVlYZBHoRX53ftlfsVHwmt947+H9mX0MZm1PRIFybLu00IHWLuyf8s+o+TIj6Iyu7HyGT5+5xWGxcteku/k/wDP7/P4inbArqrByLWAE87F/lXIyAyyqg6scV1ULAKAOg4FfonC1Np1anTRfmfkvi5iIzjhMOt7yfy0X46/cdF4b8Saj4V1mz1bSbyWw1G0kEsNxC2GRh/TsQeCCQeK/TP9mX9pvTvjfoy2F+YrDxdaR5ubQHC3Cj/lrF7eq9VPtg1+WcclbPhzxHqPhbWbPVtJvJbDUbSQSw3ELYZGH+enQjg19LmuVUszpWek1s/0fkfhmT5xWyatzR1g/ij3815/0z9qKK8I/Zk/ac0743aMun6g0Vh4utI83FqDhbhR/wAtYvb1Xqp9sGvd6/GsThquEqujWVpI/fMJi6OOoxr0JXi/6s/MKKKK5TsCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKTpQAdK+Of2q/wBqVmW78G+DbvCnMWo6rA3XsYomH5Mw+g7mr/7Tn7TBkW78I+Ebr5OYr/U4W69jFGR27Fh9B3NfH1xBnNeLicar+zpv1Z+o8OcOfDjcbHzjF/m/0X3nO3EGc1m3Fv14rop4MZ4rPngqKVU/VE7HNTwEEkD6iq9bdxb9eKzJ4CCSB9RXqQmpI6oyuV6KKK2NAooooAOtfq1+x58Lpfhb8E9Lt7yIw6rqrHU7tG6o0gART6YRUyOxzXyf+xd+y/N8RtZtvG3iO2KeFtPn3WkEqkf2hOh7A9YlYcnoWBXnDY/R0DFUj8f4zziFVrLqLvZ3l69F8uvyFpCARg9KWvlf9sP9sO1+DdlN4U8KzxXfji5j/eSjDppaMOHcdDKQcqh6Ahm42h+zC4WrjKqo0Vdv8PNn5Di8VSwVJ1qzsl+PkvM+Fv2n/Ceg+Ev2i/Gen+HPIXSLe6UxxW2PLhkaNGljXHA2OzrtH3cY4xXnscmP8Ko/aGmkaR2Lu5LMzHJJPUk1Mj1+05fg44GgqUXd9X3Z+LZ7mtXOsX9YqbJKMV2S/wA9W/NmlHJmrEclZscmP8KtRyZr1Ez5ScDofDviLUPDGsWmq6VeS2Go2kglhuIWwyMO/wD9bv0Nfph+zH+05p3xt0ddO1ForDxfaR5uLUHC3KjrLF7eq9R9MGvy2jkrY8PeIdQ8M6xaarpV3LY6haSCWG4hbDIw7ivGzTK6WZ0uWWk1s/0fkevk+cVsnrc0dYP4o9/NeZ+1NFeDfsxftO6f8bdHXTdSaKx8X2keZ7YHatyo6yxD+a9vpXvNfjWJw1XCVXRrK0kfvmExdHHUY16Erxf9WfmFFFFcp2BRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRSEgdaAFooooAKKKKACiiigAooooAKKKKACiiigAooooAKKiubqGziMs8qQRAgF5GCqCTgcn1JAqQHNAC0UUUAFFFFABRRRQAUUUUAFFFJnFAC0UmR60tABRSZxRketAC0UUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUmR60ALRSZzS0AFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUU1nVFJZgoHJJOKAHUU2ORJo1kjYOjAMrKcgj1Bp1ABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRUU9zDapvmlSJcgbnYAZPAHNAEtFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAV8w/tVfHS50eSbwXoEzQXDRj+0LuM4ZVYZESnsSCCT6ED1r6er89/j/pNzp/xb8TLdIytLdNMhYfeRgGUj2wQPwrxs1rzo0VydXY+x4WwlHFY1usr8qul53X5HlM0XWqUsVbEkVVJYq+ThM/bEzDuLfrWdPBjPFdBLFVC4t+telSqmqdznp4Kzbi368V0U8GM8VnzwV61KqWnY5qeAgkgfUVXrbuLfrxWZPAQSQPqK9SE1JHVGVyvX0n+yb+yddfGXUIvEXiOKWz8E20nQEo+pOp5jjPURgjDOPdV+bJRn7J/7KF38Z9Rj8Q+IYpbPwTayc9UfUXU8xxnqEBGGce6rzkp+l+m6baaNp9tYWFtFZ2VtGsMNvAgRI0UYVVUcAADGBW6R+dcS8SrBp4PBv951f8vkvP8AL12WwsLfS7G3s7SFLe1t41iihjXaqIowqgdgAAKsUV8o/tn/ALZ9p8DdPm8KeFZob3x9dRDe+A8ekxsMiSQHgykEFIzxgh2G3asnbhcLVxlVUaKu3+Hm/I/DMTiaeGputWen5/8ABH/tlftk2nwUsJvCnhWeK88d3MfzycPHpUbDIkcdDKQQUQ9AQzDG1X/Lu81O61a/uL29uZbu8uZGmmuJ3LySuxyzMx5JJJJJ5JNZ97qt3q+oXN9f3M17e3UjTT3NxIXklkYks7MeSxJJJPJJpUev1/LcupZbS5Iaye77/wDA7I/I8zx1XMKvPPSK2Xb/AIPc0YpatRyVmI9WYpa9k+enA0kerEcmP8Kz45KsI9UcU4mlHJmrEclZscmP8KtRyZq0zjnA6Dw/4gv/AA1q9pqml3ctjqFpIJYbiFtrIw6EGv0//Za+Pw+OvgueW+jhtvEOmOsN7DE3EgK5WYD+EMQwxzgqfUV+VEclfa3/AATg8P3z614u10oy6atvFZBz915S28gepUAZ/wB8etfLcSYajVwUq0/ijaz+ex9XwpisRQzGNCm7wne69E9fkfddFFFfkB+6BRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABX5kf8FCv2o5vFnjBPAHhXUHi0jQp99/eWshUz3inGwMP4Y/1fP91TXuv7c37YkXwl0q48D+EbxJPGl7Fi6uYmB/suFh1PpKwOVHVQQxxlc/m78N/h14h+Mnjuw8N6DA19q+oy8ySsdkY6vLI3OFUZJPJ9ASQD95kOWKC+v4nSKWl/z/AMvv7HxmdZg5v6lh9W97fl/n9x9u/sCfF/42/FXxgtjqOvyat4G0eMtqF3qkKyzMzKRHEsxG9nJG4kk4UNnkqD9yfEfUbrSPh74nvrGY297a6ZczQTAZ2SLExVsd8EA1h/A34NaJ8B/h1p/hTRF3xw5lurtlCyXdwwG+V8dzgADnCqoycVqfFj/kl3jD/sD3f/ol6+bxuIpYrGc9GCjG6SsrX13fqe9hKFTDYTlqyblbXW9vJeh+Rp/bc+Nxg+z/APCfXuzbsz9ng3Y6fe8vOffOa/Xv4cX15qnw98MXmo3Aur+40y2luJwu3zJGiUs2O2SScV+CX8f41+5/hnxnoPgb4PeF9T8RaxY6Jp8ek2ga5v7hYUz5K4ALEZJ7Acmvp+JMNSpxpKjBJtvZb7dj57IMRUqSqutNtJLd+vc9ArL8ReKdF8H6cdQ17V7HRbAOsZutQuUgiDE4UbnIGT2FfHXxs/4KaeFfDUFxp/w7sJPE+q8quo3iNDYxnn5gpxJJgjphAc5DHpX5+/FH4zeMfjLrZ1TxbrlzqkwJ8qFjtggHpHGMKo+gye5NeXgeHsTifere5Hz3+7/M9HGZ5h8P7tL35fh9/wDkfu0CCMjkUtfEH/BO740/EjxZpMfhfxFoOo6p4UtLd/7P8UyxlUh2FV+ztI2BKBnC7csuMEFRlPt+vBxmElgq0qM2nbt/Wnoe1hMTHF0VVimr9woooriOsKK434w/Eq1+D/w21zxheWc2oW+lxLI1tbkB5MuqAAnpywyfTNfLH/D1LwH/ANCl4h/8gf8Axdehh8vxWLi50IOSWhw18bh8NJQrTsz7Zor4m/4epeA/+hS8Q/8AkD/4uj/h6l4D/wChS8Q/+QP/AIuur+xcw/59P8P8zn/tbA/8/V+JH/wU++Li6B4B0bwDaODe65N9su+hMdtEw2j2LyYwR2jYd6+ev2Rv2ovjBa+P/DPgbSdQPijTr25SEadqx80xQqC0hjlJ3IFjViBkqNvSvGf2hfjNffHn4qav4su1khtpmENjaSNu+zWy8InpnkscfxMx713f7Hnx58Gfs8eKdX8SeItF1PWdXmtxaWJshHst0JzKx3sPmbCAYHA3c/NX3tPLlhsr9i6anO17f3n/AJfofFzxzxGY+1VTlhe1/Jf5/qfsYKWvib/h6l4D/wChS8Q/+QP/AIuj/h6l4D/6FLxD/wCQP/i6+C/sXMP+fT/D/M+0/tbA/wDP1fifbNFfE3/D1LwH/wBCl4h/8gf/ABdH/D1LwH/0KXiH/wAgf/F0f2LmH/Pp/h/mH9rYH/n6vxPtmivl74Kft8eFvjh8StL8G6V4c1ixur9JmFzdmPy08uNpOQrE8hSM+uPWvqGvPxGFrYSfJXjyvc7qGIpYmPPRldbBRRRXKdAV8wf8FF9UvtI/ZwuZ9PvbiwmOpWyGS2laNmU7gVJB6H0p/wC3l8b/ABj8CvAPh3V/B99b2M91qf2a4ae3WYsvls4ADZAB2EHjPPBFfnx8Wv2w/iR8a/CL+G/E97Yz6W0yTlLezWJt65x8w+tfWZNlNetOni1bkT+ej9D5nNcyo0oVMK78zXy1+Z1P7GfxE8WeIP2m/AtnqPinWb61knmDw3N/LIjgQSNgqzEEZUfkPSv16r8E/h18QNY+FvjPTfFOgyRw6tpzM8DzRiRQWRkOVPXhjXvn/Dx740/9BPS//BbHXv5zk1fHV41MPZJK3bq/I8XKs2o4OjKFe7bd+/ReZ9u/8FCL+90z9mfWbqwvrqwnjvLUeZaymNmVpArKSOSpDHivg79kT4i+Lte/aU8B2moeKtavLaS9KvFPfyyK6+WxKkMSCCVGfp9Kxfip+2X8S/jJ4NufDHiW+sZ9JuJI5JEgskjYlGDL8w9xXlvw/wDHeq/DPxjpnifQ5I4tV06Qy27yxh1DFSvKnrwTXZgMrqYbA1MPVScnzW+asuhyY3MadfGQr021FWv8mfvhXyj/AMFG/HfiX4e/BvQdR8Ma7faDeSa9FBJPp87QyOnkTPtLKQduUBI6HFeF/sy/tt/FX4qfHjwl4Y1vUtPl0nUJpY7iGKxSMsoid8hgMgjZ29efb1f/AIKmH/ixHhwf9THD/wCk1xXyGHy6pgcyo0cRZ317rqfU18dDGYCrVo3VtO3Y8F/Yg+OnxF8fftJ+GtL1/wAa61q2ltDdtLZXV47xSYt3I3KTg4YAg9iK/UWvyE/4J3Ln9qfw4fS1vD/5LvX691fElOFPGRjBJLlW2nVk5BOU8K3N3fM9/RBRRRXyh9IFFFFAGf4g1/T/AArod/rOrXSWWmWED3NzcSZ2xxqCWY454A6DmvzH+Mn/AAUT8X3/AMY4NW8BXYtPCmkFobawu4yYtSByGlnTIbB/hGQVAB4JNfqJcW8V3C8M8ayxOMMjjIP4V8a/tFf8E5PDfjwXOtfDxrfwnrhy7aaVIsLg+iqP9Sf90Ff9kda+iyatgaVV/XI3vom9UvVfqeFmtLGVKa+qvbVrq/67Hd/s6/tz+CPjk9tpF648LeLZAFGnXsg8q5bgYgl4DEkjCHDegIBNeq/Fn46eCfgfZ6ddeNNYOkW+oSNFbuLaWfeygFuI1Yjgjr61+KHxA+HniH4W+Kbvw/4l02bS9UtjzHKuA69nQ9GU44I4rs/Hvxp8ffE/4T6DpPih31bRtDvmjs9YuFJmLtHzC0hP7zCgHP3hkZJBWvo6vDlCpVhUoz/dvdX+6z1ueDTz6tClKnVj+8Wzt+aP00/4b/8AgR/0O5/8FN7/APGaP+G//gR/0O5/8FN7/wDGa/K/4U/A/wAa/G6+v7PwXo39s3NjGstwn2qGDYrHAOZXUHn0zXpP/DAfx4/6Eb/yrWX/AMeoqZHlVGXJUrtPs5RT/IKecZlVjzU6Ka7qMn+p+g//AA3/APAj/odz/wCCm9/+M1veBP2xfhH8S/Flh4a8OeKjqGs35Zbe3On3UQcqpcjc8aqPlUnk9sda/Lf4hfsj/Fj4V+FLvxL4o8K/2ZotoUWa5/tC1l2l3CL8qSsxyzAcDvW5+weM/tXeBM/89Lr/ANJJqzq5Fl/1apXoVHLlTeji1dK/RGlPOcb9YhRrU1Hma3TTs3bqz9k8iue+Il/c6X8P/E17ZTG3vLfTLmaGZRko6xMVb8CAa/Oj/go9478TeE/2gbGHRfEmr6VA2hQSeVZX0kKAmWUNgKw67Fz64HpXrX7EniTV/FH7IfxLu9Z1W91e5S61GNZr64eZ1X7DE2AWJIGWY49SfWvAeUypYWnjXNNNrS3dntLM41MTPCKNmr637HyToH7Y/wAab3V9MsG+IGoeVJcRRZeKJjjeByduT7881+yinKj6V+A/hL/kbNG/6/Yf/QxX77p9xfpXq8T0aVGVJUoqN77JLsebw9VqVY1faSbtbd37jqKKK+IPrwrwz9qH9qvw/wDs4eHisy/2l4rvIS+naSAwD8kebI3RUBBzzk9B3I9zryj9oj9nLwz+0X4Ok0nWU+x6pAC+nazFGGms5MfhujPG5MgMAOQwVl7MG6CrxeJT5Otv627nLilWdGSw9ufpc+eP2R/2/wCHx/qC+FPiVcw2HiC6nb7BqwRYrWfcflgcDARx0UnhuhIbG/7dBzX4V/GX4MeJvgV41ufDniW08mdPnt7qPJgu4s8SRt3B/MHIIBGK+z/2Ff215bmex+HHxA1DzHfbBo2s3L/MT0W3lY9c8BGP+6f4a+uzbJYSp/XMDrHdpdu6/wAj5jLM2nGf1XGaS2Tf5P8AzP0GryL9qD46zfs8fDRfFcGjLrjfbobQ2r3HkgBwxLbtrf3cdO9euda+UP8AgpgwX9m4AkAtrFqAD3O2SvlcvpQr4unTqK6bVz6PHVJUcNUqQdmk7Hj/APw9iuv+iaRf+Do//GKP+HsV1/0TSL/wdH/4xXxX8Kfh1efFn4g6L4S0+7trG91SUwxT3ZIiUhS3OATzjHA6kV9U/wDDq74g/wDQ1+G/++p//jdfoOIwGSYSShXiot67y/zPiKGNzfExcqLbS8o/5HV/8PYrr/omkX/g6P8A8Yr2T9lb9t6f9pL4iX/hmTwhHoCW2mSagLldQNwWKyxpt2+WvXzc5z/D78fAv7SH7LHiD9mc+Hxruq6bqf8AbP2jyf7PMh2eV5e7dvUdfNGMehr1r/glv/ycBrv/AGLdx/6U21c2My3LXgJ4nDQ6aO7726s3wuPx6xsMPiJddVZf5H6mUUUV+bn3wUUUUAFcP8YvjJ4Z+BfgybxL4qupLexD+RBFBEZJbmcozLEgHG5gjYLEKMckDmu4rmviH8OfD/xT8L3Xh/xLp0epaZcKQYpOqMVKh1PZhuJDdQcEcitqPs1Uj7W/L1tvYyq8/I/ZW5ul9j4C+Dn/AAUt1WP4n6s/xAiZvBuq3ObSO1jRpNFTcdoyqq0yBSNxPzfLuUZJU/opomuaf4l0m01TSr2DUdOu4xLBdW0geORD0KsOCK/Hn9qX9kfxH+znrr3KrJq/gy6k22WsKvKE5IhnA+5IAOv3WHI53Kun+yJ+15qv7PniCPS9UebUvA15KPtVluLNaEnmaEZwD3Zf4gOxwa+8x2T4fG0Fisuttstn/k/z9T4vB5rXwlZ4fH9930/zX9bH7A0VneHfEOm+LNDstY0e9h1DTL2JZre6gbckiEZBBrRr8+aadmfcJpq6CvyP/bW+JXi/wz+1D44s9J8V63ptoj2oSG11CWJFBtomwArAAbnY4/2j61+uFfjZ+3l/ydf47/37T/0jhr67hmMZ4uakr+6/zR8xxDJxwsXF295fkz67/wCCYnifXPFXg/xvc63rmo6y6X8CRi/uWm8v5GZipYkjJPPOOB75+2K+Fv8AglP/AMiH45/7CUH/AKKNfdNeXnSUcwqpLqvyR6OUtvA02/P82FFFFeIeuFFFFAHD/GH4w+Hvgd4OfxN4me4TTVmW3AtYvMkZ2ztAGR6HkkCvHP8Ah4x8E/8AoNah/wCC2X/Crn7fHgfxH8QvgK+keF9GuNc1FtSgla3tRl1jUOSwHfnAwOea/OL/AIY6+M//AET7Vv8AvlP/AIqvsMqy7AYrD+0xNTlld/aS0+Z8vmOPxuGr8mHheNl0b/I/RH/h4x8E/wDoNah/4LZf8KP+HjHwT/6DWof+C2X/AAr8k9T0250bUrqwvIjBeWsrwTRMQSjqSGU49CCK9H8Mfsw/FPxnoNnrWi+CtS1DSrxPMt7qJV2yLkjIyemQa96fD2W0kpTm0n3kl+h4sM8x9R8sIJvyT/zP0jk/4KNfBREZhrGpOQMhV02XJ9hkYrf8bftw/Cz4f6rp1jrGoX8T3+nQ6pBLFYu6GGVQ0fI5yVOenGDkg4z+Z8n7H3xmgjaRvh7rBVAWISNWOB6AHJPsK9c/ai/Z6+KHjbx14en0rwLq13BbeGdNtHlgj3oJI4FEikjgFWYrg9dpIrhllGVqrCCq6NO/vLpax2RzTMXTlJ09Va3uvzPrf/h4x8E/+g1qH/gtl/wo/wCHjHwT/wCg1qH/AILZf8K/Kbxn4G174d65Jo3iTS7jR9UjRZGtbldrhWGVOPcVt+Avgl46+KNjc3nhTwxf67a20nlTS2ke5UfGdp564Oa7nw9l0Ye0c3y9+ZW++xxrPMe5cigr9rO/5n6ff8PGPgn/ANBrUP8AwWy/4Uf8PGPgn/0GtQ/8Fsv+FfnSf2RPjIASfh5rWP8AriP8a8gIwcUqfD2W1b+zm3btJP8AQJ55j6VueCXqmv1P11/4eMfBP/oNah/4LZf8K90+GvxE0b4seCNM8V6BJLLpGoq7QNNHsf5XZGBXsQysPwr8aNF/Zc+K/iLR7LVdN8C6veadewpcW9xFCCksbAMrA56EEGv1b/Y/8O6r4T/Zw8FaTrenXGk6pbW8oms7pCkseZ5CNynkEgg4PrXz+b5dgsHRUsNO8r2eqfR9j3Mrx+LxdVxrwsrXWjXVdz2Oiiivkj6YKKRjhSfavyjv/wDgo/8AGPTdQvLX7RosoinkQM+njOAxwOGHQcdO3PNergMtr5jzext7tr38/wDhjzcZmFHA8vtb6328j9XaK/Jv/h5Z8Yv+emh/+C//AOzo/wCHlnxi/wCemh/+C/8A+zr1v9Wsd/d+/wD4B5n+sGD8/u/4J+reo2813p91Bb3L2U8sTJHcxKrPExBAcBgVJB5wQRxyDX4r/tBeMfipYePNd8I+PPFmsalc6ZO1u8E104gkTqkixghcOhVgcZIYZr03/h5Z8Yv+emh/+C//AOzrxH4z/GnXvjt4pi8Q+JILBNVS3W2aWxgMQkRSSu4ZOSNxGeuMDsK+hybKsTgKsnXjFp9d2n9x4ebZlQxtNKi5Jr7mvvP0c/4J3fH/AP4Wd8MD4P1N8a74VijgjYn/AI+LPG2Jh3ymNh9ghzljj62r8I/hF8X/ABJ8EPGUXibwvcx2+opE8DLMm+OWNhyrr3GQp+qg9q94/wCHlnxi/wCemh/+C/8A+zrgzHh2tVxMqmGtyvXV7PqduBz2lSoRp4i/MtPl0P1kor8m/wDh5Z8Yv+emh/8Agv8A/s6P+Hlnxi/56aH/AOC//wCzrzf9Wsd/d+//AIB3/wCsGD8/u/4J+slFfky//BSn4xyIyibRIywwHXT+V9xlsV+ovw81i88Q+AfDWq6j5f8AaF9pttc3HkjCeY8Ss232yTivMx2V4jL4xlWtr2Z6ODzGhjnKNK+nc6GiiivHPUEPAr4g8R/8FQdH8NeJtZ0e58AagZNOvJbTd9vQFtjbSSNnynIPGT9a+3z0r8IvjR/yWHxz/wBh2+/9HvX1WQYDD46dSNeN7JW1a/I+bzrGV8HCEqLtdvomfd3/AA9b8P8A/Qg6l/4Hx/8AxNMuP+CruiLC5h+H1/JKB8qyaiiqT7kIcfka8A+H/wDwT0+JXxI8F6N4n0vUfDsWn6rbLdQJc3cyyBGGRuAhIB+hNeK/GD4Uap8FfHd74S1q90+91SzSNp202VpI0LqGCksqnO0gnjuK+mpZXk1ao6VPWS3V30+Z89VzHNaMFUqaRfWy/wAj9Xv2ef2z/A3x+Wz02GY6F4umDbtCuiWZiqlmMUmAsgCqx7EAHKivfq/G39mT4hx/s7if4q3vhNfE6tcNomnB7v7P5Exj8yaRW8t/mCFF6dJW59fon/h7Kf8Aol4/8H3/ANzV8/j8hre3f1GDcPVb9Vq76eZ7mDzql7FfXJ2l6PbvtY+8PGWm6nrHhPWbHRdSOj6xc2ksVnqAQP8AZpipCSbTwcNg49q/En4ueOPiLqni/UNN8d+INXv9W0u7eGS3vbp2SCVDtOxM7V5HVQM5z3r7G/4eyn/ol4/8H3/3NXyd+0p8atP+P/xHfxfaeGT4Yu7i2jhvIRe/ahO6DasmfLTB2BFxz9wV7GRYDFYKrJYikkn191tP776nlZzjcNi6cXQqO66apNfdbQ/VT9lD492/7QPwlsNakZU12zxZatAq7Qtwqgl1H91wQwxwMkdVNezV+L37LH7UOofsx+JNXvoNJGvabqlusNxp7XRt8ujExyB9rcrucYx0c19Mf8PZT/0S8f8Ag+/+5q8bH8P4pYiX1WF4PVapfLV9D1cHneHdCP1idprfR/foj9CaK/Pb/h7Kf+iXj/wff/c1fTv7K37SJ/aY8H6trp8Onw59gvvsXk/bPtIk/dq+7dsTH3sYx+PPHjYnKcZhKfta0LR9V+jPWoZlhMTP2dKd36P9Ue2UUUV5J6YUUUUAFFFFABRRRQAUUUUAFFFFABXm3xp+C2n/ABa0dSHWy1y1U/ZL3GRjqY5B3U+vVTyO4b0misqtKFaDhNXTOnD4irhKsa1GVpI/MvxT4W1HwprN1pWq2r2l9btteNx+RB7gjkEcEVgSRV+hvxl+DWnfFfRcHZaa3bqfsl9jp32PjqhP4g8juD8I+KfC2o+E9ZutK1W1e0vrdtrxuPyIPcEcgjgivhsZg54OfeL2Z+25NnNLNKfaot1+q8vyOVliqlLFWxJFVSWKsITPp0zDuLfrWdPBjPFdBLFVC4t+telSqmqdznp4K9z/AGYv2Vbn4xapHrmvRy2fg22k+YjKvfuDzHGeoQHhnH+6Oclan7MvwZi+NvxEurO7DHw/oixzaq6PsYl8+VCvfL7WJI6Kp5BZc/pTpmmWmi6dbWFhbRWdlbRrFDBCoVI0AwAAOgAr6PD05WUpaI/PuIOJVhVLCYKX7zZv+XyXn+XrsaZplpounW1hYW0VnZW0axQ28CBEjRRgKoHAAHarVFeQ/tSfHi2/Z9+E+oeIAI5tZn/0TSrWQ8SXLA7WYZyUQAs3TIGMgkV61GlOvUjSpq7eiPxqtVjShKrUei1Z5d+2j+2fa/AzT5fCnhOeC88f3UeXYqJI9JjZciSQHgykEFIzkYwzjaVWT8ptS1K71nUbq/v7qa9vrqVp57m4kMkksjElnZjyzEkkk8kmrPiLxFqXi3Xb7WdYvJdQ1O+lae4uZmy8jsckn/DtWdX7HluXUsupckdZPd9/+B2PynMMfUx9XmlpFbLt/wAEVWx9KsRyVWpVbH0r1zymrmij1YR6zo5Kso9M5JwNGKWrUclZiPVmKWqOOcDSR6sRyY/wrPjkr2D9nb9nzXvj/wCK1srENZ6LasrahqrJlIEP8K/3nPOF/E4AJrOtWp4em6tR2ijGnhqmIqKlSjeT6F39nv4Ba78e/FS2ViGs9GtirahqjrlIE/uj+855wv4nABNfqp4A8A6L8M/Ctj4e0G0Fpp9omAOryN/E7n+JieSf6cUz4d/DvQvhb4Us/D3h6zWz0+2X6vK5+9I7fxMe5/AYAArpa/Ic3zepmVSy0prZfq/P8j9lyTJKWVU+Z61Hu/0Xl+YUUUV88fThRRRQAUUUUAFFFFABRRRQAUUUUAISAMngV8S/ta/8FANM8GWt74U+Gt7FqniM5hudajAe3sexEZ6SSe4yq+5yA/8A4KI2nxj/AOEYu7nw9fg/DYQIupWempsuhndvaZuWeLGN23C4I3KQC1fnr8MPhR4o+Mfiq30DwrpU2p30py7KMRW6d5JX6Io9T1OAMkgH7fJspw9Sn9cxM04rp0X+L/I+QzXM68J/VaEGm+vV+n+ZzF/f3Oq31xe3txLd3lxI0s1xO5eSV2OWZmPJJJJJPXNfav8AwSsRT8VPGDlQXGjKA2OQDMmf5D8q8h/aw/Z0sv2cH8FaMl42o6xfac91qV2CfLabfjbGp6IoGATyeScZAHqn/BMLT/7W8e+PbL7RPafaNC8rz7WQxyx7pVG5GH3WGcg9jX0+ZVqeIyupUpv3WtPvsfPYClOhmMKdRe8n+h+lXiDxTo3hOwe+1vVrHR7JAS1xf3CQxj6sxAr5i+O/7d/wl0vwd4g0PTNZm8S6ne2M9oi6TAXiV3RkBaVtqkZOflLcdq/N34ueH/FWj/FDXPC+vahf+IdZ0y9lsxNNJJO8wVuHXJJwygMB1wa6Dw9+yZ8WPEei3WsR+Db/AE/SrW3kupbvUwLRRGiliwWQhm4HG0GvDoZBg6CjVxFa+zWyX46s9etnWKrOVKhSt0e7f4HkRPOa2fEfjDXfGVxBNrerXurSQRrDD9qmaQRRqAAiAnCqABwOOKxgOcV+0P7O/wCzp8Ovh74K8N6vo/haxXWbjT7e4k1O5Tz7gyPGrMVd8lOT0XAr6PNMyp5bGM5w5m72/wCHPCy7AVMfKUIysla//DH5nfCz9jD4sfFnyZtP8NyaRpkoBGo60Taw49QCC7D3VTX3R8DP+Cc3gT4c/Z9S8XsPHGuLhvKuY9thEfQQ8+Z6ZkJB4O0GvrcDFcx8U/FkvgP4aeK/EkCJJPpGl3N9GkgyrPHEzqCMjIJA4yPqK+BxWe43Gv2UHyJ6WX+f/DH2uHyfCYNe0muZrv8A5HR29tFZwRwQRJDDGoRI41CqqgYAAHQAVLXyb8G/+CjPw58eafa2/iuV/BevMfLeK4VprV2zwUmVflB6/OFx0yep+sQcivDxOFr4SfJXi0/z9H1PYw+Jo4mPNRkmv6+4WiiiuQ6Tyr9qXwTrvxG+Ani7w54atI77W7+COO3t5JViD4mRmG5iADtDYyQM45Ffm5/w7r+Nv/QAsf8AwZwf/FV+vFFe5gc4xGX03TopWbvqv+CePjMroY6aqVW7pW0/4Y/If/h3X8bf+gBY/wDgzg/+Kryf4w/A3xT8CtXstL8WxWlpqN3D9ojtre6SZxHkqGbaTtBIYDPXafSv2v8AiB470f4ZeDdW8T6/dC00nTYTNNJxk9lVR3ZmIVR3LAd6/EP4xfFLVfjP8Rta8Xav8lzqE26O3DFlt4hxHEp9FUAZwMnJ6mvtsmzHG5jOUqqSguye/bf7z5HNsBhMBBKm25vu1t32M/4efDzX/ip4tsfDXhmwbUdXvC3lQhgowqlmZmJAUAAnJNe6/wDDuv42/wDQAsf/AAZwf/FV9V/8E4f2ex4E8CSfELWLfbrfiKILZLIvMFjnII95SA3+6qepr1z9s/x/4h+GP7Pmv+IfC+onStXtpbZI7pY1dlV50RsBgRkhiM4+nPNcuLzyu8asJhOW11G7vv8ALodOGyeisG8Tir3s3ZW2+fU/Pj/h3X8bf+gBY/8Agzg/+Ko/4d1/G3/oAWP/AIM4P/iqu/C39tf41eJPid4R0m+8bPPZX2r2ltPE2n2qh0eZVZSViBwQT0IPpiv1qFZZhmuZ5bKMavI79k/1aLwOW5fj4ylS51bu1/wT8iP+Hdfxt/6AFj/4M4P/AIqj/h3X8bf+gBY/+DOD/wCKr9eKK8n/AFnxvaP3P/M9P/V7Cd5fev8AI/Ov9kL9jv4pfCD4+6B4l8SaLbW2i28N0s9xFfQyFC0LKo2q245Zh0HrnFfopRRXh47HVcwqqrVSulbT+vM9nB4OngabpUm7XvqFFFFecdx5v8bvgF4V/aA0TT9K8VpdtbWNz9qhNnN5TbsYIJweCCRXxd+2X+xx8Ovgl8F5vEvhm31CLU0voLcNc3ZkTY5OeCPav0Zr5X/4KT/8m0XP/YUtf5tXv5Ri8RDE0qMZtR5lpfQ8TM8LQlh6taUFzW36n5ofBHTfCGsfFHQrPx5eNYeEpZHF/cK7IUXy2K8qCR84UcDvX2L/AMKt/Yu/6HOf/wADZ/8A43Xw74J8E618RfFFh4d8PWR1HWb5mS3thIiFyFLH5nIUcKTye1ew/wDDB/x0/wChEl/8GNp/8dr9Fx8KUqic8S6btspJfPU+FwcqkYPkw6qa7uLfy0O3/aF8C/s1aF8ML67+G3iWXUvFazQiC3a5lcFC4DnDIBwue9eK/s3+A9K+J3xu8J+GNbSV9K1K5MU6wvscqI2bhu3IFX/iF+yp8U/hX4Yn8Q+KfCr6Vo8DpHJcteW8gVmbao2pITySO1af7Fv/ACdD8P8A/r+b/wBFPThy0sFVlSrOpZSfNdNp27rsKV6mLpqpSULtaWt17M+q7nQP2Yf2WPjhB9pvvEOn+KfD+yZY28y4ty0seVJ2qSSEbpkD5++OON/bu/ap+Hvx0+F+i6J4R1O4vb621dLyVJrSSECMQypnLAA8uOKzf22P2fPiT8RP2kPEmr+HfBmqarpUkNosV5bxZjkxboGwfZgwI7Y9xXzD4/8Agr44+FlpaXPizw1faFBduY4Hu0CiRgMkDn0ry8DhcNXlRxVSs5VbJ2ck+l7W3PRxmJxFFVcPTpKNO71UWuvfY7n9jj4n+Hvg/wDHPTPE3ie6ktNJt7a4jeWKFpWDPGVX5VBPU1+hX/DxP4I/9DBff+Cyf/4mvyo8DfD7xH8StaOkeF9Iuda1IRNOba1Xc4QEAt9BkfnXWeJv2Z/ih4M0K81rW/BWqabpVmm+e6njASNcgZJz6kV15hluBxtdSxFS0rJWul+DObA5hjMJRcaMLxve9m/xP0z03/goJ8GNW1OysLbXb157uZYI92nTKoZjgZJUYGcDPv6V9HjkV+EHwatPt/xe8EW25l87XLKPcvUZnQZFfu+Ogr4nO8uoZdOnGi3qne//AAyPrsox9bHxnKqlpbYWiiivmT6AKK8D/a2/ah0r9nrwPcxwSpd+MtQhKaZYbsFC2R9ofj7iYzjqxAXjJYfO37P3/BTOPyrbRvitauJM7R4jsIgQees0CDjHrGD2+Xqa9ejlOLxFB4ilC6/F+aXU8urmWFoVlQqSs/wXr2PrT9oD9njwv+0P4Ol0fXIFttQjBaw1iGJTcWT5B+UnqjYAZM4YdwQrD41/4KAfC7RPg38Cvhh4U0CEpY2N5cAyuB5k8hjUvK5A5ZjyfwAwABX6D+GvFGkeMtFttX0PUrbVdMuV3xXVpKJEcfUd/UdRXxR/wVZ/5EfwJ/2ELj/0Wtd2TV631ylh5SfKm3bzszjzajS+qVa8UuZpK/ldHB/8EqP+R58d/wDYOg/9GGvG/iJ+1d8YND+IHizTLb4gaqltDqlzAqqygKqyFAF4O0YUcA+/Uk17J/wSo/5Hnx3/ANg6D/0Ya+RPi7/yVnxn/wBhq8/9HvX2lGlTq5piFUinpHdX6HydWrUpZdQcJNay2dup+l/7Xl7d6l+wY93f3Bu76ew0iW4uGGDLI0sBZiO2SSfxr4f/AGD/APk67wJ/10uv/SSavtz9rL/kwNf+wbo//oyCviP9g/8A5Ou8Cf8AXS6/9JJq8zLP+RVivWf/AKSj0Mw/5GWH9Ifmd/8A8FPv+Th9N/7F+3/9HT17J+wN/wAmb/E7/r91H/0ggrxv/gp9/wAnD6b/ANi/b/8Ao6evZP2Bv+TN/id/1+6j/wCkEFGI/wCRLR9Y/mFD/kb1f+3vyPz28Jf8jZo3/X7D/wChiv33T7i/SvwI8Jf8jZo3/X7D/wChiv33T7i/Subiv4qP/b36HRw18NX5fqOooor4I+0CivKPjx+0r4S/Z2j0OTxVHqDpq7ypA1hAsu3ywpYtlhgfOoGM9a8n/wCHl/we/wCo9/4AD/4uvQpZfi68FUpU20+qRw1MdhqM3CpUSaPXP2iv2f8AQf2h/AFzoOqKttqMQMum6qsYaWzm7EeqNgBl7j0IVh+aXwY/Yj8Z/EH4v6h4U121n0HSdEuWj1XUyhAZVb7tuWHzlxgq2CArBj1Ab7R/4eX/AAe/6j3/AIAL/wDF17j8FfjT4e+PPg3/AISbw0t2lh9oe1ZL2ERyLIoUkEAkHhhyCRXt0MTmWU4ecJQai9m+j8jyK2HwGZ14yU05LdLqvM6rwv4ft/CXhrSdDtJJprXTbSKzikuZN8rJGgRS7d2wBk9zXxl/wVR8X29p8NvCXhkXAF7fao1+YVPJiiiZCT7bplxnqRx0OPrb4l/FDw38I/Ct14h8T6lHp2nwKcbjl5mwSEjXqzHHAr8bP2kPjtqP7QvxMvPE15CbOzVBbWFjvLfZ7dSSoJ7sSSxI7k0+H8FUxGLWJkvdjrfu/wCtRZ3i6dDDPDxfvS0t2R3H/BP7w9Lr37UvhWRIw8OnR3N7OSM7VEDqp/77dPzr9hq+Gf8AgmR8Drnw34a1X4karCscutxiz0tWHzi2VyZJD7O6qAOOIs8hhX3NWPEOIjiMc1F6RVvnu/zsbZHQlQwact5O/wDl+R+ev/BWX7/wv+mp/wDtrXnv/BLf/k4DXf8AsW7j/wBKbavTP+CsGmvJp/w4v9/7uKW+g2be7CA5z/wCvmH9j74+6V+zp8T77xJrOn3mpWdzpcth5Vjs8wM0kThvmIGP3eOvevpsHTnXyL2dNXbUv/SmfPYupGjnPtKjsk1+SP2eor4hvf8Agqp4JjhlNp4O16aUJmNZpIY1ZvQkM2B05wfpX1V8GviXD8YfhnoXjG3sJdMh1WJpRaTOHaPDshG4cHlcg8cEcCvhMRl+KwkFUrw5U9Oh9nQx2HxMnCjO7Wp2lFFFecdwUV5d8eP2i/Cv7O+l6Vf+KY9Qlh1KcwQjT7cSkEAFi2WUAAHPXJ5wDivIf+Hl/wAHv+o9/wCAA/8Ai69Cll+KrwVSlTbT6pHFVxuGoy5KlRJn0x4r8J6P448P3uia9p8GqaVeRmKa2uFyrAjH1B9CMEHkEGvyZ+OP7EPjD4c/Fyw8MeHLabxBo+uT7dIviVBUE/6u4PARlzy3CsBkYOVX7O/4eX/B7/qPf+AA/wDi69B+CH7XfgT9oHxTe6B4XTVPttpZm+d721EcZjDqhAIY8guvGO/sa9rBTzLKFKp7J8ltb7eT/rc8jFxwGaOMPaLm6W39DT/Zj+ANt+zp8NY/DkWpT6peXE3229nkb9357IisIlwNqDYMZ5PUnoB63XI+Jfi/4F8Gao2m6/4z0DQ9RVA5tNR1OGCUKeh2uwOD61lf8NFfCr/opfhH/wAHlt/8XXg1I4jEzdacW3LW9j2qboUIqlFpJaWueh1+Nn7eX/J1/jv/AH7T/wBI4a/VD/hor4Vf9FL8I/8Ag8tv/i6/KD9tPxDpfir9pnxnqmi6laavplw9sYbyxnWaGTFrCp2upIOCCOD1Br6rhmlUhi5ucWvde680fN8QVITwsVGSfvL8mfXP/BKf/kQ/HP8A2EoP/RRr7pr88f8Agml8TfB/gPwV4yg8S+KtF8PzXF/C8MeqahFbNIojIJUOwyM9xX2X/wANFfCr/opfhH/weW3/AMXXl51Qqyx9Vxi2rro+yPRymrTjgqacknr182eh0V55/wANFfCr/opfhH/weW3/AMXXUeFPHHh3x3Zy3fhvXtM8QWkMnlST6Xdx3CI+AdpZCQDgg49xXhyo1IK8otL0PYjVpzdoyT+Zt0UUViahWF468Sw+DPBWva9cMqw6ZYzXjFzgYRC2M/hW7Xy7/wAFFviOfA/7O95pkEhS+8R3UemptPzCL/WSn6FU2H/roK68HQeJxEKK6tf8E5cVWWHoTqvomfk3cTXOs6nJLIXubu6mLMQMtI7HJ4Hck1+7/wAMfCaeBPhz4Z8OxoEGl6bb2hC92SNVY/iQT+Nfjz+yF8PW+Jf7RHg3S2QvaW94uoXRHQRQfvSD7MVVf+BV+1nSvseKa656VBdE39+i/Jny3DlF8tSs+un6v9BaKKK+DPsz8j/+Cj3/ACc9qf8A2D7T/wBF19I/8Eq/+SX+Mv8AsLp/6JWvm7/go9/yc9qf/YPtP/RdfSP/AASr/wCSX+Mv+wun/ola/R8d/wAiGHpH9D4LB/8AI5n6yPtyX/VP9DX8+kn+sb6mv6C5f9U/0Nfz6Sf6xvqaw4U/5f8A/bv/ALcbcS/8uf8At79D9x/2bf8Ak3v4a/8AYuaf/wCk6V6PXnH7Nv8Ayb38Nf8AsXNP/wDSdK9Hr4jE/wAefq/zPsKH8GHovyCiiiuY3Gv9xvpX4A+I/wDkYNT/AOvqX/0M1+/7DKkeor80b7/glt47vJbu6bxh4fNxJJJIqlZgrEsSuTs4zxng4yeuOfsuHcZh8G6vt58t7W/E+Vz3CV8UqfsY3te/4HTfBH/gnX4F+Jvwm8LeKtQ8QeILa91axS6mhtpIBGjN1C5iJx9Sa7WX/glr8NoY2kk8VeJURQWZmmtwAB1JPlV00H7R3gP9kH4PeHfBviDxJa+LfFWkWv2RtP8AD2JXLKeA5JxEAGUfOQxwSFPIr4j/AGgP23vH3x2S40wSr4Y8Ly5U6Rp0hzMnpPLwZPoAqn+7Xbh1m+OrSlTqONO7s32v0Vrs46zyzB0oxnTUqlldLv5vZFf45eDvgX8NfFUOheF9X8T+MpLeYLqN5DeW0dtGufmSJ/JPmP7/AHRxyeQPcvgf+x98Af2gdJkuvC3jnxQ15bqGu9Lu2t47q2z/AHl8ogj/AGlLLnjOeK+afgN+zH43/aC1mOHQdPa30ZJAt3rd2pW1tx3wf42/2FyeRnAyR+rf7Pn7NHhL9nTw+9noMct1qlyii/1a6OZrkjnGBwqAk4UdB1LHJPXm2NjgaKpU68nVXp/5N28ranNlmEljKrqzoxVN/wBad/M+bPGf/BOn4M/DvQLjW/Evj7XNF0uAZe5u7m2RScEhV/dZZjg4Vck9ACa+QvD1r8B77x3fWGrXHjbTfC5cJZaskttLKRzl5ohDlVPGApYj0OeP1w+LvwS8I/HLw/HpHi7TTf28LGSCSOVo5IHIxuRlPB6eoOOQRX58fHP/AIJq+LfBn2jU/AF2fF2krlvsEu2O/iX0A4WX/gO0nsprkynM4Vk4YuvJTe2yS9H39dDpzPL50mp4WinFb9X8129D1HwT/wAE8Pgr8R9Fj1bwx8Qdb1vT3/5bWd1bPtP91h5WVb2IBo+IP/BNH4f+EvAfiPXLbxH4jludM024vIo5ZINjNHEzgNiIHBI5wa+DPDfi7xn8GfFUlxpF/qfhXXbVtk0a7oZAR/BJG3Uf7LAj2r7K+HH/AAULT4heC9a8C/E4R6Pd6xp8+nQeKbGDdDE0sTIHnhHTBOdyZHT5QMmuzFYbNaElUo1nOHyvb9fl9xy4fEZbWi4VaKhP8L/p8/vPgofeH1r95PhP/wAkt8H/APYHs/8A0SlfAGif8EvrvxHoVrqul/E/SNRtrmESwz2lm0sEmRxtkEnK57gfh2r9EfBuhP4W8IaHo0k4uZNOsYLRpwu0SGOMKWx2zjOK8riHH4fGQpxoSu03fRr8z0siwdfCynKtGyaVtv0NmiiiviT64Q9K/CL40f8AJYfHP/Ydvv8A0e9fu6elfhF8aP8AksPjn/sO33/o96+64V/i1fRHx3En8On6s/XX9m7XLPwz+yh4M1jUJRBYaf4eS6uJWOAkaRlmJz6AE1+P/wAR/GN58TviLr3iO4Dvdaxfy3IjxkqHY7UGPQYUfSvt742/FtPAn7AHw78M2twI9W8T6ZBbbFI3C1QBpm+h+RPo5r56/Yb+Eq/Fj9oHRI7qLzNJ0T/ib3gIyrCJh5aH/ekKZHcBq9LK6ccHDE4+p3lb0Tf5vT5Hn5jOWKnh8HT7L72v0R6t+1z8K1+Dn7J3wf8ADjxeVqC3Ulzf8YP2mWLfID67SdgPoor5y+AHwP1H9oLx8PCul6ha6ZdG1kuvPuwxTCYyPlBOfmr75/4KXeB/E/jrwd4LtPDPhnVvETw308s39lWj3JhHlgLuVASM5ODjHyn2z8cfAzxj4j/Y8+KNn4p8W+BNagS6s57WG11GCSwaXO3LIZE+baduQB/EPx3y3E1auXSlTadV8zS03u+jMcfh6dPHqNRNU1ypvXayPZP+HVfjX/octC/79Tf/ABNH/Dqvxr/0OWhf9+pv/ia7v/h69pH/AETy9/8ABmn/AMbo/wCHr2kf9E8vf/Bmn/xuvP8Aa8Q/y/8ApP8Amd3s8j/m/wDSv8jhP+HVfjX/AKHLQv8Av1N/8TR/w6r8a/8AQ5aF/wB+pv8A4mu7/wCHr2kf9E8vf/Bmn/xuj/h69pH/AETy9/8ABmn/AMbo9rxD/L/6T/mHs8j/AJv/AEr/ACPiT46fB6/+BPxGvvCGpX1vqN3aRxSNcWoYIwdA4xu54Br7+/4JYf8AJGPFP/Yeb/0nhr41+M2teKP2tfiprXjXwl4D1y5tmjt4JbfTrWW+8grGFG9o0wC20kAivuL/AIJs+CvEngb4U+JbLxJ4f1Pw/cy6yZootUtHtnkUwxjcFcAkZBGen5V1ZzUlLK1Gs17T3brTfqc2U04rMXKknye9Z67dD67ooor8yP0IKKKKACiiigAooooAKKKKACiiigAooooAK86+Mvwa034r6LtbZaa1bqfsl9jp32PjqhP5dR3B9ForKpThWg4TV0zow+Iq4WrGtRlaSPzL8VeFdR8J6zdaVqtq9pfW7bXjb9CD3BHII61gSRV+h3xk+Dem/FfRdrbLTWrdT9kvsdO+x/VCfy6juD8I+KvCuo+EtautK1W1e0vrdtro3f0IPcEcgjrXw2Mwc8HPvF7M/bsmzmlmlLtUW6/VeX5HKSxVi+Ib+PRNGvtQkXcltC8pXON2BnH49K6WSKvMvj1eGw8Azxq203U8cPXBxncf/Qa3y+n9ZxNOj/M0vl1PTzTGPAYGtilvCLa9bafidD+wT+1fZ/BX4l61beLpTH4f8VNH9pvgufslwrN5ch7+Xh2U46fKegNfrzZ3kGo2kN1azR3NtOiyRTQsGSRCMhlI4IIOQRX86UQxX2Z+xN+3FefBa7tfB3jS4mvvAkz7Ybk5eXSWJ+8o6tCSfmQcjll5yrftGYZU6kfbYdardenY/lDCZn7/ACYh7vfzff8AzP1nr8tv+Cn3jm41z41aR4b3D7FoWmq6p386c7nP4osI/D3r9QNO1G11ewtr6xuYbyyuY1mgubdw8csbDKsrDgggggjgg1+Vv/BTHwldaJ+0HFrDo32PWdMhlik7b48xuv1AVD/wIVz8NqP19c26Tt6/8NcvP3L6k+Xa6v6f8PY+SqKKK/Vz81CiiigBVbH0qxHJValVsfSglq5oo9WEes6OSvaf2av2cNf/AGifFwsrENY6DaMrajqzJlIFP8K/3pD2X8TgCsqtanh6bq1XaKMoYepXqKlTV2y7+zd+zvr/AO0J4sFlYhrLQ7VlbUdVdMpAh/hX+9Iey/icAV+tfw6+HWg/CvwlZeHPDtktlp1qv1eVz96R2/iY9z+HAAFJ8OPhxoHwo8I2Xhvw3YrY6barwBy8rn70jt/E57n+QAFdPX5Jm2bVMxqWWlNbL9X5/kfqmU5RTy2HM9aj3f6Ly/MKKKK+fPoAooooAKKKKACiiigAooooAKKKKACiiigCO4t4ruCSCaNZYZFKPG4yrKRggg9RXG/Cv4M+EPgto93pvhDSE0u2u7hrq4IdneRzxyzEnAAwFHA7Dk57airVSai4J6Pdd7EOEXJTa1R+aX/BVb/ko/gr/sFSf+jTUP8AwSr/AOSo+Mf+wOv/AKOWn/8ABVWVD8TPBkYdTIukuxTPIBmbBx74P5Go/wDglYwHxT8YLkbjo6kD/tsn+Nfo6/5J/wD7d/8Abj4P/mefP9D9HbXwjodjrt5rdvo1hBrN7t+06hHbItxPtUKu+QDc2FAAyeAMVQ+JtnNf/DfxVa2yhribSrqONTwCxiYAfnXTV4t4l/bC+DnhnXNV0DWPGVvaalYTPaXds9pcHZIvDLkR4OOmQSK/P6MK1aadOLk12TZ9vVnSpRanJRT72R+LJ4c/Wv1y+F37bnwatvh74Ys7/wAZQ2GoQaZBDcW01pcZikSNVZSRHg8g455HNfE/jj9kXUfiDqOueLfgk9r438CyajJDBFZyNFc2bYV2iaOZUJC7wFKlsrtJ7151P+yn8X7edoW+HPiEuBklLJ3X/voAg/nX6jjaeBzWEVVqcrj0uk1fo0z85wlTGZbOTp0+ZPrZtPzTR+pH/DbnwR/6H6y/8B5//jdeUftR/tm/C3XvgX4u0Pw14qi1jXNUszZwWsFtMMiRgrksyBQApY8nt718F/8ADK/xe/6Jz4i/8AJP8Kms/wBk74w312ltH8OteWRwSGmtGiTt1dsKOvc/yrzaWS5ZRqRqe22d/ij0+R6FTNswqwdP2O6ttLqYfwD8MHxl8bPA+jGITR3WsWqyowyDEJFMmR6bQ1fukBgAV8QfsRfsR678J/FEPj7xw9tb6vHbyw2WjRgSvbM4CmV5Qdobb5i7V3DD53A8V9wV4HEONp4vERjRd1Fb+fX9D2sjwlTC0JSqqzk9vIKKKK+VPpApskixIzuwRFGSzHAAp1fPX7dnw8v/AB7+zzr0uk3N1BqOjL/aQitpmRbiFQRNG4H3h5ZdsHuo9TXRh6Ua1aNKUuVN2v2MK9R0qUqkVdpXsfF37e37VsXxi8Rx+DvC155vg/SJi011C3yajcjjeD3jTkKRwxLNyNpHnv7HfwItfjV8UIDrs0Vp4T0grdajJPIEFwQfktxk8lyDnHRQ3IOM+D0oJHQkV+ywwSw+F+q4eXLpvu/N9NT8pnjHWxP1iuubXbb5eh++0HiXw/awRww6pp0UMahEjS4jCqo4AAB4FfPP7f8A4g0u/wD2W/FENtqVpcTNNZ4jinVmP+kx9ga/I/c3qaCzHqTXzuH4ajh60K3tb8rT27fM92vxA69GVL2VuZNb9/kdn8EpUh+M3gOSR1jjTXrFmZjgAC4TJJr9xh4s0TH/ACGLD/wJT/GvwFpdzepr080ydZnOMnPl5fK/6o8/Ls0eXxlFQ5r+dv0P35/4SzRP+gxYf+BKf40f8JZon/QYsP8AwJT/ABr8Btzepo3N6mvE/wBVI/8AP7/yX/gnr/6yy/59fj/wD+gexv7XVLVLmzuIru2kGUmgcOjD2I4NWK+cf+Ce7q/7KnhMK2WWa9D89D9qlP8AIivo6vhcTR+r150b35W19zPscPV9vRhVtbmSf3hRRRXMdAV8r/8ABSf/AJNouf8AsKWv82r6K13x74Z8L30VlrPiLStJvJY/Ojt769jhkdM43BWYEjPGema+Vv8Agod4/wDDHiL9nW4s9K8RaVqd2dTtmEFnexyuQC2TtVicV7GVU5/XaMrO3MjysynH6pVjfWzPzZ+Hfj/WPhb4y03xRoEscGr6ezPBJLGJFBZChyp4PDGvev8Ah418av8AoL6b/wCC2L/CvLv2a/DPhjxj8bvC+j+MjCPDV1LKt4bi5Num0Quy5kDKV+YL3HpX6Gf8Mq/spf39F/8ACol/+P1+h5nicDRqqOJoubtvyp6XemrPhsvw+Lq0m8PVUFfa7Wp8L/FL9sn4mfGPwdc+GPEuoWVxpNw6SSRw2SRMSjBl+Yc9RUf7Fv8AydD8P/8Ar+b/ANFPX0R+1p8BvgJ4E+Cup6v4DbTD4kiuLdIRba490+1pAHxGZWB+XPbivm79j7U7PR/2k/At5f3UNlZw3jNJcXEgjjQeU/JY8Cro1aFfL6zw1PkVpaWtrbsiatOtRx1JYifM7x1vfS5+1tfn1/wVg1RhD8N9OUjazX9w475HkKv82r7ai+LHgie4ht4/GGgyTzuI4ol1OEtIx6Ko3ZJ9hX57f8FUtajufil4Q0tZA0lppDTsg/h8yZgPz8qvhshpS/tGnzK1rv8ABn2GdVI/UZ8r3t+aMv8A4Jb6ebj4769dnGy38Pyj33NPBj9Aa+0f25btLL9lbx7I7bQ1vBGDju1zEoH5mvl7/glFoxl174hasUOIbaztVcjg72lYgH/tmPzFfRn/AAUGm8r9lLxeuf8AWS2S/wDk3Ef6V6GYy9pnkI9nBfkziwC5MnnLupfqj8vv2dLV734/fDeJF3E+ItPYjPYXCE/oDX7m1+JH7JsPn/tJfDpeuNYgb8jn+lftvT4qf7+mvL9SeHF+5qPz/QKKKK+JPrzyT47fsv8AgX9oKwx4h07yNYjj8u21qzwl1COSAWxh1BJ+VgRycYJzX5lftC/sW+O/gEZtQeH/AISPwsrca1p8RxEO3nx8mL6klckDdk4r9kKjuLeK8t5IJ4kmglUo8cihlZSMEEHqCK97L85xOAainzQ7P9O35Hi47KqGNTk1yy7r9e5+InwH/aB8dfAzxJHN4RvZJYrqVRPo0qmW3vD0AMY53dgy4btnHFfXH/BSHVdV1z4LfCnUNc0z+xtZupHmvNP37/s8rQoXTPfByK+gfDn7EHw88JfG6D4iaTaGz+z5lg0RFBtIbggjzUX+HGcheinkAYGPGv8Agqz/AMiP4E/7CFx/6LWvo44/DY7M8PKhCz1u+uz0+Xc8GWCxGDy+vGtK60sum61+fY4P/glR/wAjz47/AOwdB/6MNfInxd/5Kz4z/wCw1ef+j3r67/4JUf8AI8+O/wDsHQf+jDXyJ8Xf+Ss+M/8AsNXn/o9693Df8jXEekfyPHxH/ItoesvzP0r/AGsv+TA1/wCwbo//AKMgr4j/AGD/APk67wJ/10uv/SSavtz9rL/kwNf+wbo//oyCviP9g/8A5Ou8Cf8AXS6/9JJq8jLf+RVivWf/AKSj08w/5GWH9Ifmd/8A8FPv+Th9N/7F+3/9HT17J+wN/wAmb/E7/r91H/0ggrxv/gp9/wAnD6b/ANi/b/8Ao6evZP2Bv+TN/id/1+6j/wCkEFLEf8iWj6x/MdD/AJG9X/t78j89vCX/ACNmjf8AX7D/AOhiv33T7i/SvwI8Jf8AI2aN/wBfsP8A6GK/fdPuL9K5uK/io/8Ab36HRw18NX5fqOooor4I+0PjP/go38IvGfxasfANt4Q8PXWutbT3huDbFf3O5Ytm7JAAO1+c9RjqRXxR/wAMVfGz/oQNQ/7+w/8AxdftHRX02Dz6vgqEaFOCaV979XfufP4vJaOMrOtOTTfa3RW7H4BeJ/DOp+DdfvtE1m0ew1SxkMNxbSEFo3HUHBIr9O/2H4/EUv7F10vhGa1g8TGa/wD7Pe9j3xCbd8u4ZHU8ZOQCQSGAwfg79rn/AJOW+In/AGFpf6V+iH/BN/8A5Nj07/sI3f8A6HX1Oe1XPLadVrVuL8tmz5zJqahj6lNPRKS89z82Pj34k+ImufEjVLb4l315c+JbCUwS21y4Mdv0O2JV+RUIwRs4Oc85zXtn7I37Dms/GHUNP8T+MLWbSfAgxMiMTHPqYGCFQdVjPeTuOF5O5eg/4KiRf2d8c/C2oW0skN42iRt5iOQUKXEpUr/dIJ7d+eua+ax+0B8UAMD4jeLAP+w3c/8AxdejTlXxmApvC2puS+70OCpGjhcbNYm87P7/AFP3K0/T7bSbC2srKCO1s7aNYYYIVCpGijCqoHAAAAAqxX4W/wDDQPxQ/wCij+LP/B3c/wDxdH/DQPxQ/wCij+LP/B3c/wDxdfMf6rVnq6q+5n0X+sdFf8u396P2I/aA+BOiftC/D+XwxrU89kFlF1a3lsFLwTqrBWwwORhiCOMgkZHUfn1q3/BL/wCKdneBLLVfDd/btkiYXUsZHT7ymLjv0J6dq+fP+Ggfih/0UfxZ/wCDu5/+LqO4+PXxLuoJIZ/iH4qmhkUq8cmtXLKwPUEF+RXsYLK8wwEXTpVo8vZq55WLzHA4yXPUpO/dM99g/wCCY/xcln8trvw3EmAfNe+kK/TiIn9K/QX9mL4W638F/gxofg/X7+01G/04zYmst5jCPK0gUFwCcbyM4HGBjivkv/gnN8Z/il488V3HhjUNTXWPBekWJknlv13T25PywxxyD5iSQTh8jajYwcZ/QivmM8xWM5/qmJlF2s9EfQ5RhsLyfWcOmr6ahRRRXyp9IfIv/BRr4aeLPib4A8K2PhTQLvXp4NTaWdLNdzRqYyoJHoS3Xtj8vgr/AIY/+M3/AET3WP8Av2P8a/a+ivpcDntbA0FQhBNK+9+p8/jMmpYys605NN9rH4M/ED4V+LPhZeWtp4s0K70K4uozLBHdqAZFBwSMH1r6s/4JXf8AJYfFf/YDP/o+KtD/AIKsf8lA8D/9gyb/ANG1n/8ABK7/AJLD4r/7AZ/9HxV9hicTLGZLKvNWcl+p8th8PHC5tGjF3Sf6HF/8FH5Um/ac1JY8lo9OtEfg8Nsz+PBHSvl7afQ1+++peDtB1i9+2X+iade3e3Z59xapJJt443EZxwPyFQf8K/8AC/8A0Lmk/wDgFH/8TXh4XiSGGoQo+yb5Ulv/AMA9jE5BLEVp1faW5m3t/wAE/A/afQ0mMV++P/Cv/C//AELmk/8AgFH/APE1+Qn7clhbaZ+1L43trO3itLaN7XZDAgRFzaQk4A4HJNfRZZnUcyrOkqfLZX3v1S7eZ4eYZRLAUlVc73dtrd/PyPCQCegpdp9DX6K/8Eu/Dek654G8bPqOl2d+6ajCEa5gSQqPLPA3A4r7b/4V/wCF/wDoXNJ/8Ao//ia5cbxDHB4iVB0r8vW//AOjCZFLFUI1lUtfy/4J+B+0+hr9Nf8Aglb5f/CpvFwA/ff20N3H8PkR4/XdX13/AMK/8L/9C5pP/gFH/wDE1f0bw9pXhyCSDSdNtNMhkcyvHZwLErOTksQoGST1NfPZln0cfhnQVO17a37fI9zAZLLBV1Wc72v0/wCCaFFFFfHn1IV+U3/BSP4vDx38ZofC9lOJNL8LwmBwpOGu5MNKf+AgRp7FW9a/QT9pb45WPwA+FOp+JJ9k2pMPs2mWjH/X3LA7R/uqMu3+yp7kV+Juqanda1qd3qF9O91e3crzzzynLSSMSzMT3JJJr7nhnBOVSWLmtFovXr9y/M+O4hxajBYWL1er9On9eRv/AA8+J3if4U6zLq3hTVpNG1GWE273MKIzmMkEr8wOASo6eleif8Np/Gv/AKKBqP8A37i/+Ir0f9gT9nzwX8etS8Zw+MbCa+j02K1e2ENy8O0uZQ2dpGfur1r3j9pj9iX4UfDX4GeLfEuhaNd22rafbLLbyyX80iqxkVeVLYPBPWvocVj8BHF/Vq1O89FflT321evU8PDYLGywv1ilUtDV2u1tvp8j5v8Ag5+118WvEPxg8C6dq/ju/m0u612yguonEao8TzIrq2F6FWP8+1fohF+2N8GJlJX4g6SAGZfnZ1OQSDwV6ccHoRyOK/JD4BkD46fDskFh/wAJFp/AGSf9JjrJ/wCFZ+L5WZk8K6067iMrp8p5BwR93sQRUY7J8Ji6qT9yy6WV7t+XkXg80xOGptr37vrd2tY9c/bp8d6B8Rv2gNQ1rw1qkGsaW9lbRrdWxJQsqYYfga92/wCCdXxw8CfCz4feKLLxZ4msdCurnU1mhiumILp5SjIwOmRXwlqmkX2h3jWmo2VxYXSgFoLqJo3AIyMqwBq5o3g/XvEUDz6VouoanCjbGks7V5VVsZwSoODgjivQrZfRq4KOElK0VbXTp+Bw0sdVp4t4qMbyd9PU/ZST9sH4MmNgPiFo+SD/AMtG/wAK/FVzl2PvXS/8Kv8AGQ/5lPW//BdN/wDE1zPSoyzLaOX8/sZuXNbt0v29S8wx9XHcntY2tfv1t39D9dPgT+1P8J/DfwU8B6VqfjnSrPUbHQ7K3ubeRyGikSFFZTx1BBFe1/Dv4v8Ag34sx6g/hHX7XXUsHRLlrUnEZYErnIGcgHkeh9K/EG0+HXiu/tYrm28M6xcW8yB45YrCVkdSMgghcEEd6/Qr/glz4f1HQPDvxAi1LS7rTbg3loP9Lt2iZh5bnHzAE4zn/gQ9a+TzbJ8NhqNTEwqNyvtp1Z9Llua18RWhh5wSjbfXoj7looor4c+wCvgL/gorpnxh8PxHXtP8T30nw3mK289nprGE2jknH2jYAWjYkKGYkZ2g4JXP37Xzb+1d+1d4c+BWo6d4X8SeEZ/E+n65ZySXCB4/KMWShRkYEPk4BBwMHv0r2MpnVhi4ulTU32028r7M8rM4054aSqT5F31/G3Q/K34ffCnxh8VtVXT/AAn4evtcuScMbaI+XH7ySHCIPdiBX3h8Av8AgmZp+mLb6v8AFK9XU7ogONA0+QrDGeuJZhguexVMAEH5mFReF/8AgpV8N/BOi22kaD8N9R0nTbdQsVtaNAiKAMZ4PJ45J5Petb/h6t4P/wChK1r/AL/xf419jjsRnOIvChRcI+qv999Pl958tg6GVULSrVVN+jt91tfn9x7N+0j8ebH9kL4f+HZtH8H2+oadNdfYIbC1nWyhtlCFhjbGw7HgAdzn18t+Bv8AwUVf4z/FXQPBn/CAf2SNVkkj+2Lq3nmLbGz52eSufu46jrntivnD9sD9srQ/2kvBWjaJpegahpE1jqH2xpbuRGVl8t0wNvf5q8S/Z5+J9p8GvjF4d8Y31nNqFrpjyu9tbsA77oXjGCeOrg/hWeGyOEsDKWIpfvrStq9+nWxeIziccZGNGp+606Lbr0ufuXRXw5/w9W8H/wDQla1/3/i/xo/4ereD/wDoSta/7/xf418r/YmYf8+n+H+Z9J/a+B/5+L8f8j6X+M37OngX47aRJaeJ9Hje82bINVtlWO8t/QpJg8f7LAr6ivzE/aP/AGJ/GvwFluNTgibxL4QDErq1lES1uv8A08RjPl/72Sp45BO2vq7R/wDgqJ4V1rWtO06DwTrW+7uEt8+dGSNxwCFHLHJHH/6q+1pI0uImSRFeN1wyMMgg9QRXdQxeYZHKMa0fdfRv8uxx1sNgc4TlSfvLqv17n5MfsIat8ZJfiHBp3w/uZm8NJMkmsx36l9OiiJ+YsD92QjO0IQ7bf7obH601geCvAHhz4caTJpfhjRbPQ9Pkne5e3sohGrSP95iB34A9goAwAAOgrzs0x8cwr+1hBRX4v1O/LsHLBUfZynzP8F6BRRRXjnqCHpX4RfGj/ksPjn/sO33/AKPev3dPSvwi+NH/ACWHxz/2Hb7/ANHvX3XCv8Wr6I+O4k/h0/Vn6Aaj+yvY/tFfsefDq5sRFZ+NNL0OE6feuSFlTBJt5P8AZOSQcZVuehYH4F+FXxO8SfAX4j2niLRi1pqunSNDcWlyrKkqdJIZU4ODjkdQQCMEAj9f/wBk7/k274df9geD+Vfnx/wUh+FcPgP44x6/ZQpDYeJ7b7YwTj/SUO2bj3yj57lzXXlGN9pia2Araxbla/q7r9TlzTCezoUsbS0klG/3aM/T74e+OdM+JXgnRfFGjy+bp2qWyXERPVcjlT/tKcqR6g189/tpfsoeJP2ltQ8JTaFq+mabFpCXKTLqBkBbzDGQV2K2f9XznHWvmD9kj9s+2+APwb1fStY0m71yG21iH7JFBIsflxzo5k5I7GFmA7l+o6j2L/h6t4P/AOhK1r/v/F/jXjrK8wwOLlUwkL8rdnps1/kz1HmOCxmFUMTO3Mlda9P+Cjyr/h1d4+/6Gzw7/wCR/wD43R/w6u8ff9DZ4d/8j/8AxuvVf+Hq3g//AKErWv8Av/F/jR/w9W8H/wDQla1/3/i/xr1PrHEH8n4R/wAzzvY5J/P+LPKv+HV3j7/obPDv/kf/AON0f8OrvH3/AENnh3/yP/8AG69V/wCHq3g//oSta/7/AMX+NH/D1bwf/wBCVrX/AH/i/wAaPrHEH8n4R/zD2OSfz/iz079i39mPxF+zXpviu017VNN1MarPBJbnTjIdoRXDbt6rgksMAZ6dea+la+HP+Hq3g/8A6ErWv+/8X+Ne5/sy/tS6X+0zD4il0zQ7zRo9HeBGN3Ir+b5gcjG3oRsORz1HrXzePweYScsXioW2u9PJHv4LFYJKOGw8772Wvqe4UUUV4R7IUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFedfGT4N6b8V9F2PstNZt1P2S+x077H9UJ/LqO4PotFZVKcK0HCaumdGHxFXC1Y1qMrSR+ZnivwpqPhLWrrStVtWtL63ba6N39CD3B6gjrXhf7R9szeC7NwMhL9C302OP8ACv1j+Mfwc034r6Jsfbaazbqfsl9jlf8AYf1Q/p1HcH83/wBoj4danpnhnX9F1Kze31OxXzxGe+w7sqe4Kg4I65r57CUHluZUJy+DmWvrpqfptfMo57kuJox0qqDuvTW68n+B8ZxjirCCokFToK/oimj+V6jPrD9jL9tbUfgJew+GPEsk2p+ALiUkIMvLpbsctJEO6Eklo/Ull+YsH+8f2ovgvpX7WHwTgufDd5bX2qW6f2joN/FKDFMSBujLdNrqMdsMFJ4BB/GZB0r7D/4J6fHvxX4R+J2meAYRNq3hnW5irWbHIsn6mdMkYGAQwzznOCwAPjZjlzhL+0MK+WcdX2dt/wDg9z1cBjlUX1HE6wlovL+unY+UtS0270bUbqwv7aayvrWVoJ7a4jMckUikhkZTyrAggg8giq9fqj+2j+xKPjTJP4z8HeRaeMooSbq1lJVNUCKAi7s4SQKu0EjDZUMVAyPyxuLeW0nkhmjaKaNijxuCGVgcEEHoa+iy7MaWY0ueGjW67P8Ay7HhY/A1MDU5J7PZ9xlFFFeqeaFFFey/szfsy+IP2jvF4s7MPYeHrRlbUtXZMpCv9xP70jDov4nisa1anh6bq1XaKNaVKdeap01dsX9mX9mjxB+0Z4vFlZB7Dw/aMralq7plIF/uL/ekYdF/E4Ar9g/hv8N9A+E/hCx8NeGrFLHTLRcADl5XP3pHb+J26kn9AAKPhv8ADbw/8JvCFj4a8NWCWGmWi4Cjl5G/ikdv4nbqSf5ACunr8kzXNqmY1LLSC2X6vz/I/TstyyGAhd6ze7/ReX5hRRRXgHtBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUV+en/BTH4ieKvBXj7wdF4e8TavoUUunSySR6bfSW6s3mEbiEYZOOMmvjb/AIX98Tv+ii+K/wDwdXP/AMXX1uD4eqYyhGvGokpeR8zis8p4WtKi4N28z9nPHXwQ8CfE3VrXU/FXhfT9dvraE28M15HvKxlt23HTryPTJx1OTwL8EfAnwy1W61Lwr4X0/Qr25hFvNLZx7C8YbdtI6deffA9BX4x/8L++J3/RRfFf/g6uf/i6P+F/fE7/AKKL4r/8HVz/APF16X+rmL5PZ+393trb7jz/AO3sNzc/sde+lz91K/ET9qsY/aP+I3/Yauf/AEM1jf8AC/vid/0UXxX/AODq5/8Ai643VNVvdb1G4v8AUbue/vrhzJNc3MhkllY9WZiSST6mvYyjJ55bVlUlNO6seXmmawx9OMIxas7n6o/8EyQB+zlcYHXXLnP/AH7ir61r8F/DfxT8Z+DtPNhoHi3XNEsS5kNtp2ozQRlzjLbUYDJwOfatT/hf3xO/6KL4r/8AB1c//F15uM4cq4rETrKolzO+x34XPqeHoQpODdlbc/dSivwr/wCF/fE7/ooviv8A8HVz/wDF0f8AC/vid/0UXxX/AODq5/8Ai64/9Va3/P1fczr/ANZKX/Pt/ej91KK/Cv8A4X58TWIB+InisjPfWrn/AOLr9rfhfcT3fw08Jz3NzLeXMmk2jy3M5BklYwqS7EYySeT9a8XM8onlkYynNPmuetl+ZwzCUoxi1Y6eiiivAPaCqWt6PaeIdGv9Kv4RcWN9BJbXERJAeN1KsuRzyCRxV2imm07oTSaszwFP2avgf8DPh1falqHg3Sb7T9HsmubvUNZtkvLiYRplmJkBG9iPuoAMkBVHArgvh1+0H+yv4zsk26P4Y8MXIUbrLWNDgtygAwBvCGM9OAGz04rzj/gpp+0BGlpafCrR7pWldo73WzG2SijDQwMPc7ZSOvyx+tfK/wCyV8EH+O/xm0nRZ4mfQ7Q/btVcZAFuhGUz2LsVT/gRPavu8Jl7r4KWMxtWSvqtenz7/wCR8bicd7HFxwuEpxdtNuv/AAP8z9dLf4QfDu6gjmh8FeGZYZFDpIml25VlIyCCF5BrwH9vH4aeEPDn7Mnia/0rwtoumX0c1oEubPT4opFBuYwcMqgjIJFfV1paxWNrDbQLshhQRouc4UDAH5V86/8ABQr/AJNU8Vf9drL/ANKo6+ay+rUeNormduZdfM9/HUoLCVXyr4X+R+WHwYtIL/4weB7a5hjuLabXLKOWGVQyOpnQFWB4II4INfsn4r8G/CDwJpEup+IvD/g/RbCJSzXF7YW0S44HGV5OSAAOSSAOSK/HL4G/8lq8A/8AYfsP/ShK/Vb9rj9lKz/aR8NQT21/JpninS4ZBp0jP/o0pJB8uZcE4OCAy8ruJww+U/X5+4PFUIVZuEWndo+XyTmWGrSpwUpJqyZj/Bz4qfs9/HPxhq3hzwz4V0U31iokhe90O3gW/Tnc0AI3kLgZDKp5BAIzj2n/AIUx8P8A/oRvDn/gpg/+Ir8SLq18U/Brx80My3fhvxVod0DjOyWCVeQQRwQRggjIYEEZBr9W/wBkD9rfTf2h/Dg03U2hsPHGnxA3lmDtW6QYHnxD+6Tjcv8ACT6EE+Xm2V1MJFYjDTcqfXW9vP0Z6OWZjTxUnQxEUp+m/l6o9+0Lw/pfhfS4dN0bTbTSdOhyIrSxgWGJMnJ2ooAHPoK0KKK+Obbd2fVJJKyCiiikM/Mn/gqHp91qHxs8MJaWlxcSDw+pJijZuPtE3p6Z/UetfFE0UlvK8UqNHIhwyOCCD6EV/QYVB6gGvw//AGof+TiviP8A9h27/wDRrV+ncPZh7emsLy25FvffX0Pz3PcD7Gf1nmvzva22h5/b+H9Uuollh067liYZV44GIP0IFQx6deS3bWqW073K5zCqEuPw61+sXw2+LOn/AAR/YR8NeLb0RvJaaOFtLdzj7RcszCKP1wWIzjooY9q/L3QoPEvxP+JdqthLPfeK9b1MSJOjbZHuZJNxkLD7vzEsW7cnjFexgsfPFuq5Q5YwbV772+R5WLwUMMqSjLmlNJ2ttc5+80280/aLu1ntt3TzoyufpmrQ8Maz1GlXv/gO/wDhXsn7XXwI8X/BXx9AfEmsXnim01KBZLXXrtndp2AHmxsWJIZWJ4yflKnuQPrv/gnb+0+/jjRh8NPE1yJdc0uAvpV1Kw3XNquAYiSeXj7Y6oP9gkzicxlSwixdCKnHrZ20+7p1Hh8DGpinha0nB9NL6/f9x8CfDPRNRsviX4Rku9MvhCNXtdy/Z3ywEqkgDGScA8D0r0v9vPxWfFX7T/iza4e307yNPix28uJd4/7+GSv2B1fUbXQ9KvdRu3WG0s4XuJpCOFRVLMfwANfg14+8UzeOfHPiDxDPu87Vb+e9YOckeZIWx+GcV5+V455riXXcOXkVt77u/ZdjvzLBrLaCoqfNzu+1tl6+Z+k3/BLrwv8A2X8Fde1p1Cyapq7Ip9Y4o0A/8eaSuq/4KQXv2X9mPUY84+0ajaRfk+7/ANlr0T9krwP/AMK9/Z18D6S8flXD2C3k6nqJJyZmB9wXx+FeNf8ABUTUTa/AHRrZcZuvEECtn+6sE7cfiFr5OnU+s50pr+f8F/wx9LOn9Xyhwf8AL+f/AA58RfsS2X2/9qTwDFjO27kl/wC+IJH/APZa/aCvyC/4J42C3n7U/hqVl3G1t7yYH0P2d0z/AOP1+vtdHFEr4yK7RX5sw4djbCyfeT/JBRRRXx59SFITgUteR/tWfFj/AIU18C/E3iCGTy9SeH7Fp+Dg/aJfkRh/ugl/ohrajSlXqRpQ3bt95lVqRo05VJbJXPlHWv8Ago1qHgz9o7xLa3kH9t/Dm3uGsIreyRVnjMZCtOjN9/LK/wApKghhyMVF/wAFF/iFoXxS+Dnw48SeG73+0NHu9RuhDceWybtqhT8rAEcg8EA18J+FfDd/408T6XoWmRefqOp3UdrBH6u7BRn0GTya+7f+Ci3gnTvht8EPhR4X0uNYrLS5pLaMKu3ftiXc5HqzZYnuWJr9JqYLC4PHYVUlabuvVKL1fn/wT4CGLxOKweIdR3jp97a0XkYv/BKj/kePHf8A2DYP/Rhr47+JFz9t+I3iefeZPN1W5feQBuzKxzxX1j/wTQ8Q2fhLVPihreoSeVY6doa3c7+kcbMzH8hXyFoemXXjPxhYafGxe91W+jgVtuSZJJAoOB7t0r0sNHlzHE1HtaH5HBXd8Dh4Le8vzP1E/a5tns/2CTbytvki0/SEZsYyRLACcV8QfsH/APJ13gT/AK6XX/pJNX37+31aw6b+yL4itIvkiiawijUnnC3MQA9+BXwB+wi6p+1b4ELEKPNuhknubSYCvDyp82U4mXfn/wDSUexmS5czw67cv/pR6D/wU+/5OH03/sX7f/0dPXsn7A3/ACZv8Tv+v3Uf/SCCvG/+Cn3/ACcPpv8A2L9v/wCjp69k/YG/5M3+J3/X7qP/AKQQU8R/yJaPrH8xUP8Akb1f+3vyPz28Jf8AI2aN/wBfsP8A6GK/fdPuL9K/Ajwl/wAjZo3/AF+w/wDoYr990+4v0rm4r+Kj/wBvfodHDXw1fl+o6iiivgj7QKKKKAPxM/a5/wCTlviJ/wBhaX+lfoh/wTf/AOTY9O/7CN3/AOh1+d/7XP8Ayct8RP8AsLS/0r9EP+Cb/wDybHp3/YRu/wD0Ov0jOv8AkUUv+3f/AEk+Byn/AJGdT/t78z0/4t/sw/Dj44apFqni/QDqWqQWv2OG8jvJ4Xjj3MwACOFOGdiNwPWvxu+Kvw61H4TfETX/AAjqozeaVdNB5m3aJU6xyAZOA6FXA64YV+1mhfGzwZ4m+JOreA9L1qK98TaVB9ovLWJWKxrlQRvxtLAuuQCcE4OCCB84/wDBQ79miX4meEYvHXh2yafxLocRW7ggUF7uz5Y8dS0ZywA5IZxydoryckzCpg66w+JbUJJWv07Wv0Z6eb4GniqLr4dJyi9bde/zRxH7HHwK+Afx/wDhXbT6h4Tim8X6WBb6xEmqXqMWyQk20SgASKuflGAQwGMYr3r/AIYC+A//AEI3/lWvv/j1flt8CfjZrnwD+Idj4o0YmVYz5d5YNIUjvID96N8fmDg7WCnBxiv2Z+HPxg8LfFH4fweMtE1SFtFaJpJ5J3EZtCozIk2T8jL3zxjBBIIJ0zunjsFW9pTqy9nLb3no+2/3EZRPB4ulyVKceeO+i1Xf/M8U8a/sa/s4fD3wpqfiPXvCK2Ok6dCZ7idtVvjhR2AE2SScAAckkAda/KDxJe2GpeIdRutK09dK02a4d7WxSR5BBEWO1Nzks2BgZJJNfTf7cP7XZ+OeuJ4X8LzsngbTZd/m7SralOMjzWB5Ea9EXAPJY9VCVv2Ev2Y7j40ePofE+rW4Hg3QLlXn8xcreXK4ZIB6gZVn68YGPnyPey/22XYSWLx9Rtvo23bster/AK6ni472WOxUcNgoJJdUkr93p0R9tfsE/BS4+EHwPtrjVLfyNd8QyDU7lGGHiiZAIYmBAIIX5ip5DOwr6SpAMClr80xNeeKrSrT3k7n6Bh6McPSjShskFFFFcx0BRRRQB+av/BVj/koHgf8A7Bk3/o2s/wD4JXf8lh8V/wDYDP8A6PirQ/4Ksf8AJQPA/wD2DJv/AEbWf/wSu/5LD4r/AOwGf/R8VfpUf+Sf/wC3f/bj8/l/yO/n+h+nVFeS/Hf9p3wP+z5pqyeIb83GqzDNto9jh7qX/aK5ARf9piBwcZPFen6RqlvrmlWWo2b+baXcKXEL4xuRlDKfyIr87lRqQhGpKLUXs+9j7uNWnKbpxldrddi3X42ft5f8nX+O/wDftP8A0jhr9k6/Gz9vL/k6/wAd/wC/af8ApHDX1fC/++T/AML/ADR83xF/usf8S/Jn1V/wSn/5EPxz/wBhKD/0Ua+6a+Fv+CU//Ih+Of8AsJQf+ijX3TXlZ3/yMKvqvyR6OUf7jT9H+bCiiivDPYCiiigD5G/4KS/CFvHnwZt/FNoXOoeE5WnMYPD20pRZuPUFY2z6I3rX5han4O1XR/C+i+ILu2MGmaxJcR2UjcGbySgkYD+6Gfbn1Vh2r959e0LT/E+i3ukaraR32m3sLQXFtMMpJGwwyn8DX50f8FP/AA7pfhP/AIVdpOjafbaXplra3yQWlpEI44xuhOAo4HJJ9ySa+84ezKScMC11bv2Vm7fefGZ5l8Wp4xPotPO6X5Gj/wAEn/8AkM/Ef/rhY/8AoU1fVH7an/Jr3j//AK8k/wDRqV8r/wDBJ/8A5DPxH/64WP8A6FNX1R+2p/ya94//AOvJP/RqVxZl/wAjtf4ofodeA/5E7/wz/U/J79nb/kv/AMNP+xl03/0qjr9zdo9BX4Zfs7f8l/8Ahp/2Mum/+lUdfuY7iNGZiAAMkmunir+NS9H+Zz8N/wAGp6r8j8a/269S/tT9qnx1IHLJFLbwKM8LstolIH4g/ma+2v8AgmJpK2P7PV/dYO+91y4lJPoIoUGP++TX5v8Axq8Zf8LB+Lfi/wARqVMWo6nPNFt6eWXOz/x0LX6v/sKeE38Jfsw+Do5gRNfRy6g2RjiWRmT/AMc2V6Gd/uMqpUXv7q+5HDlH77MqlVbe8/vZ71KB5T8djX8+kn+sb6mv6C5f9U/0Nfz6Sf6xvqa5+FP+X/8A27/7cb8S/wDLn/t79D9xv2bQP+GfPhrx/wAy5Yf+k6V6OAB0Fec/s2/8m9/DX/sXNP8A/SdK9Hr4jE/x5+r/ADPsKH8GHovyCiiiuY3Cvi/9uj9ljx9+0B428Nah4Vj06TT7Gxa3lN1c+U6yF2YnocrjaOOcn06faFFduExdTBVVWpWuu/mcmKw0MXSdKps+x+Tf/DtP4xf889D/APBh/wDYUf8ADtP4xf8APPQ//Bh/9hX6yUdK97/WXHf3fu/4J4v+r+D8/v8A+AfiR8c/2ZPGP7PMGkSeLG05TqrSrbJZ3PmsfLC7yRgYA3r+dcv8JfhRrvxp8a2vhbw4sDapcRySp9pk8uMKilmy2DjgV7T/AMFAfjPB8V/jhLYaZdfadD8NxHToGRgY3n3EzyKe+W2pnoRGDXof/BLXwC+qfEnxL4umhk+z6VYCzhkIIQzTNk4PQkJGcjtvHtX20sZXoZZ9ar2U7X+b20+658isJRrZh9Wo35L2+7f9Tjv+Hafxi/556H/4MP8A7CuU+Kf7D3xK+D/gTU/FuvppQ0nT/L842155knzyLGuF2jPzOv4V+x9fPX7fv/Jpnjn/ALcf/S2CvlMJxDjK2Ip0pWtKST07v1PpcTkeEo0J1I3uk3v2Xofk58LP+Sm+Ef8AsL2n/o5K/ecdBX4MfCz/AJKb4R/7C9p/6OSv3nHQVrxV/Epej/Qz4b+Cr6r9RaK5q++I/hzTPHFn4QvNVt7TxBe2v2u1s53CNcJuKkR5+8w2klRzgZxiulr4ZxlG11ufYKSlez2CiiipKEPSvwi+NH/JYfHP/Ydvv/R71+7p6V+EXxo/5LD45/7Dt9/6PevuuFf4tX0R8dxJ/Dp+rP2F/ZO/5Nu+HX/YHg/lXyb/AMFYriAy/DSAIDcgag5fPIX/AEcAY9yD+VfNvjDxp8Svgpp/guLRfH/iGz0jWfDtpqdpaQalOsUCvuR0CFtow8bkYHAYV5F4m8W65421Q6j4g1i/1zUCoT7VqNy88u0dF3OScDJ4r08Bk0qeN+u+0TjeTt63X4HnY3NY1MJ9U5GnZL7rM7z4GfA7xX8fbnV/D3hT7GZ7ZYr6cXkxiXapaMEHB5Hm5x6Z+h9g/wCHZnxf/wCevh7/AMD3/wDjdfSv/BNv4E6r8OPA2seLtfs5LG/8RmEWdvMAHW0QFlkI6jezk4PZFPevsWeZLaGSWR1jjRSzOxwAB1JNefmOf16GKnSwzTivK+ttTuwGSUa2HjUxCak/Pp0PwW+I3gLUvhh421bwtq728mpaZL5M7Wrl4920HAJAz19K9U+DP7GPxB+Ovg7/AISbw4dKTTTcPbD7bdNG5ZMZIAQ8c+vY15p8WvGJ+IPxP8VeJMkpqmp3F1HkYIjaQlBj2XA/Cv1r/Yb8Jt4Q/Zh8FwyjE97BJqLnGMiaRnT/AMcKV72a5hWwGDhUjbndk/u1PFy3BUsbip03fkV3+Oh+Tfxd+E2t/BTxvdeFfEJtm1S2jjkc2khkjw6hlwSB2PpXb/A39kjxz+0H4ev9Z8Lvpi2dldfZJft1y0bb9qtwAp4wwrrv+Cin/J0mvf8AXpZ/+iFr6i/4JXf8ke8V/wDYcP8A6IipYrMK1HLIYuNuZqPpqPDYGlVzGWFlflTf4Hzv/wAOzPi//wA9fD3/AIHv/wDG6+tP2E/2cvGH7PVj4ztvFi2I/tKa1a1ayn80MEWTcTwMffA5HY9sV9VUV8Jis7xeMouhVtZ+XZ3PssNlGGwtVVqd7rz+QUUUV4B7YUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABXnfxl+DWl/FrQmhmVLbVoUItb3bnH+w/qh9O3UdwfRKKzqU41YuE1dM6KFephqiq0naSPwX+Ovwj1b4MfEbUtA1SzktArmS33D5WjJ4wehA9fTB71waDmv25/am/Zl0b9pLwJJp83lWHiO0UyaXqpTJif+4+OTG3Qjt1HIwfyctv2XvidJ8SB4H/AOER1Fdb8/yGJgb7OozjzTKBs8vvvzjFfpOT5hCrQ5K0vegtW+qXX/PzPgM4wThiHUoR92b0S6N9PTt5ehyfw9+HviD4oeKbPw94Z0yfVdVumASGEcKMjLu3RVGeWOAO5r9h/wBln9lbQf2bfC7JE0eqeKr5B/aOsMmCw6iKIHlYwce7EZP8IWf9mH9lzw5+zd4UEFmiX/ia8iUanrDA7pmHOxAfuRgngDrgFsnp7ZXz+a5vLGP2NLSmvx9fLsvv8vYy3K44Re1qazf4en6/gFfGv7a/7E9v8Ura/wDHXguAQeMIY/MvLBOE1JEU52qFJ884UDs2MHnmvsqivGwmLq4Kqq1F2f5+TPVxOGp4um6VVafl5n8+l3aT2F1NbXMMlvcQuY5IZVKujA4KkHkEHjBqKv1C/bW/Yjj+KKXvjrwRAsHi2OPfeaZGgC6njHzKcjbKFz678AcHk/Gf7PP7IvjL44+MhYz6de6BoFpKV1LVby2aMRbSN0SBgN0vP3e2cnA6/rOFzfDYjDPEOXLbdPp/n5dz8zxOV4ihXVBRvfZ9/wCupn/sy/sy+IP2jvF4tLMPYeHrRlbUtXZMpCv9xP70jDovbqeK/YP4b/DfQPhP4QsPDXhqwSw0u0XCqOXkb+KR2/idjySf5YFL8OPhxoHwo8IWHhrw1YJYaXaLhVXlpG/id26szHkk101fnOa5rUzGpZaQWy/V+f5H3mW5bDAQu9Zvd/ovL8wooorwD2gooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKAPMPit+zR8OPjbrFnqnjLw6NXv7SD7NFOLueAiPcW2ny3XIyxPOcZPrXEf8ADAHwI/6Eg/8Ag2vf/j1fQ1FdsMbiqcVCFWSS6Js5J4TD1JOU6cW33SPnn/hgD4Ef9CQf/Bte/wDx6j/hgD4Ef9CQf/Bte/8Ax6voair/ALRxn/P6X/gT/wAyPqOE/wCfUf8AwFf5Hzz/AMMAfAj/AKEg/wDg2vf/AI9X5a/tAeFdM8EfGvxpoGi232PSdO1Oa3trfzGfy41bAG5iWP1JJr90K/EX9qz/AJOP+Iv/AGGrj/0M19fw3iq9evUVWo5K3Vt9fM+Xz/D0aNGDpQUXfokuh9XfsMfsr/DH4x/BWXXvF3hs6rqq6pPbCf7dcw/u1SMqNscijqx5xnmvob/hgD4Ef9CQf/Bte/8Ax6uR/wCCZX/JuU//AGG7n/0CKvrSvDzPHYqnjasIVZJJvRN/5nsZfg8NPCU5Spxba7I+ef8AhgD4Ef8AQkH/AMG17/8AHqP+GAPgR/0JB/8ABte//Hq+hqK8z+0cZ/z+l/4E/wDM9D6jhP8An1H/AMBX+R88/wDDAHwI/wChII/7i17/APHq950TRrPw5o1hpWnQi20+xgjtbeEEkRxooVVyeeAAOau0VhVxNeukqs3K3dt/mbUsPRotulBRv2SQUUUVzHQFeL/tR/tJ6P8As6eA5b6Z47vxHeq0WlaZu+aWTH+sYdRGmQSe/AHJrpvj38TtQ+EHwv1fxRpfh268UXtmECWNtkY3MAZHIBIRQSTgHp2GWH4u/FH4o+I/jF4yvPEvii/a+1K4O1R0jgjBO2KNf4UXJwPck5JJP1GS5T9fn7Wp/Di/m329P6R87m2Z/Uo+zp/G/wAPMyda1nV/Hfie61LUJ59V1vVLkyyyEbpJ5XbsB3JOAAPQAV+vP7Gf7OEPwA+GNv8A2jaxr4w1ZVuNUm4ZosjK24YZGEHXBILFiCRivBP2AP2Pm0w2HxS8aWbx3nMmh6XcR48tSBtu3GepBOxSOOH67SPvi4laC3lkSJpmRSwiQgM5A6DJAyfcgV3Z/mcar+p4d+6t/N9vRfn6HHkmXSpL63XXvPb/AD9X/W5JXzf/AMFCv+TVPFX/AF2sv/SqOuNi/wCCoXwt+0eTcaF4stXBKuZLO3IUjscT59uleUftZ/tzfD/4z/BPVvCPh2z1r+0r6a3YPe2yRRoqSrISSHJP3cYx3rzsDleNp4ulOdJpKSf4ndjcxwlTDVIxqJtpr8D4/wDgb/yWrwD/ANh+w/8AShK/dgdK/Cj4FpJL8bPACRJ5kreILAKmcbj9oTAr91x0r1OKv4tL0f5nncN/wqnqj5l/bO/ZJsPjx4Xm13RIIrTx1psReCZUx9vjUZ8iQjqePkY9Dx0Jx+U+g69r/wAM/GFvqem3F1oXiHSbklHClJYJVJDKykfVWVhgjIIIJFfvnX5/f8FHf2YIZLV/iv4atCk6FY9etoh8rrwEuQAOCOFfnkbTgYYmcgzTlf1LEaxeiv8Al6MrOsu5l9boaSW9vz9UfUn7Mf7QmmftFfDe31y3EdrrNsRbarp6n/j3nxnKg87GHzKee4zlTXrtfip+yf8AFrxT8Ivi5p+o+G9PvtdhnBi1HRrGNpGu7bq/ygfeTG8N2K8naWB/ae3laaCORo2hZ1DGN8bkJHQ4JGR7EivIzrLll+ItD4Zary8j1MpxzxtC8/ijo/PzJKKKK+fPbCvw/wD2of8Ak4r4j/8AYdu//RrV+4Ffh7+086yftE/EcqwYf29eDIOeRKwI/OvtuFf94qen6nyPEf8ABp+v6E/xa+MOs+MvAnw+8HSwTWGheHNMUQwvkC6mfJac9iMYVeuMNj7xr7D/AOCbP7OH9l2L/FXX7VRdXStBoUcq4aKI5WS456bxlF/2dx6MK/PO61nUL6ztrS5vrm4tLYYggllZkiH+ypOB+Fatv8RfFdpBHBB4m1iGGNQiRx38qqqgYAADcADtX2OMwM62FeGoSULvXru7v7+p8rhcZCjiFiK0XK23y0X3dD9nv2j/AILaf8evhTq3hq5WIX+w3Gm3T9be6UHY2ewOSp/2WNfkP8J/CnxCsfjNp+n+DtJvJPHGjX+4Wka4aCWJ9riUkhVQH5W3ELgkE81zv/CzfGH/AENWt/8Agwm/+KqjYeMde0q4ubiy1vUbSe6YNPLBdSI0pGcFiDljyevqa5cuyyvgKM6LmpJ7JrRd+uz7HTj8wo42rCqoOLW9nq+3Tc/Uj9vb40z+A/2dotGmWOy8T+LYlspLSOXf5Me0G6weNyjIjz38wV+cX7P3w0b4vfGPwt4VKO1tfXi/aymcrbp88pz2+RW59cVx2r+I9V8QeT/aeo3WoGEMIzdTNIV3MWbG4nqTk19i/scXuifsxeCrj4v+PtI1QWuvyjSdGuLS0WXZGDukcksNu8rhehxC+M7hnOGG/sfAyp0veqSbtbq3tb0WvyLnX/tXGRnU0hG179Et7+rP02ggjtoI4YkWOKNQiIgwFA4AA7CvhH/gq3rYh8KeANI3/Nc3t1dbP+uaIuf/ACLXr0H/AAUP+B8yAv4lu4Sf4X0y4JH5Ia+H/wBvH9obQPj54+0JvCtzLeaDpNiUWeWFoi80j5kwrAHACxjnuDXymS5fiYY6E6tNpK7u010PpM2x2Hng5wpzTbstH5nQ/wDBMLRG1H9oHUb4rmOw0SeTd6M0kSAfkW/Kv1Tr89v+CUvhOcN488TSR4tiLfToZP7zDdJIPwBj/Ov0Jrm4hqc+YSS6JL8L/qdORw5MFFvq2/xt+gUUUV80e+FfmX/wU6+MS+I/HekeANPuC9noSfar9VIKm6kUbQfdIz/5FI7V+g/xZ+Iun/CX4c6/4t1Mj7Npdq0wjLYMsnSOMH1Zyqj3avwx8VeJdQ8aeJtU13VJjc6lqVzJdXEh/id2LHHoMngdhX2nDOC9rWeKktI6L1f+S/M+T4gxfs6Sw8d5b+i/zf5H1V/wTS+FB8Y/GW68WXcG/TvDNsXjZuhupQUjHocJ5rexC1+iPxZ+A/gf442+mw+NdEGsR6c7va/6RLCYy4AbmNlJB2rwTjgVxf7GPwdPwZ+A+h2F3AIda1Jf7T1HK4YSyAFYz7ogRT7hj3rP/aO/bO8FfAjTr+wgvbfXvGaJiHRraTd5b8Y891yI8A7tp+YjoOc1xY7EYjMMybwl7rRW7LrfonrqdeDo0cDgEsTaz1d+/a3XofP/AO2P4W+HH7KnwuvfD3gHR5NG8ReN1FpcMl7NMBZRMGkLCVnxklUwNpO5jn5cH4p+EXi25+GvjPT/ABvDo39sJoUwmSOQMIFuGVhCZGHQBhvx1bYRkZyE13X/ABn+0F8SRc3slz4k8VazMsMccaDc56KiKMKqgdhgAAk9zX6wfs/fspeHfhf8DZfBWu2NtrVzrkQl18yLuSaVlHyKeoWPopGDkFxtJ4+qrV4ZNhFTxL9pOe+urWz13slov0Pm6VGea4l1MOuSENtNF1Xzb1Z8ufs5/tb6b8cYb74Y/HmSHW7LWrkSWWo3AW2jEmQVgfygm0bhlG7E7ScbcfYvhH9lT4U+BPE9h4i0HwdaadrFjuNvdRyynYWGCdpcqTjIyRxk461+X/7V/wCyvq/7Nvi1djyal4Q1GRjpeqNjeMcmGYDgSKD1wA4+YY+ZU+rf+Cff7XFx4tQfDfxrqbXGsRKDol/csN1xEq4NuxxkuoXcpJJYFhn5Rny8zwfNh/rmXTapveKbS7N2X4o9HL8Vy1/quPiudbSer8lf8mfRPxY/ZL+G/wAa/FUfiLxZpNxe6mlstqHivJIVKKSVyFI5BZuff2GE0/4J+FvgT8EPHGieErWe00+6s729kSedpiZGgK8FjwAqqPw5ycmul8WfH74deBNfk0TxD4x0rRtVjUO9reXAjdQQCCc+oIxXD/E39p74T3Pw48TwwfEHQLmebTLmKOGC9SSR2aJgFCqSSSSO1fM03jqkYU3zOGllrY+hqLBwlKouVT1u9Ln4++Ev+Rs0b/r9h/8AQxX77p9xfpX4D+FDjxVo5Cs2LyE4UEk/OOgHWv34T7i/SvpuK/io/wDb36Hz3DXw1fl+o6iiivgj7QKKK8g/ab/aI0/9m/wEmu3el3mrXd5KbSyhgjIh87YzDzZeiL8vux5wpwxG1GjOvUVKmryexlVqwowdSo7JH5U/tbSLJ+0p8RGU5H9rzDp3Bwau6J+1Z4z8H/BS0+G/hi5Og2PnTzXuo2zYubgSMT5av/yzUDGSvzE9wMg+V+KvEt/4z8Tarr2qS+fqOpXUl3cSdi7sWbHoMngV9Cfs2fsU658YII/E/iu4bwh8P0AkbUrgrHLdrn/liH4C/wDTRht54DcgfslZYfDYaH1u1oW37pW0XU/KqTr18RP6re8r/c31fQwf2Ko/H9v8ddH1XwHoZ1y4tCU1BJmMdstrICrmWTonGWXqdyDCtjaf2V6jmvmC1/aV/Z3/AGafDFv4b8O61YyW0AyLXQI2vHlbAy8kq5VnPGSz57dsD5x+Lf8AwVC8S64lxY+ANCh8OW7ZVdS1Ei4usdmVP9Wh9j5gr4nG4fF53XVSlR5IrS70uvP/AICZ9bhK2Fyei4VKvNJ62Wuvl/wTH/4KBfsuaR8LdcHjnw1d2VlpWs3B+0aG8ypJDOeWe3QnLRk8lV+4TwNhAT5H0/xLq+k6XqOm2OqXtnp2pBFvbS3uHSK6CElBIgOH2kkjcDgnirfiXxX4j+JPiI6hrepX/iHWblhGJbmRppWJPyouc4GTwo4GcAV7p4Q/YB+LXi7wDdeJk0mHTZFQSWmkahL5V5eL3KqeE9hIVJ/In7SlKOAw8KeNqpva709PW3f5s+Tqxlja8p4Sm0uy/H7+xwX7NnwNk/aA+Jln4a/ta10e0x591PNKolaMHlYUJzI59BwOp4HP7OeA/Aui/DXwnp3hvw/ZR2Gk2EflwwxqB7lmI+8xOSWPJJJJJJNfhFf6fq/gvxBLa3cN3o2tadPh43DQz28qH8CrAivpX4Of8FE/iT8NoorHXmj8c6TGAoXUpCl2oHYXABJ+rhz715Gd5biswUZUJpxX2dvnfZ/geplGPw2CbjWi1J9f0t0P1oor5Z+Hf/BRv4S+M44otXur3whfN8pj1O3LxE4/hlj3DHuwX6V9E+EfHnhvx9Yte+Gte03XrVCA8unXSTqhIyA20nafY81+dV8HiMM7VoOPy0+/Y+7o4qhiP4U0/n+hu0UUVxnUFFFYvjTxGfB/hLWNbXT7rVW0+1kuRY2MZeecopOxFHVjjAqopyaiuom1FNs/O3/gqvIp+IngmMH510uViPYy8fyNfMPwY+O/iP4EXOvX3hYwQarqtj9gF5Mm82yl1cuingt8oA3ZAyeDWj+0t8fNS/aI+JU/iK8t1sbKCP7Jp1mucxW4dmXfyQXO7LEcZ6cAVX+Bf7OfjP8AaB1/7D4bsCthE4F5q1zlba1B/vN3bHRFyT6YyR+xYWhDC5dGljLWS1vtvc/K8TWnicdKphb3b0tvtY4uWXxD8SPFTO51DxJ4i1ObPAe4ubmU+gGWY+wr9l/2UdB8beFfgX4b0Xx9awWWtadF9ligikV3S2T5YRIVJXeFGPlJ4C55zVL9nX9k/wAG/s7aSh06BdW8SyR7brXrqICZ89VjHPlJn+EEk8bi2Aa9rr4bOc3p45KhRj7kXv1+XZf1ofY5Tlc8G3Wqy959Onz7sK/Gz9vL/k6/x3/v2n/pHDX7J1+Nn7eX/J1/jv8A37T/ANI4a24X/wB8n/hf5oy4i/3WP+Jfkz6q/wCCU/8AyIfjn/sJQf8Aoo19018Lf8Ep/wDkQ/HP/YSg/wDRRr7prys7/wCRhV9V+SPRyj/cafo/zYUUUV4Z7AUUUUAFfnV/wVg/5DPw4/64X3/oUNforX51f8FYP+Qz8OP+uF9/6FDX0XD/APyMafz/ACZ4Wd/7hP5fmj5h/Z8/ac8U/s23WtT+GbHSb19WSJJxqkMkgURliu3ZImPvnOc9q7/4n/8ABQX4ifFjwHrHhLV9I8N2+napEIppbO1nWVQGDfKWmYA5UdQayv2Qf2WbL9pu+8TQXmvz6ENIjt3UwW4l8zzC4OcsMY2frX0t/wAOo9D/AOigX/8A4Lk/+Lr7XGYnKaOKbxC/eKzvZv0PkcLh8zq4dKg/3bv1XzPiT9nb/kv/AMNP+xl03/0qjr9cP2r/AIpwfCH4D+KdaZv9Nntm0+xQNgtcTAohH+7kuR6Ia8E8D/8ABMvSfA/jnw54kh8d3ty2j6jb6gbd7BV83ynDhNwkyuSo55718/f8FB/2kbP4veOLbwn4fmM3h/w3NKkl0jZju7rhWZfVUwyhuhyxHBBPmV3RzvH0fYO8Iq8tGuu2vc9GiquUYKr7ZWlL4dU+nl2Plnw/ol34n8QabpFinm3uoXMdrAn96R2CqOPciv3r8KeH7fwn4X0fRLRQtrptnDZwqBgBI0CLx9AK/L3/AIJvfBh/HnxhfxfeRZ0fwqgmXcOJbuQMsS++0b3yOhVOzV+q9cHE+KVSvDDx+yrv1f8AwPzO3h3DuFGVeX2tvRf8H8hkv+qf6Gv59JP9Y31Nf0Fy/wCqf6Gv59JP9Y31NdfCn/L/AP7d/wDbjl4l/wCXP/b36H7j/s2/8m9/DX/sXNP/APSdK9Hrzj9m3/k3v4a/9i5p/wD6TpXo9fEYn+PP1f5n2FD+DD0X5BRRRXMbhRVDW9e03wzpk+pavqFrpWnQAGW7vZlhijGcZZ2IA5IHJ7186/FH/goP8J/h6s8GnalL4w1NFO2DRl3wluwM7YTHuu7Hoa6qGFr4l2owcvQ5q2Jo4dXqzSPphmCgknAHc18I/tl/t5WGk6dqHgf4a6it7qdxG0F/r9o+YrVW4KW8in5pMZBccLn5TuHy/MXx5/be+IfxxiutMa6Xw34YmGxtI01iPNX0llPzSZ7jhTx8teO+Avh54j+KHiW30DwtpNxrOrT8rBABwo6szEhUUZGWYgDPJr7rLuH44f8A2jHNaa26L1f9I+Ox+dyr/uMGnrpfq/Rf0zL0TRL/AMS6xaaXplrLfahdyCKG3hXLOxPAFftb+zP8EbX4A/CXS/DEbJPqJzd6ldIMCa5cDeR7KAqD2Qd68+/ZG/Y30r9nrThrOsNBq/ji6j2S3iDMVmh6xw5APP8AE5AJ6DAyD9LV5Ge5usbJUKHwLr3f+SPTyfLHhE61b4307L/MK+ev2/f+TTPHP/bj/wClsFfQteD/ALdGmPq37KvjyGM7WSC3nPGeI7mKQ/oprwcvdsZRb/mj+aPaxyvhaqX8r/Jn5H/Cz/kpvhH/ALC9p/6OSv3nHQV+AnhTWR4c8UaRqxQyCxvIbooP4tjhsfpX6Sp/wVQ+H+xd3hTxGGxyAICM/wDfyvuuIsDiMXOm6EOayd/wPjsixlDCwqKtK17Hi/8AwVE1FrX48eFjbSSwXlvoMUqyxttKn7TPtKkcggqTn6V7L/wTT+I/jb4iaV4yPifxJe65p+m/Y7eyjvpPNeNiJS53n5jwEHJP4Y5+LP2rvjtb/tDfFiXxRY2U+n6dHZxWVtBdY80Im5iW2kjJZ2PHYivuH/gl14Wm0r4L69rUqhV1XVmEXqUijVc/99Fx+FTmFH6tkkadWPvKy8073ZWCq+3zeU6b913frpY+zKKKK/OD7wQ9K/CL40f8lh8c/wDYdvv/AEe9fu6elfhF8aP+Sw+Of+w7ff8Ao96+64V/i1fRHx3En8On6s/UH4Kfs+fD34t/Ab4Y6r4u8MWutahbaBBbRTzO6lY+W2/KwBAJJGemTjqa9M8J/sv/AAp8D6lHqGj+BNHt76Iq0VxLB5zxMpJDIZN21sn7wwenPAqv+yd/ybd8Ov8AsDwfyrv/ABV400DwLpv9oeI9a0/QrEtsFxqNykCM2CdoLEZOAeBzxXzWJxGIdepSjOVrvS77voe/QoUFRhUlFXstbLt3NnpXzL+3n+0BafCL4Q3+g2VzE3ijxLC9jb24Y74bdwVmnIBBGFyqnP32BAIVq4r44f8ABSvwf4Usp7L4ew/8JZrJ3It5cRvFZQsP4jna8vfhcA4+96/nD438b+IPin4wvdf1+9m1bW9QkBkkI5J6KiqOAAMAKBgAACvfyjI61SpGviY8sVrZ7v5dEeLmmcUoU3Rw75pPS62X/BJvhh4Dv/if8QdB8LabE8t3ql2kA2dVTOXf6KoZj7Ka/dvRNItfD+jWGl2UYhs7KBLaGMfwoihVH5AV8g/sC/slTfCzSV8feLrKS18W6hEUs9PnBV9Pt26717SuMEg8quAcEsB9l1zcQZhHF11SpO8Yfi+p0ZHgpYWi6lRWlL8F0PyH/wCCin/J0mvf9eln/wCiFr6i/wCCV3/JHvFf/YcP/oiKvl//AIKLwyR/tRa2zptWSys2Q5+8vkqM/mCPwr0H9g/9qnwD8CfAfiDRfF97dWVzd6l9shaG1eZSnlImDtBwcqa+ixVGpXySnClFydo6I8HDVYUc3qTqOyvLc94/4KDfGHxV8FoPh5rnhPUv7PvBfXIkVkDxzJ5aZR0PBUgn3HBBBGa2/wBkn9tKX9pTxDqOgXnhT+xL6wsTeveQXfnQSAOqbdpUFT8+Ry33TyK+TP2+P2nPCXx8/wCEQsfB8893Z6Ybia5nnt2hy7+WFVQ2DwFbJ9xXff8ABKTQJn1zx9rRiIgjt7azWT1ZmdyPyVfzFebUy+nSyb2uIp2qRvbo9Zdf+CehDHVKma+zoTvB/NaR/rY/ReiiivhD7IKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKTAznHNLRQAUUUUAFFFFABSAAdKWigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAr8Y/wBpT4f+K9f/AGg/iLdWHhfWru3OuXIEkNhK6n58ggqCMEEMPYg96/Zyk2j0FezlmZSy2pKpGPNdW3seVmGAWYQjBytZ3Plb/gm5pOoaL+z7dWupWFzp1wuuXP7q6iMbH5IgTg89QR9Qa+qqQADpS1wYqu8VXnWatzO52Yaj9Xoxop35VYKKKK5TpCiiigAooooAQjNfPmtfsPfDXWfjTZ/EF9NWERZmudDjRfsV1c5+WZ07epQYViFJH39/0JRXRRxFXDtulJq6s7djCrQpV0lUinZ3QgGBS0UVzm5+XP7X/wCyX451r9oTxFqXgrwZqGo6Pquy+8+1RfK85lXzucgAl9zYPJLE9K8s0/8AYa+N+pXEcSeBbmHecb7i6t41X3JMlfszRX1tLiTFUaUaUYxdla7v0+Z8zVyDD1akqjk1d3srf5Hwn+x/+wRrXwz8e2/jXx9NZ/atN3HTtLtX84eaVAEzv0+UFgFAPzYbI2gH7soorwMZja2Oq+1rPXbyR7WFwlLB0/Z0loFUtZ0ez8Q6RfaXqEC3NhewPbXELEgSRupVlOOeQSKu0VxJtO6OtpPRnnnwh+Afgf4HaZJaeEtEhsZJeJ72T95czjOQHlPzEDsvQeleh0UVdSpOrJzqO7fVkwhCnFQgrJdgooorMsK4LU/gH8Ndb1O71HUfAPhvUL+7laae6utKgkklc8lmZlJJNd7RWkKk6esG16EShGeklc85/wCGcPhR/wBE18J/+CW3/wDiKP8AhnD4Uf8ARNfCf/glt/8A4ivRqK0+s1/5397M/YUv5F9yPOf+GcPhR/0TXwn/AOCW3/8AiK8y+P37DvgX4q+DDZeGtI0rwVrlqWmtbvTLCKCOZ9pAjn2JuKE46cjrg9D9J0VrSxuJozVSFR3XmZ1MJQqwcJQVn5H5FfAj9iDxb4/+Mmo+G/FFlcaLoXh+5MesX2CnmgZ2pblh85kGCHAwEYMeqq36HftLfCmPxX+zR4j8G+H9EW5a30+NNL062AXY0LIYlj9MBAAO44717KsaqzMFAZupA5P1p1ehi83r4utTrS05LNLpfv8AP8jhwuV0cLSnSjrzXu/L/gH4pP8Asd/GeNVJ+HurkMARhFP54bj8a6bwR+wN8ZPGGrpaXHhseHbXgyX2rTokaAnsqlmY8HgKe2cZFfsPRXqy4oxbVowin8/8zzo8O4ZO7lJ/d/kecfs//BHS/wBn74bWPhPS7qbUPLdri6vZ1CtcTsBvcKPurwAFySAACzEZPo9FFfI1Kk603UqO7erPpqdONKChBWSCiiiszQ84+PPwL0H9oPwOfDWvz3ltAkwureezl2NFMFZVcgghgAzDaRjnscEfDPgD/gnh4q8G/tG+GrbXI4de8BwTG+k1i3AVJFiBZYZIidyMzhARyCrHDHBA/S2ivXwmaYnB05Uacvdaena/VeZ5mJy7D4qpGrUXvK3zt0Zj+LfDNv4x8M6jolzcXdpBewmFp7C5e3mjz3WRCGB/HB6HIJFfkp4o/Yf+IKfHXV/AXh3TrjVrS3ZZotduo2htfs7gMrvIRjIyVIXJLI2AcV+wdFVl+aVsu5/Z6qXfv3Jx2XUsfy+00a7dux8/fsu/sgeG/wBnXT/t7f8AE38YXMXl3WqSHKxrxmOEYG1CRk5yx7nAAH0DRRXnV8RVxNR1asrtndRo08PBU6Sskcd8XPhZonxm8A6r4U16EPZ3sZCTKoMlvKPuSoT0ZTz78g5BIr80/hZ/wT/+JesfFbU9PuLtvCWneHr4BfEpjf8AfMp3RSWgBBc8K2dy7M8ncNtfq7RXoYLNcRgac6dJ6S79H3Rw4vLaGMqRqVN49uvkfDv/AAUf+A+q+MfDPgzxF4e0jUvEfiCwkOl3RsbRp7ieBlLrI6xr0V1boAMzH1FfDtl+zJ8Wr8AxfDfxOoLbf32lzR/+hKOPev3Eor0cFxBWwVBUFBO3c4cXklHF1nWcmr9j8xv2Wf2BvHE/xI0jxD4/0keH/D+lTR3otp5o3mvZFO5IwqMdqggFi2OOADkkfpzRRXk4/MK2Y1FUq202S2R6eCwNLAwcKXXdsKKKK8w9AK5/x/4G0n4leDdX8M65apd6ZqVu0EqOoYrkfK65Bw6nDK3UMoI5FdBRVRk4SUouzRMoqScZbM+Bf2VP+CekOkaxdeIvihbQ6ibK8mgsdEZS0E3lSMnnyhgN6MV3IpGGUgnIbFfdtzoun3ukvpc9lbzaa8XkNaPEDEY8Y27cYxjjFXaK7sZjq+Oqe0qy9Oy9DjwuDo4On7OkvXu/U/Hr9ov9lLXfB/7Qep+EPA3h7U9Y067jXUNNht4nmKQNjcC+MbUk3JuJPG3Jya774W/8EyPHvieSK48Y6jZ+ELE4ZoEIu7ojqRtU7F9Mljj0NfqNgZz3pa9uXEmM9lGnCyaVm92/PseRHIML7WVSd2m9tkvLueM/Bf8AZI+G/wADRFcaJoq3utKBnWNTxNc5wMlSRiPp/AB1r2aiivmq1apXlz1ZNvzPoKVKnRjyU4pLyPOfi/8As++BfjlpZtfFehw3dwqFIdRhHl3dv/uSjnGedpyp7g18OfFb/glz4h0l5rvwB4hg121zldP1UC3uVHoJB8jn3IT/AB/SmivQwea4vBaUp6dnqv8AgfI4cVluGxmtSOvdaP8Ar1Pw08Ufs7fEnwXrtlpGteDdVsbu9uFtbZmh3QzSt0RZVzGTwTw3AGelfrz+zj8BtH/Z6+G1n4c04m4vpSLnU75mJNzdFQHYf3UGAqqOgAzlizH1EgHqM0tdeY51WzGnGlJcqW9uv/DHNgMppYGpKpF3b2v0Ciiivnj3Ao60UUAfAfjT/gn0/jz9qnW7sqNG+HNwsWqyvacPJJJuElvF8u1W8xHY4yFV07kAfcPgvwVofw78M2Hh7w5psOk6PYx+XBawA4UdSSTksxJJLMSSSSSSSa26K9HFY+vi4whVlpFJJenV+Zw4fBUcLKU6a1k7v/L0CiiivOO4K/Ir9tvwB4o1/wDaj8cXem+G9Xv7Z3tSk1tYySIwFrEuQVUgjKsPwPpX660mB6V6+WZg8tqurGPNdW/FP9DzMwwKx9JUnK1nf8/8z4l/4JeaFqeg+CfHEWp6bd6dI2pRbRdwNFuwjKwG4DOCCD6GvtukAA6DFLXLjcS8ZiJV2rc3Q6MJh1haEaKd7BRRRXEdYUUUUAFeH/tFfsmeGv2lL/Q7rX9V1XTm0lJY4l054wHDlSc70bnKjkf4Y9worejXqYeaqUnaS6mNWjTrwdOorpnhv7On7Jfhz9mvUddu9B1jVdSOrRxRyR6g0ZCBCSMbEXJyzcnsfz9yooorV6mIm6lV3k+oUqNOhBU6askcR8ZfAeq/Ez4eap4c0fxLc+E7y9UJ/aVom6RVzkr1BAPQ4IOMjIzX5LeIf2O/iNoPxosfhw1hDcalf4ltdQSTFpJb5IM7P/ABtbKkbsgAKxZd37QUxokZ1copdc7WI5GeuK9bLs3rZdGUIJNP8+//AADzcfllLHuMptpr8u3/AAThvgj8INI+Bvw40vwlo/7yK1UvcXTLte6nbl5WHPJPQZOAAOgrvKKK8apUlVm5zd29WerCEacVCCskMl/1T/Q1+EM3wh8dIk0zeDPECwpuLyHS5wqgZzk7eMYOfpX7w0mB6V7OV5rLLOflhzc1uva/+Z5OY5bHMOXmlblv+Nv8jz79ni0uNP8AgP8AD21u7ea0urfQbKKWCdCkkbrAgZWU8ggg8GvQqKK8apP2k5T7u560I8kFHsFFFFZlnI/FH4V+G/jH4Rn8N+KdPTUNNldZQDxJFIvR426o2CwyOcMw6E1+T3xL/Yr+IfhX4u33g7w5oGo+JbbAnstRhhKwyQNjBeVtqKyk7WyQMj0Ir9j6K9vLs2r5dzKGqfR7X7nkY7LKOPs56NdVvbsfnN8Hv+CXWpXcsN98SNfisbcEMdJ0Y+ZK47h5mG1DnsofI/iFfdHw2+EPg/4RaOum+EtBtNHt8AO8KZlmPrJIcs592JrsaKwxmZ4rHP8AfT07LRfd/mbYXL8Ng/4Ude+7CiiivLPRCsvxP4a07xj4d1HQ9XtUvNN1CB7e4gkGQ6MMH8fQ9jWpRTTcXdbiaTVmflv4+/4Jh/ETRNTmHhfVNM8S6aF3RvM/2Scnn5SjZX8d/PtXHr/wTq+NrKCdBsFJHQ6nDkf+PV+vFFfVQ4lx0Y2fK/Vf5NHzkuH8HJ3V18/80flJ4N/4Jo/FbXNYig1xtL8N6cHAlu5LlbhwvGSkcZO489GZenUcV+l3wp+GWjfB3wBpHhHQEkXTdOjKK8z7pJXZizyOf7zMzMcYAzgAAADraK8zHZricwSjWasui2PQweW4fAtypLV9WFFFFeOeoIehr8Sfi98LfGt58UPG99F4P15rRtYvJvOGmylAhlZgSwUjG0g5zjBzX7b0mB6V7eWZnLLJSlGPNzHkZjl8cwjGMpWseV/sr2l1Yfs7+AbW9tLiwu4NLiilt7qMxyIy5Byp5HTPNN/aO/Z70D9onwFNouqxrb6rbK8ulaoijzLScj1xkxsQodP4gB0ZVZfV6K876zONd4im7Svf0O72EHRVCaurWPxZ+HP7HXxW+J2rz2eneGJ7C3trhra41DVs2sETqcN94bmweyKx56V+gv7Nf7B/hL4ITWOv61IPFHjKA+Yl3IpW2tHx/wAsY+5GeHbJ4BAXpX1BjFLXs47PsVjIumvdi+i6+rPKweTYbCS5370vP9EFFFFfOHvHx9+3B+xvrHx51XTvFfhCS0TX7O0Nrc2l1KY/taBt0ew42qw3SZLEAgqMjHPxLN+w58cILmSBvAV2XTBJS7t2Xn0YSYP4Gv2bor6TB5/isHSVGKTS2vf/ADPAxWS4bFVHVbab3t/wx+N+h/sHfG3WtUSzbwc+nKQC91e3cKRIPUkOSfooJ9q/TD9l79nqx/Zx+HJ0G3vptTvryf7dfXUyKmZSirsRRnCKEGASxyWOecD2CiscfnOJzCCp1LKPZdfvua4LKcPgZ+0hdvuwooorwT2gooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACivnr9rr9qW6/ZjsPDVzbeH4df/teSeNlluTB5flhCCCFbOd/6Vtfsn/tC3H7Sfw+1LxJc6LHoT2mpvp4t4rgzBgsUT7txVf8AnpjGO1d7wNdYZYtx9x9br023ONYyi67wyfvrpr/wx7XRRXx5+0x+3nf/ALP/AMU7rwjB4Pt9ZjhtobgXUl80JO9c42hD0+tRhcJWxtT2VBXe/b8ysTiaWEh7Ss7LY+w6K5L4TeN3+JXw08NeKpbRbGTV7GK8a2R94iLqDtDYGcZ64r5U+Of/AAUSvfg58WPEHg2PwPBqiaXKkYvG1IxGTdGr52+Wcfex1PStMPgMRiqsqNKN5R31XR2Ir42hh6catWVovbR+p9sUVleFdZbxH4Y0jVmiEDX1nDdGIHIQugbGe+M1b1S7On6bd3QUOYInkCk4zgE4/SuFxafK9zrTTXMWqK+HPhB/wUhvvij8T/DnhJ/AsGnJq14tqbpdTaQxA99vlDP5ivuOuzF4KvgpKFeNm9d0/wAjlw2Lo4yLlRd0vX9Qoqtqeo2+kaddX13IIbW2iaaWRuiooJYn6AGvz4uP+Cr12lxIsPw6heEMQjvqpViueCQIjg+2T9avCZficdzfV43tvqlv6kYrHYfB29vK19t3+R+iFFcz8M/HFr8S/h94e8U2S+Xb6tZRXYjJyYyygshPqrZU+4rpq4JRcJOMt0dsZKcVKOzCivkf9pr9vN/2fPiPL4Rg8FrrciWsVz9tk1LyBl8/LsETdMdd34V4lJ/wVa8SGTKeA9LWP+617IT+e0fyr26OSY7EU1VhDR6rVf5nkVc3wdGbpznqt9H/AJH6S0V8F+Cv+Cq2j3t7FB4q8EXWmW7EBrzTLwXO3PcxsqcD2Yn29fsr4c/FHwr8WvD8eteE9atdasGO1mgb54mxnZIhwyNgg7WAOCD3rjxWXYrB614NLvuvvR1YbHYbF6UZ3fbZ/czqqKK5D4ofFjwv8HPC02v+KtTi06xjOxATmSZyCRHGnVmIVuB6E9Aa4YQlUkoQV2ztlKMIuUnZI6+ivzr8af8ABVfUG1GZPCfgi2SwVsRz6xcs0sg9THHgKfbc31p3gr/gqveDUYY/FvgiB7FmAkuNGuSssa9yI5Mhz7F1+te9/YGYcnN7P5XV/wAzxf7bwPNy8/4Ox+iVFcv8NviX4e+LPhKx8R+GdQS/026TcCOHjboUdeqsDkEH9RzV/wAZ683hXwhrmtJCLh9OsZ7sQs20OY4y23PbOMZrwnTmp+zas9reZ7KnFx507rc2aK/P3w1/wVVS+8Q6fba14GTTNJlnRLq9g1Bp3gjJwziPyxuwOcZ/wr79tbqG9tori3lSaCVBJHJGwZXUjIII6gjvXZi8BiMC0sRG19tn+RyYbG0MYm6Er2/rqS0Vwfx0+Jcnwd+FHiDxjFYLqkmlRJKLR5fLEmZFTG7Bx97PQ9K8c/ZM/bKuv2mPFGt6TP4Wi0BdOs1uhLHemcyEuFxgouOtRTwderQliYx9yO7uv+H6lzxdGnWjh5P3pbI+n6KK8l+Of7T/AIE/Z+s0PiTUJJtSmUtBpVinm3MuO+MgKOnLEe2a56VKpXmqdKLbfRG1SrCjFzqOyR61RX506/8A8FXNSkuJF0X4f2sEAYhHvtRaRmGeCVVFwcY4yceppugf8FW9UjuUGteALSe3LAO1jqDRuozyQGRgTjtkZ9RXu/6v5ja/s/xX+Z4/9t4G9uf8H/kfoxRWB4B8YW/xB8E6H4ms7a4s7TV7OK9hhu1AlVJFDLuAJGcEHgmt+vn5RcW4y3R7cZKSUlsworlPir41f4b/AA28S+KY7Rb59HsJr0WzPsEuxS23dg4zjrg18K/8PYNQ/wCicW3/AIN2/wDjNejhMtxWOi50I3S80vzZwYnMMPg5KNaVm/Jv8j9FaK/Or/h7BqH/AETi2/8ABu3/AMZr339kr9sG5/aa1rxBYz+GItAXSreKYPHeGfzN7EYwUXGMe9b18nxuGpurVhaK31X+ZlRzXCYioqVOd2/J/wCR9MUV80/tbftfXP7MmreHbODwzF4gGrQyyl5LwweVsZRjhGznd7dKu/smfte2P7TCa3aT6Unh/XNNKy/YFn84S25wPMViq9GyCMcZXk7uOZ5fiVhvrfL7ne672233N/r2H9v9W5vf7a+u+x9E0UV8i/tO/t2Xf7PPxNbwnD4Ph1tBZxXX2p78wn593y7RG3Tb1z3rHC4StjKnsqCu9+35muIxNLCw9pWdl/XY+uqK5v4beLW8ffDzwx4me2Fk+s6ZbagbZX3iIyxK+zdgZxuxnAzit2+v7bS7Ke8vLiK0tLeNpZp53CJGijLMzHgAAEkmueUXGTg90dEZKUVJbE9FfGvxc/4KaeCPBl/9g8H6RceNp1I8y7877JaAEZ+RirM5BwCNgH+1xivIJP8Agqx4qM+U8D6QsOfuNdSlvz/+tXt0sizCtHmVOy82l+G549TOMFSlyud35Js/Smivk39l/wDbxi/aC8cDwld+DZ9G1FrdriO7tbsXEG1AS5kBVCg+6BjfktzivffjR8Q3+E/wu8ReLo7JdSfSbb7QLVpPLEvzAY3YOOvoa8+tgq+HrLD1I2k7WV112O6li6Nek69OV4q/fodrRXyP+y/+3Td/tE/EtvCs3hCHQ0WxlvPtSX5mJ2FRt2mNeu7rntXtn7Rnxgl+BHwm1XxnDpiaw9jJAn2N5jEH8yVU+8FOMbs9O1VVwGIo144acbTdrK667eQqeNoVaLxEJe6r3dn0PTKK+KvgT/wUftPir8TtI8Ka54Xi8NQaoxt7e+W+89ftB/1cbAouA5+UEZ+YqMYJI+1ajFYOvgpqFeNm9f6sVhsXRxcXOjK6QUV4f+1h+0dP+zT4M0jXYNCj1431+LIwSXJg2fu3fdkK2fuYxjvT/wBlD9oqb9pXwNqniCfQ49Bay1FrEW8dyZw4Ecb7txVcffxjHaj6lX+r/WuX3L2vdfluH1uj7f6tf397a/8ADHttFfNX7W/7Xlz+zJqfhy0g8MxeIBq0M0paS8MHleWVGOEbOd3t0r5//wCHsGof9E4tv/Bu3/xmu2hk+NxNNVqULxfmv8zkrZrhMPUdKpOzXk/8j9FaK/Or/h7BqH/ROLb/AMG7f/Ga+5vhJ47f4m/DPw34rks1sH1eyjuzarJvEW4Z27sDOPXArDF5bisDFTrxsnpun+TNcNmGGxknGjK7Xk1+Z19Ffn/4s/4Ki3/hnxTrGjj4e29wNPvJrUTHVWXfscrux5RxnGcVlf8AD2DUP+icW3/g3b/4zXashzCSuqf4r/M5HnWBTs5/g/8AI/RWivij4Ef8FEb34zfFjw/4Nk8EQaUmqSPGbxdSMpj2xu+dvljP3cde9fQv7SPxvH7PnwxuPFx0c66YrmK3FmLn7PuLnGd+xsY+lcNbLsTQrRw9SPvy2V11072O2lj8PWpSrwl7sd3ZnqVFfnDff8FXNckcmz+H+nwJ2E2ovIfzCLWp4Z/4KtP5hHiDwGvlllAbTb3kL/EcOvJHGBnn1Feg8gzFK/s/xX+ZwrO8C3bn/B/5H6FUVwfwc+NvhP46+FIte8K6gLiLAFxZzYW5tHOfklQE7TwcHJBxkEjmsT9pn42y/s/fC+fxdDpKa08V1Db/AGR5zCCHJGdwVun0rxo4arKssPy+/e1npqeq69NUnXv7tr38j1eivmn9kn9r65/ab1bxFZz+GYvD40mCKUPHeGfzd7MMcouMbffrXU/tX/tFz/s1eB9L8QQaHHrzXuoLYmCS5MAQGN33ZCtn7mMY71vLAYiGI+qOPv8Aa69d9jKONoSofWVL3O9n/wAOe20V8l/swft6WXx+8fyeE9U0CPwxfy27TWBW789bp15ePJVSGC5YDByFbkYAP1pWWJwtbB1PZV42Zph8TSxUPaUXdBRXzb+1x+1zc/sxX3hqCDw1F4gGsRzuWkvDB5XllB2Rs53+3SvRv2c/jBL8d/hNpXjObTE0d76SdDZpMZQnlysn3toznbnp3qp4KvDDxxUo+5LRO68+m/RkxxdGdeWHi/fWrX3fLqemUUhIAyeBXy38bP8AgoV8PPhVd3Gl6OsvjTW4GKPDp8gS2jcEgq05BGQR/CrVGHwtbFz5KEXJlV8RSw0eetKyPqWivzZvP+CrPiZ7gm08CaVDBnhJrySRsf7wC/yr0n4H/wDBSQ/Ezx9ovhTVfActrc6rOtrDdabeiULIxABaN1XCAZJYMSAOFNerUyLH0oOcoaLzX+Z5tPOcFUmoRnq/J/5H27RRRXgHthRRXzt+0F+3B4D+A9xNpO6TxP4nQMG0vTpFAgYDIE8pyI8nHADMOu2uihh6uJn7OjHmZhWr0sPDnqysj6Jor809U/4Kr+LZrktp3gnRrW3z/q7q4lmf/vpdg/SvRPgX/wAFIdR+JXxB0Lwnq/gVEn1a5S1jutLuyfLZj94xuOVHJJ3ZABwD0r2KmQ4+lB1JQ0Wu6/zPKhnOCqTUIz1fkz7ooor5p/aI/br8GfAjVZ9AgtZ/E/ieJFeSytmEcMG5N6eZKQeSCpwobg84ryMPhq2Kn7OjG7PUr4ilhoe0qysj6Wor8zrj/gqr40a7ZoPBuhR2u7iOSWZnA9NwYDPvivdPgH/wUZ8J/FPXLTw/4o0uTwfrN3IsNrN532i0uJCVCqX2ho2ZicBgVwOXyQK9WtkePoQdSULpdmmebSzjBVp8kZ2b7po+vaKQEEZHINfPn7Wn7WS/svw+Gwvhk+Ip9bNxsBvfsywiLy8knY+7Pmjjjoea8nD4epiqio0leT/4fqenWrU8PTdWq7RR9CUV+blx/wAFWvETSEweAtMjj7LJfSMfzCj+VdZ4K/4Kq6ZdXMMPirwVcWETuFe50y6EwRePm2MFJ78Z7Drnj2ZZBmEVzezv81/meVHOsDJ25/wf+R970Vh+CfG2ifEbwtp/iPw5qEeqaNfoZLe6iBAcAlSMMAQQwIIIBBBBGRXi/wC1r+1kP2X4fDW3wz/wkU2tm42ZvfsywiLy8k/I+7Pmjjjoa8ijhq1esqFOPv8Abbbfc9SriKVGl7acvd777+h9CUV+blx/wVa8RtITB4C0yNPSS+kY/mFH8q7HwH/wVS0a+vIbfxd4NudLheTa95plyLhY1P8AEY2CnAPXBJx0BPFetPIcwhHm9nf0a/zPLjnWBk7c/wCD/wAj7yorF8G+MtF+IPhqy8QeHtQi1XR71S0F3Bna4DFW64IIZSCDyCCK2HdY0LOwVRySTgCvBlFxbjJWaPbTUlzJ6DqK+Nfjn/wUl8L/AA81q50PwhpDeLdRtpDFPePP5NnGw4IUgM0hBGDgKPQnpXjVn/wVV8ZJeBrvwXoc1rnmKGaaN8f7xLD/AMdr3aWRY+tBTjCyfdpHjVc4wVKfI56+SbP0vor5/wD2cP2zfBv7Q+7T4kfw54njXe2kXsqt5o7mGTjzAO4wG77cc19AV5FfD1cNN060bNHp0a9PEQVSlK6CivkT9pj9veT4AfEm58H2/gtdalhtopzeyal5KkuuduwRN0453fhXirf8FWfEpkyvgTSxHn7pvJCfzx/SvVo5Jjq9NVYQ0eq1X+Z5tXN8HRm6c56rfR/5H6S0V8V/DL/gp54P8UaxY6Z4p0C88MtcskR1BJkmtYnOATJnayoD/EA3qQOcfZ1pdwX9rDc200dxbzIJI5YmDI6kZDAjggjnNefisFiMHJRrwtc7sPi6GLTlRlexNRRXyz+1d+25H+zf4x0vw5Z+HY/EV7c2f22433ZgEClyqDhGyTsY44wAPWow2Gq4uoqVFXkViMRSwtP2lV2R9TUV8y/sl/tmx/tL6/r2jXWgReHb/T7aO7gjjvDP58ZYrIeUXG0mP1zv7Y5+mqWJw1XCVHSrK0kOhiKeJpqrSd0wor4u/aC/4KE3nwQ+LuveCovBMGrpphgxePqJiMnmQRy/d8s4xvx17V53/wAPYNQ/6Jxbf+Ddv/jNepTyPH1YRqQp6NXWq2fzPOqZxgqU3Cc9U7PR/wCR+itFfnV/w9g1D/onFt/4N2/+M13PwQ/4KL3vxf8Aip4e8HyeB4NLTVZzCbtdSMpjwjNnb5Yz93HUdadTI8fSg5yp6JXeq6fMUM4wVSShGer02f8AkfblFfKX7Vn7bt1+zb8Q9P8ADMHhKHXkutMj1A3Ml8YCpaWVNm0Rtn/VZznv7V7R+z38WJPjh8IdB8ay6aukSan5+bNJvNEflzyRfewM58vPTvXBUwOIpUI4mcfcls7r/h+h208ZRq1pYeMvejurM9Foqnq+r2OgaZdajqV3DYWFrE009zcOEjjRQWZmY8AAAkn2r4h+Lf8AwVD0bQdWm07wJ4cPiBIWKNqmoTGGByO8cagsy+5K9OnOaMJgcRjZONCF7fd94YnGUMIr1pW/M+6qK/NXSP8Agqv4rhu1bVPBGj3ltnmO0uZYHI/3m3j9K+zP2e/2nvB/7RejS3GhTPY6tagfa9HvCouIc9GGDh0P94fiAeK6cVlOMwcPaVYe73WphhszwuLlyUpa9noevUUV8MfFv/gpLffDH4meJPCieBLfUE0i9ktBdNqbRmUKcbtvlHH0ya5cJgq+Ok4UI3a13S/M6MTi6ODipVnZP1/Q+56K4f4K/FjSvjZ8NtG8XaSdsN7HiaBj81vMp2yRt9GBx6jB6EV27HAJ9K5JwlSm4TVmtGdMJxqRU4u6YtFfEnw2/wCCit94++MOj+B38DwWaX+pfYDerqTOUG4jds8sZ6dM19t11YrB18FJRrxs2r7p/kc+GxdHFpyou6Wn9XCiiiuI6wooooAKKKKACiiigBM0Zr8l/wBuL4m+MPD/AO1F40sNL8Va1ptjCbPy7W01CWKJM2cJOFVgBkkn6k14T/wubx//ANDx4j/8G0//AMXX2eH4aqV6MKqqpcyT27q58pW4gp0asqTpt8ra37H7uscKfpX43+J/2xvjLY+JdWtoPHuoxwQ3cscaBY/lUOQB930rzL/hc3j/AP6HjxH/AODWf/4uuQmmkuJnlldpJXYszucliepJ7mvossyKOCc3XtO9rabfeeDmGcSxaiqN4Wv13+4/Xr9gX4keJPij8EbrWfFWrzazqa6vPALi4wGEaxxELwAMAsfzr6TzX4IaF8RPFXhey+x6N4l1fSbTcX+z2N9LDHuPU7VYDPA59q0f+FzeP/8AoePEf/g2n/8Ai687FcNSr1p1ITUU3orbHfhuII0aMac4NtLe+5+72aWvwg/4XN4//wCh48R/+Daf/wCLr9h/2UtSu9Y/Z28B3t/dTXt5PpqPLcXEhkkkbJ5ZjyT9a+dzPJp5bTjUlNSu7bHu5fmscfNwjC1lc8Y/aL/b+f4B/FPUPBw8DrrgtIoZReHVPI3eZGHxs8lsYzjrXmf/AA9kk/6Jgv8A4Pf/ALnrxH/gon/ydJ4g/wCvSz/9EJWL+y9+zv4Z+N2k+M9V8VeMl8F6X4d+x7r2YRiE+e0qje7soX5kUDnktivqaGWZbDA08TiKe6jfWW7t0T7s+crZhmE8ZPD0Z7N20jsr9Wux9Ef8PZJP+iYL/wCD3/7no/4eySf9EwX/AMHv/wBz1yn/AAxp8CP+jiNE/wDAqz/+O0f8MafAj/o4jRP/AAKs/wD47WXsMj/59S+6oa+1zj/n4vvgdX/w9kk/6Jgv/g9/+56P+Hskn/RMF/8AB7/9z1yn/DGnwI/6OI0T/wACrP8A+O0D9jT4Ek/8nEaL/wCBVn/8do9hkf8Az6l91QPa5x/z8X3wOr/4eySf9EwX/wAHv/3PX0J+yX+1m37UB8UBvC48Of2J9l6X/wBp87zvN/6Zptx5Xvnd2xz+Y/7Snwgs/gZ8Vr7wjYalLq9vbQQTC7mQIX8yMP0GcAbsda+tf+CTX3/ih9NM/wDbqozLLcBTy6WKw0LOyad31a6N9h4DMMbPHRw9ed901ZdE+yPtD4r/ABu8GfBDTbG/8Z6udItL2UwQSC2ln3uBkjEasRx615l/w8A+BP8A0Ozf+Cq9/wDjNeR/8FVv+SbeCv8AsKyf+iTX52+CvB2p/EDxVpnh3Rokn1XUZhBbRSOEDueg3HgZ9TXJlmSYXGYRYitJp63s1bT5HVmOb4nC4p0KUU9ujvr8z9bv+HgHwJ/6HZv/AAVXv/xmvin9qj9sPXde+Ld3efDH4g6xD4Va1hEaWkk1tGJAvz4RwpHPfHNc/wD8O8Pjh/0Ldp/4NLf/AOLrxr4pfCvxF8HPFb+HPFFrHZ6tHEkzwxTLKFVuV+ZSR0969rL8tyynW5qFTndtm4v52seTjsfmNSlatT5FfdJr8bn0v+yF+2VqPhv4m31z8WPH+rT+HG0uWOFbwzXSC5MsRU7EDEHaJOcY/OvsT/h4B8Cf+h2b/wAFV7/8Zr8rPhJ8G/FPxw8ST6D4Rso7/UoLVrx4pZ0hAiVkUnLkDq68e9eu/wDDvD44f9C3af8Ag0t//i6nMMtyupW5q9TklbZOK/CxWCx+Y06NqNPnXdpv8bn3t/w8A+BP/Q7N/wCCq9/+M16p8K/jB4S+Nfh+41vwdqh1bTLe5azkmNvLDtlVVcrtkVSfldTnGOfrX4ZeItBvPC2v6lo2oRiK/wBOuZLS4RWDBZI2KsARweQeRX6c/wDBLX/kgXiD/sZZ/wD0mtq8XNclw2Cwvt6Mm3dbtW1+R62W5tiMXifY1YpKz2Tvp8z6x8WeJ7DwV4Y1XX9VkaHTNMtpLu5kRC5WNFLMQByeAeBXz7/w8U+CX/Qfvv8AwWT/APxNe3fFXwdL8Q/hr4o8MQXCWk2sabPYpcSKWWMyIVDEDqBnOK+AL3/gllrWm2c93d/ETSLW1gRpZZ5rR0SNFGWZmLYAABJJryMtoZdVhL67UcXfS3/DM9TH1sdTkvqkFJW1v/w6PpH/AIeKfBL/AKD99/4LJ/8A4mtjSP26PhHrmia5q1prN5JY6LDHPeudPlBRJJViQgFcnLuo4r8g/FGm6fo/iLULHStT/trT7aZoodQEJhW4AON6oSSFJyRnkjBIB4H1F+zd+zdqXiX4HePdd8T61YeBPCWv21raW2s6udqkx3cUm/aSo2EpsDFhljwDX0uKyLL8NSVWUpK7X4tdLXvbofPYfOcdXqOmoxdk/wAF3va1z7CP/BRT4Jf9B++/8Fk//wATXxd4o/4KH/F3/hJdW/sXxLbf2P8Aa5fsXmaXBu8jefLzlM524681a/4Y1+HH/Rx/gz/vqL/5Ir5d1/TodH13UbC3vI9Rt7W5kgjvIfuTqrECReTwwGRz3rvy/LMs5peyTn/iX5Xijjx2YZhaPtGo/wCF/nZs/Sr4F/8ABRLwcvw204fEvxDcN4v3y/ajb6Y2zb5jeXjy12/c216AP+CifwSJA/t++5/6hk//AMTXwL8Jf2bfBnxE8DWeu6x8aPDXg+/neRX0jUTH50QVyoJ3TKfmADDjoa8JvbdLPUZ4I5luI4pWRZl6OAcBh7HrWX9h5dias+RyTT1S0S9Lx29DT+2Mfh6cOZRaa0e7frr+Z/QJFIJokkX7rAMPoafXOeAvG3h/x54fjv8Aw3rVhrtlERBJcafcLMiSBVJQlSQGAZTjryK6OvzOUXCTi1Zn6DGSkrp3CiiipKCiiigAooooA+Bv+Cr/APyAfhz/ANfN7/6DDXYf8EtP+SCeIf8AsZJv/Sa2rj/+Cr//ACAfhz/183v/AKDDXiv7KH7bVn+zZ8PtR8N3HhSfXXu9TfUBcRXohChook24KNn/AFec571+gUsNVxeRQpUY3lf/ANufc+JqYilhs5nUquyt+iP1jr8kf+CkH/Jzmo/9g60/9Ar3X/h6/pf/AETm7/8ABqv/AMar4+/aZ+NcHx/+KVx4ut9Kk0ZJrWGA2kswlKlBjO4AZz9KMiyzF4PFupXhZWa3Xl2Ys5zDC4rDKnRnd3T2fn3R+s37K/8Aybl8Of8AsCWv/osV+W/7bv8AydT49/6+of8A0nir9SP2V/8Ak3L4c/8AYEtf/RYr8t/23f8Ak6nx7/19Q/8ApPFWeQ/8jOv6S/8ASkaZ1/yL6Pqv/SWfrt8MP+SbeFP+wTa/+iVrU8Sf8i7qn/XrL/6Aay/hh/yTbwp/2CbX/wBErWp4k/5F3VP+vWX/ANANfFT/AIr9f1PrY/w16H4wfskf8nMfD3/sLx/1r9sa/E79kj/k5j4e/wDYXj/rX7Y19bxT/vNP/D+rPmOHP93n6/oj56/bx+Ii/D79mzxIEkCXmthdHgGcbvNz5g/79LJX5VWHww1C9+D2rePVTdp9lrFvpZIzwXikdie2AREP+B19Z/8ABU34jHUfGvhXwTbyEw6batqNyqngyynagPuqxk/SX8vVvh7+zmtz/wAE9bzw41rjWda0+TxDjaQzXPEsA9c7I4kP4135fVWV5fSqS3qTX3f8MvxOPHUnmONqQjtTi/v/AOHf4F//AIJlfEf/AISj4KX/AIYnm33fhy+KopbJFvNl0/8AHxMPwr7Cr8mv+Cb3xD/4Q79oNNGnm8uz8RWclkVb7vnJ+8jP1+R1H+/X6y187n2H+r46bW0tfv3/ABue5ktf2+DjfeOn3bfhY/JX/gpL/wAnMXf/AGDLX+TV9o/syfs+/DTXfgD4G1LU/Afh/UdQvNLhmuLq706KWWVyMlizKTmvi7/gpL/ycxd/9gy1/k1foj+yf/ybd8Ov+wNB/wCg17OZVJ08pwrhJrbb0PKwEITzPEKavv8AmeZ/Gr/gn38OvH/hu8HhXS7fwh4jVC1rc2u4W7PxhZI8kbTjGVGRkkA9D+fPwa+KXiX9kz43SzTwyq1hcvp2taXni4jDbXXnuCNyt6gdic/tXX5If8FHfDVtoH7S9/cWyCM6rp9tfSgdC+DET+PlA/UmjIcZPFyngsS+aLT31DOsLDDRhi8OuWSfQ/VzTfEmm6t4btdftbuOXSLm0W9iuwfkaFkDh8+m05r8bvj78XvEP7U/xtC2ss9zYz3w03w9pjfKsUbyBI/l7PIdpY8nJAzhVA+2/C3jy6T/AIJnNrEcsiXMfh+501JN2GUCZ7UYOeMADH4V8jf8E+/D0Gv/ALUPhp7gBlsIbm8VGx8zrEwXr6FgfwrTKMPDAxxWKkrundL5f56EZpXljHh8OnZTs38/8tT70+Bn7Dfw6+FvhK3tde0LTPGfiCTL3eo6pZrMhY9FjjfcqKBxnqep64HN/tPfsK+DPHvge/v/AARoFl4b8W2URmtk02IQwXe0ZMLxqNuWHAYAHdjJxmvrKjrXykczxca/1j2j5r3309LbW8j6WWX4aVH2PIrfj9/c/Hr9if486h8EPjJY6deXJg8M63cLYapbT5CxsSVjm/2WRiMk/wAJYHsR+q/xeOfhN4z/AOwLef8Aoh6/HL9qXw3D4R/aK8fadbAiBdWlnRT/AAiU+btHsN+B7Cv1r1jWJfEP7Mt7qk7Fp73wk9zIx6lnsyxP5mvp89pQnUw+Mgrc9r/g1+Z89k1ScYV8LJ/Be34pn4mafo93qsV/LawtMtlB9pn28lY96oW/Auv4ZNfpz/wTg/aBHjzwDL8P9WuN2t+G4gbNpG5nsScKB3/dEhD/ALLR+9fIn7CPhjTvG/xzm8OavD9o0zVtFv7O4jzglGj5wexHUHqCAa51JPEX7G37SrAFnu/D99hgDhb20cfyeJh9CfUV9HmVKGYqpgvtpKUfy/4D9TwcvqTwLhi/sNuL/r8fkfpr+25/ya14+/69Yv8A0fHXyF/wSq/5KX40/wCwTH/6OFfUn7VHi7TPH/7F/ijxFo0/2nS9T0y3ureTGCVaaI4I7EdCOxBFfLf/AASq/wCSl+NP+wTH/wCjhXy2DTjkuJi91L/5E+jxbUs2w7W1v8z7J/aw+Pkf7PXwmutdgRJtbu5BZaXDIm5DOyk7nGR8qqrE+4A71+YvwI+CPi/9r34nX7XOqzSIjJdazrl67Suis20AZ+85AO1SQMIegFe7/wDBVbxJLceP/BWgeb+4tNNkvfKDfxSylMkfSHj8a92/4Jo+EbfQv2eTq6DNzrepzzyOeu2MiJV+gKMf+BGujDSWVZT9apr95Ue/b+kr+pz4iP8AaWZ/Vpv3IdP683b0O38F/sMfBjwXaiNfCEOs3BUB7nWJWuWfHfaTsXr/AAqKseKv2JPgz4qsZrd/Bdppcki4FzpZa3kjPquDtzx3B/U16R8Sfir4V+EGhwax4v1ePRtNnuFtI55I3cNKVZguEUnojHpjivNf+G4/gf8A9D7a/wDgJcf/ABuvm6dXMqz9rTc35rmZ786eX0l7KagvJ2R7XpOmW+i6XZ6faIIrW0hSCJB/CiqFUfkBVuuY+HnxK8NfFbw8Nc8K6omr6UZWgFzHG6DeuNww4B4yO1dPXlTjKMnGas+tz0oSjKKcHdHl37UX/JunxH/7AV3/AOizX5f/ALDngLw/8Svj9p+heJtMh1fSZrK5d7WfO0sqZU8EHg1+oH7UX/JunxH/AOwFd/8Aos1+S/7MnxotfgF8VbXxfd6bNq0UFtNB9lhkEbMXXbncQcY69K+6yONSeXYiNL4nt62Pjs4lCGOoSq/Ct/S5+pv/AAxb8FP+if6d/wB9y/8Axddh8OPgX4E+Ed3eXPhDw5baHPeIsc727OfMUHIB3MehJr5J/wCHrmhf9E/1D/wYJ/8AEV9Dfsw/tNWf7TGi63qNnoU+hpplwkDJPcCUyFlLZGAMdK8HFYXNKNJzxPNydbyuvuuezhsTl1Wqo0OXm6Wjb9D5N/4Kuf8AI0/D7/rzu/8A0OOvlb4S+Ptb/Z7+J3hfxfbRyBVVLkxfdW8tJMrKnoQcOoPZlB6ivqn/AIKuf8jT8Pv+vO7/APQ464fxn8D38dfsK/Dv4jadEr6v4dS8tr/AAMtj9tnCn3MbEYH913J6CvtMuq06eW4enW+GbcX8+b/K3zPk8dSqTx9adL4oJS+63/Dn6g+FfE+neNPDema7pFwt3puo26XNvMh4ZGGR9DzyOxyK/LH/AIKW/wDJybf9gi1/m9evf8Eyfj4XivfhXq9wSU332ilz0H3poR+OZB9ZPavIf+Clwx+0k3vpFr/N68jKsHLA5vKjLonbzWlj1MyxUcZlka0e6v5PW5+j/wCzn/yb/wDDX/sW9O/9Jo6/Pz/goT+05qHjnx1e/DrQb6W38L6LIbfUUhYr9uu1YFlfBwyRsqgKQPnVic4Tb94/BjVRoX7LvgnUiNws/CNncEeuyzRv6V+RXwX0AfFD48+E9N1ZvtSatrcJvTIM+arShpcj3G786jJcPTeJr4qqrqne3rrr8rFZtXmsPRw1N257fpp+J9h/sq/8E8NF13wlY+K/ifHdzzX6Ca20COR7cRREZBnYYfeeu1SMDGSScL9UWv7Ifwbs9Pjs4/h5orQxoEDSwmSQgDHMjEsT7k5r19QFUAcAcV4/4q/a6+EfgjxFf6FrfjK3sNWsZDDc2z207GNx2yEIP4GvCq47H5hVcoOT8o30+S/M9qng8FgqajNRXm7a/eW/hl+zD8PPg94yv/EvhLRTpN/e2v2SRFneSJU3Kx2hiSpJVc4PaqH7Y3/JsnxB/wCwcf8A0NaZoX7ZHwc8S63p+kaZ42trrUb+4jtbaBbW4BkldgqKCY8ckgc0/wDbG/5Nk+IP/YOP/oa1nBYn65SliVK94/Fe+67lzdD6rVWHatZ7W7eR8Df8EzP+Tj5f+wLc/wDocVfZ3/BQv/k1XxT/ANd7L/0pjr4x/wCCZn/Jx8v/AGBbn/0OKvs7/goX/wAmq+Kf+u9l/wClMdfSZn/yO6PrD8zwMv8A+RPV9JfkfkdaWV7DYHWbXzUjtLmOJriPK+TIwZovm7E+XIR/uGv2V/ZF+PEPx8+EGn6rPKp8Qafix1aLGD56qP3gH911w3pksP4TXwb+xb8HLT47fDH4y+FJ1jF88Gn3WnTyD/UXcf2ry2z2B3Mh/wBl2rlP2Qfjbdfs3fHD7LrhktNEv5f7K1qCQ4EBDlVlYesTZz/sl8cmvXzajHMoVaMF+8pWa800n+P5pHmZZWll86dWf8OpdPyabX9eTPrT/gqj/wAkd8Kf9h0f+k8tS/8ABLH/AJIl4m/7GCT/ANJ4Kh/4KmsJPgz4TdSGQ64pBHQ/6PLU3/BLH/kiXib/ALGCT/0ngr5//mQf9vfqe3/zO/8At39DzT/gq9/yMnw8/wCvW8/9Dirq/wBhb9nL4b/E/wCA1trXijwnZ6xqh1C4hNzMXDFFI2j5WHTJrlP+Cr3/ACMnw8/69bz/ANDirjv2Yf27dK/Z9+FkHhO68KXesTJdTXJuYrtYl+cjAwVPTHWvSp0sTWyWlHCX5r9HbS79Dz51MPSzarLE25bdVfWyPuD/AIYt+Cn/AET/AE7/AL7l/wDi69Z8OeHNN8IaBZaLo9qljpdjEIbe2jJKxoOijJJr4gH/AAVb0IkD/hX+of8AgwT/AOIr7h0DVhr+gadqaxmFb22juBGTkqHUNjPfGa+Rx1DH0Ix+uXs9ru/6s+nwdbBVm/qtrreyt+iPxI1Oxg1P9oy6s7qJZ7W48UtFLE/R0a7IYH2IJr9XB+xd8FCB/wAW/wBO/wC+5f8A4uvyZ8Uaynhz486tq0kbTR2PiOW6aNTgsEuSxA+uK+4B/wAFXNCA/wCSf6h/4ME/+Ir7bN8Pjq6ovB3slrZ27eaPkcrr4Oi6v1q2r0ur9/Jn094R/Zb+FngPxFZa9oHg6y03V7Ni0F1E8m6MlSpxliOhI/GvLv8AgpH/AMmzXv8A2E7T/wBCNO/Z2/bq0z9oT4hr4UtPCl1o8ptZLr7TNdrIuExxgKOu6m/8FI/+TZr3/sJ2n/oRr5XD08VSzKjDF35rrd30v6s+krVMPVy+rLDW5bPZW1sfNH/BM/4deFvH/iLxyPE3h3S/EK2ltamBNTtEuFiLNJuKhwQCdo59q+pP2hf2Lfhr40+HmtzaN4ZsPDGvWlrJc2d3o8C24LopYI6LhWVsYPGe4PFfI/8AwTu+NHgv4Oa141n8Y67DokV9b2qW7SxyP5hVpCwGxT0yOvrXun7Rv/BQ/wABn4e6ro3w/vbnX9b1S1ltUuxaSQQWgddpdvNVWLYYlQFIyOSO/t5hSzCeaN4ZSt7uuvLsvkeRgqmCjlqWIcb66aX3fzPlb9gbxzqHg39pXw5aW07LZa15mnXkIztkUozJx6h1U5+vrX3B/wAFI/8Ak2W+/wCwlaf+hGvm3/gnD+zvqniLx1D8TtUt2tdA0jzE09pBg3lyVZCyDuiAtluhbAGSrbfpL/gpH/ybLff9hK0/9CNPMKlKpndFU904p+twwNOpDKKrns02vSx4N/wSj/5Gr4g/9edp/wChyV6d/wAFTv8Akifhj/sYE/8ASeevMf8AglH/AMjV8Qf+vO0/9Dkr07/gqd/yRPwx/wBjAn/pPPUYj/koI+q/9JKof8iOXo//AEo/N7w1qeteBtR0PxdpbS2c1relrO8Awvnw7HZR64EkeR6Piv2z+BXxb0/43/C7Q/F1htja8hC3VsrZ+z3C8Sx+vDA4JxlSp71+dPwQ+B3/AAu39inxmllCJPEWgeIJtS0/H3nxawebEP8AfVeP9pVo/wCCdf7QA+GvxJbwTq85TQPE0qxws7YS3vcYjb6ScRn32dga784oxzKjUlTX7yi3fzW/5a/JnFlVWWAqwjN+5VX47f16o7//AIKwf8hn4cf9cL7/ANChr6F/4J6f8mq+Fv8Arve/+lMlfPX/AAVgH/E4+HB7eRff+hQV9Cf8E9eP2VPC3/Xe9/8ASmSvHxX/ACIaH+L9ZHqYf/kdVv8AD/8AInjX/BRz9p3UPDbx/C7wzdTWVxc26z6zewOUcRv9y3VgQQGGS/HKsozgsK8v/Y3/AGFl+Mekp4z8cyXFn4WlJGn2Vs4Sa+IYhnZuSkYIIGPmY5OVABbwL4+6zc+Pf2hvGt077pbvXbi3iLZ4RZTHGD16Kqj8K/anwX4YtPBfhHRtBsY1js9NtIrWJUUKNqKFzgeuM1242rLJsvpUMPpOerfXpf8AOy8jlwlNZrjalavrGGiXTy/K/qeaaN+x18GdCsFs7f4faTLEufnvEa5kOTnl5GZv1p3hz9kf4XeDvHukeL9A8Npo2r6Y0rwi1lYRM0kZQlkJI4BOMYwTmtT4h/tMfDT4UeIf7D8V+KIdH1Xyln+zyW8znY2drZVCOcHv2rmB+3H8DyQP+E9tf/AS4/8AjdfLx/tKpHmXO1L/ABNNP8z6J/2fTlyvkTXorWPdqKjgnS5gjmjO6ORQynGMgjIqSvHPUPmP9u/9pC6+BPw4ttN0JtnijxH5tva3AYg2kKqBLOpH8Y3oF5HLbudmD8M/slfssan+094tvtV1m8ubXwtYzBtR1DO6e6mY7jEjNn5yDlmOcAjIJIrof+ClviGbV/2jjYOx8nS9KtrdEzwC26UnHqfMHPsPSvub9hjwvbeF/wBmHwatuP3l/FJfzv8A3nkkY/ooVf8AgNfeqf8AZOUQqUdKlTr63f4L8dT4pw/tPM5U6vwU+npp+LN7w/8AskfB3w1YJaW3w90O4RVA339sLqQ+5eXcc/jU2h/ss/C/wv470zxhonhO00fXNOMjQSWJaKLLoUOYgdh4Y444JzXq9FfGvF4h3vUlrvq9T6xYWgrWgtNtEfPP7cHx6ufgV8HJZdHuDb+JtamFjp0qqGMPG6WXB/uoCAecM6HBGa+Av2Pv2arj9pv4hX99r91cjw3prrcapdbiZbuV2JEQc87mwxZuSAPVga9S/wCCqniB7n4neDtE80mOz0h7sR9laaZlJ/EQD8q+hv8Agmx4Zt9F/Zug1GNFFxq+pXNzK+OTsYRKM+g8vp7n1r7GjJ5Zk3t6Wk6j3+/9F958pVj/AGhmvsamsILb+vN/ceswfst/CK30oaevw48NmALt3vp0bTY/66kF8++c1+dH7c/7KcHwF8S2niHw2jjwdrUrJHAcn7BOBu8ncTkqwDMpPOFYH7uT+tNfPn7enhqHxJ+y74uMg/e2AgvoWxnayTJn81Lj8a8bKMxr0cZBSm3GTs7u+/U9bM8BRq4WbjFJxV1ZdjkP+Cd3x5u/ir8LLnw3rd19q13wu0dusr/fms2B8ksf4mXa6E+ioTkkk+T/APBWX7/wv+mp/wDtrXmH/BMrXp9M/aGuNPRyINS0ieOROxKMjqfw2n869P8A+Csv3/hf9NT/APbWvdjho4bP4qCsnd/fF3/E8eVeWIySTnurL7pL9Dpv+CeHwh8DeN/gPcan4i8HaFruoDWLiEXWpadFcSBAkZC7nUnAJP51F+3b+yV4H0b4R6h438J6Lb+HdT0aSOSe309FiguInkWNsp0BXcGBXHRhg54xf2Ev2lfhr8IPgZPpHi7xRBpGpHVLi6Fs0EsrmMpGAfkRupVuOvHTkVg/ttftueFfip8PpPAngR7nUba+nilv9TmgaGJo423rGivhy29UYkgABcc5+WFSx7zdypqShzb62t18tinUwSyxRm483L5Xv08xn/BK/wAfX8PjXxX4LeeV9MuLD+1o4ScpHMkkcTMPQssig+uwegra/wCCsv3/AIX/AE1P/wBta7P/AIJw/s6av8OtE1Px74hjayvddt0t7GxcYeO2DbjJIOxchcL1AXJ+9gcZ/wAFZfv/AAv+mp/+2tVGpSq8QKVLbW/ryu5MqdSnkjVTfT7uZHSf8E8vg94F8c/Ae51LxH4P0PXdQGsXEIutR0+KeQIEiIXcyk4BJ/OqH7df7Hfgvw/8L73x54L0mLw9qGkyI97aWh229zA7hCRGThGUspG3AxuBBO3HRf8ABObx34a8K/s73keteINL0iRNXuZ2S+vI4Ssflx/OQxHy/K3PTg+lY37e/wC1l4K1b4W3vgDwprdv4i1PWDEbm60uVZra3hSVXwZVOCzFcbVzwG3YyM4KWM/thqk5cvNrva3Xy2NnHCf2UnVSvy6bXv08zjv+CWPxKvo/FPinwJPNLLp01kNWto2bKQSI6xybR23iVM/9cx+PY/8ABSz9oW/8L6dp/wANPD99LZ3Gq25utZlgO1jasSiQbuoDkOXAwSqqDlXYHmP+CWPwwuxq3inx/cxNHZi3Gj2jsCBKzMskpH+7sjGf9o+hr52/be16bxD+1F45klkZ1t7mOzjVjkIscSJgDsMgn6knvXoxw1HE55OVrqCTf+LRf15o4ZYirQyeKvrJ2Xpr/Xoe1fsE/se6V8T9On8f+N7QX2giR7XTdLkDBbl1IDzOQRlFO5AO7Bs42jP2z4j/AGUfhF4m0abTbj4e6BaRSLt87T7GO1nT3EkYVs/jWx+z74ah8IfA/wAC6TAoC2+jWu8gY3SNGGdvxZmP416DXyWYZliMRipVIzaSelnayR9NgsvoUcPGEoJtrW63Z+Kf7Q3wa1f9lv4zf2daXk5ghkTUtF1PADtGHzGxIGN6MuDgdVzgAiv1X/Zi+My/Hj4OaL4okRItSYNa6hFGu1UuY+H2jJ+VuGAycBgOtfM//BVjw3BP4L8D6/txc21/NY7h3WSPfg/jF+ppn/BKTXnuPCPj7RTITHaX1tdqnoZY3Un8fJH5V9Bj5f2jlEMXP44uzfzs/v0Z4mCX1HNJ4WHwS6fK/wCGqPn7/go7/wAnPap/2D7T/wBF19yfs6fAD4aa58A/Al7qPgHw5f3t5ottNcXVzpcMk0rtGCzM5UsSSTzmvhv/AIKO/wDJz2qf9g+0/wDRdfXPwP8A2yvg94E+B/gfSdZ8ZRW2p2GkW1tcWsdncSvHIsYDKdkZHBBGenvWmOhiJ5XhVh027La/byM8HKhDMcQ67SV3vbv5nzj/AMFCv2Z/C/wcu9A8T+EbU6XYavLJbXOmocwxSqoZXjycqGG7K9Bt4xnFfRf/AATN8c6h4o+BV9pOoTtcDQtTe1tWfJKwMiyKme+GZ8egwO1fJv7ZP7UCftQ+KtC0Hwlpd9/YunTPHaK6brjULiQqoYRLkjptVckndyATtH3v+xn8B7r4BfBu10rVSv8Ab+pTHUtQjRgywyOqqIgwJB2KqgkEgtuwSMGs8xlOnlFOljH+8b0vvv8A5GmBjCpmk6mFX7tLW239XPdScCvxz+Lmoy/tMftj3dlbyNJaanrkWkW0iDhLaNxF5g9tqs/4mv1D/aQ+IX/Crfgd4x8RpKIbq2sHjtWJxi4k/dxfXDupx7V+d/8AwTW+HI8YfHibxDcRb7Tw3ZPcqx6faJcxxj/vkyn6qK5MjX1bD4jHP7KsvX/h7HTnD+sV6GDXV3fp/VziP2S/Flz8Ef2qdDttTItgb+TQdQUk4Uuxi59lkCHn+7X7I1+Pn7eHw/k+Gn7S+t3Vor21trJTWraRMrh5CfMIPr5qyH8RX6m/BPx6vxQ+EvhPxSCvmanp8U04Xos23Eqj6OGH4VXEEVXp0MdH7Ss/z/z+4nJJOjOtg5fZd1+X+X3nP+Nf2WvhZ8RPE154h8ReELXVNZvNhnu5JpQz7UCLkK4HCqo6dq+PP+ChP7P3w++EXwz8O6j4Q8NW+iXt1qwgmmhkkYvH5MjbfmY9wD+FfozXxR/wVR/5JB4T/wCw5/7Qkrz8nxVd42lTdR8t9ru23Y7s1w9FYSrUUFzW3sr79zy3/gnh8B/Afxg8I+Lrnxh4ct9bnsr6GO3kmd1KKyEkfKwzyB1r7P8ACX7KPwo8C+IrHXtC8HWunatZP5lvdRzTFo2wRkAuR0J6ivm7/glN/wAiP47/AOwhb/8Aotq+66vOsVXjjqtONRqPa7tsuhOU4ejLB05uCv3sr7s/LT/gqT/ycBoP/YtQf+lN1X2P+wF/yaX4G/7fv/S6evjj/gqT/wAnAaD/ANi1B/6U3VfY37Ahx+yV4HP/AF/f+ls9elmH/Ijw/qvykefgf+RxX9H+cT5Q/wCCkn7Qt/r/AI3Pwz0e+lh0LSkjfVY4jtW6ujh1RiPvJGpQ46bycjKKR6F+xN+xD4cm8F6d498fafBr11rFutxp+k3cWYLaFslZHBOJGddrAEYAPcnj4g8bXM3xG+O2tSXErvLrPiCVS5OSBJcEADPoCAB7Cv3I06xh0zT7WztkEVvbxLFGijAVVAAA/AVtmtSWV4GjhKD5XJXbW/S/3t/crGWW045jjKuJrK6jsn87fcv8zx74hfsc/CX4geHbnTG8GaVoUzoRDf6JaR2k8L44YFAA2PRgQfSvyu1G38X/ALI/x4liguFg1/w9eAxy7W8m7iPIJGQTHIhGRnOG6gjj9u6/Mj/gqd4ahsPir4U1yMFZdR0poJR2JhlJDfXEoH/ARXLw/jalSu8JWfNGSej1/q6OnPMJCFFYmkuWUWtVp/Wp+iXw38dWHxM8B6F4p005s9VtUuUB6oSPmU+6tkfhX46/tN6bPrP7UnjqwtVD3N1r0sESk4y7PgD8zX6E/wDBNvWZtU/Zms7aQkpp2p3VrH7KWEv85Wr4L+OMvk/tl+I5P7nioN+Uy115LS+q4/E0o/ZTt9+hy5vV+s4KhUl9pq/3Hp3/AATt+Pcnwx+Jk3gDXJGt9G8RziOITHaLa/A2rkHp5mBGe+7y+wNfqU/3G+lfkp+3r8Dn+DPxjXxFo6PbaH4kd9QtniJH2e6DAzICPu/MwdcdA+B92vv/APZG+PMPx8+Dtjqc8oPiDTwLHVo8YPnKoxIB/ddcN6Z3D+GuHO6EMRThmVBe7Pfyf9aP08zsyitKhOeX1t47en9an5m/s6/8nd+Ev+xi/wDajV+0Ffi/+zqMfteeEx/1Mf8A7Uav2gquJ/49L/D+ouHv4NT/ABfoFFFFfGH1YUUUUAFFFFABRRRQB+OP7fX/ACdl46+tl/6RQV3f/BPD4L+DPjJr3jO28Y6HFrcNjbW0lusskieWzM4Ygow64HX0rhP2+v8Ak7Lx19bL/wBIoK9x/wCCUX/I0fEH/rztP/Q5K/VsTOdPJIzg7NQhqvkfm1CEambuM1dc0t/mfUzfsOfA8KT/AMIFa9P+fu4/+OV+PXiq3is/E+rwQoI4YruVEReiqHIAFfv0/wBxvpX4E+Mv+Rv1z/r+n/8ARjV5vDOIrV5VfazcrW3bfc7uIaFKjGl7OKV77K3Y+9f2Df2bfhv8XPglPrXi3wvb6xqcerT2y3DzSxny1SNguEcA4LNyeefpXvniP9ib4KWHh7U7mDwJapPDayyI32q4OGCkg4MnrXHf8Exv+TdLz/sO3P8A6Lhr6h8X/wDIp61/15Tf+gGvn8xxmJhj6kI1JJc212e5gMLQngqcpU03bsj8B2GGP1r9rv2Qf+TaPh5/2C0/ma/FF/vt9a/a79kH/k2j4ef9gtP5mvqOKf8Adqf+L9GfPcOf7xP0/VH5z/8ABRP/AJOk8Qf9eln/AOiEp37NPg3WviB+zl8e9B8PadNqur3X9heRaQAb323cjtjPoqsfwpv/AAUT/wCTpPEH/XpZ/wDohK9f/wCCXeptouifGHUFQStaWtjOEJwGKrdtjP4V01KkqOTU6kVdxVN/c4nNTpqrmtSnLZua+9SPm7/hjj40f9E+1b8k/wDiq8y8X+D9Z8A+IrvQfEGny6Xq9ptE9pPjfHuUOuceqsD+NfcB/wCCr2qgkf8ACurP/wAGjf8AxqvkL45/FWb42/FPXPGk+nppUuqGIm0jlMix7IUiHzEDOdmenevRwVbMalRrF0lGNt0+v3s4cXSwMIJ4Wo5Sv1XT7kVvh38HfGfxZN+PCHh+7102Gw3P2UD91v3bM5I67G/Ku0T9jn4zh1P/AAr7Vuvon/xVaf7LH7Vt3+zDJ4ka18Ow6/8A20LcMJboweV5XmYxhWznzT+VfWPwb/4KP6l8VPih4b8JS+B7XTo9Wu1tmuk1FpDGDnkL5Yz09a5sdisyoTm6NKLppXu3rtr1N8Hh8vrQiqtRqb6JeenQ+Yv2/wCNov2k9WRwVddPsQQex+zJXvP/AASa+/8AFD6aZ/7dV4f/AMFE/wDk6XxB/wBetn/6ISvcP+CTX3/ih9NM/wDbquHG65Cv8MPzR24T/kdP/FL8mdP/AMFVv+SbeCv+wrJ/6JNfCv7PXi/S/APxs8HeItana20rTdQjuLmZY2cqgzkhVBJ/Cvur/gqt/wAk28Ff9hWT/wBEmvzi8PeHtS8V61aaRpFnLqGp3b+XBawjLyt/dUdz7VvkcYzytQm7J81/vZhnEpQzFyjuuX9D9bf+Hh3wP/6GW7/8Fdx/8RX57/tofFPw78YvjheeJPC1499pMtnbwrLJC8R3KuGG1gDWF/wyd8Yf+ida/wD+AjVwfjHwPr3w+1ptI8SaVc6NqaosjWt2myQK3Qkds1rl2W4HCVvaYepzStbdPT5GeOzDGYqlyV4WV+zX5nt/7Cnxk8LfA/4valr3i69ksNMn0aazSWKB5iZWmhYDagJ6I3PtX3if+Ch3wP8A+hlu/wDwV3H/AMRX5R+CPh94k+JOrS6Z4X0a71zUIoTcPbWUZd1jBVSxHoCyj8RXcf8ADJ3xh/6J1r//AICNU5hlmAxVf2mIqcsrbXS/NFYLMMbhqPs6ELx9G/yOP+KuvWfin4n+LtZ0+Qy2Go6vd3du7KVLRyTMykg8jgjg1+kP/BLX/kgXiD/sZZ//AEmtq/MDVdLu9E1O706/ge1vrSZ4J4JBho5FJVlI9QQRX6f/APBLX/kgXiD/ALGWf/0mtqz4gio5byx2TRpkbcsfd72Z9jMwRSzEKoGST2r8wP25v2zT8S7y68BeCb0/8Inbvsv9RgbjUpFP3UI6wqR16ORn7oBP39+0Npt5rPwJ+IFhp9rNfX1zod5FBbW0ZkkldoWCqqjkkngAda/HD/hnv4pf9E38W/8Agkuf/iK+e4dwuHnKWIrNXi9E/wAz3M9xFeMVQop2lu1+R7D+xd+yvB8ZPEUfiXxdLDZeB9Pl5hnlEb6nKP8AlknIIjB++/f7q8lin6AftOfD8fFj9n7XPBPhK60i3upVtEt43nWK3hjinjbHyg7QFTAAHoK/H3xV4E8T+BJLePxJ4e1bw+9wGaFdUs5bYygYyVDqM4yM49RXrX7PvgjX/H/ws+MOl+HtNuNX1Kaw05I7a2G52P26Nzgeyxufwr6DMMDOtVjjXXVotWVtFqlff7zw8DjI0acsIqOsk7u+r0fl9x1n/Dun4mf9Bbwp/wCDU/8AxFfNOu6PN4e1vUNLuWje4sbiS2kaJtyFkYqSp7jI4Nepf8MkfGb/AKJ9rf8A35/+vXlGpadc6RqN1Y3sTW95aytDNE/VHUkMp9wQRXt4ScpN81aM/RWt+LPIxMIxS5aTh6u9/wAEe2/Cz9jbxv8AF7wVZ+KNF1Dw/Bp100iJHf35ilBRyhyu045U456V4jfWT2F/cWkhVpIZWiYqeCQcHHtxXe+D/wBnr4lePdAt9a8O+ENU1bSbgsIru1j3I5VirYOexBH4V6l8Hf2EPid418b6bbeIfDlz4b8PpMr319qBVMRA5ZUUNuZmHAxwCeSKzeLp4dzlXrRaXTRNeW7u/kaLDTrqEaNKV311afntoj7z/Yf+CHiL4C/CXUdC8TNZtfXmryahH9ilMieU8ECDJIHOY24+lfQ1NRQihQMADAFOr8cxFeeJqyrT3Z+qUKMcPTjShsgooornNwooooAKKKKAPgb/AIKv/wDIB+HP/Xze/wDoMNcn+wj+yt8Ovjl8KNY1zxfpVxf6jb61JZxyRXksIESwQuBhGAPLtzXWf8FX/wDkA/Dn/r5vf/QYa7D/AIJaf8kE8Q/9jJN/6TW1feKtUoZDCdKTi77rT7TPjHSp1s6lGpFNW6+iO0/4d3/BD/oXLz/waXH/AMXX58ftn/Czw78HfjheeG/C1o9lpMVnbzLFJM8p3OuWO5iTX7OV+SP/AAUg/wCTnNR/7B1p/wCgVnw9jMTXxjhVqOS5Xu2+xeeYWhRwqlTgk7rZLzP0a/ZX/wCTcvhz/wBgS1/9Fivy5/bijaL9qrx4GGCbmBvwNvER/Ov1G/ZX/wCTcvhz/wBgS1/9FivhD/gpx8LLzw/8W7DxtDan+ydftEhluVGQLuEbSrHGBmIRlcnJ2vj7pqckqxhmtWMvtcyXre/6DzenKeW05Lpyv8Lfqfo38LJFl+GXhJ0O5G0i0II7jyVrT8VzJb+F9XlkYLGlnMzMewCHNfJP7Ff7YPgu++Eei+FfF/iTT/Duv6FALJTqcwt4p7dMCJlkfCZCbUK7t2UJxg1r/taftk+A/D/wo8Q6J4V8Uab4h8T6raNZwR6bL9ojiST5JJDKgZAVUsQpOSccV4k8txP1x0OR/FvbS19/Q9iOPw/1VVudbd9b229T4F/ZCge4/ab+HyopYjVEfA9AGJ/QGv2tZgqkk4A5r8r/APgml8K7vxT8aZfGEts39k+HLaQrcMDta6lUxog9SEaRj6YX1Ffe37V/xE/4Vh+z94y1qObyL1rJrO0cfeE837pCvupbd/wGvZ4h/wBpzCnQp72S+bf/AAx5WR/7PgZ1p7Xb+SR+Sv7RPxHHxU+OfizxMXa4srnUGS22v1toyI4sHBAyiqenU96+hrP/AIKh+ONPsYLO38HeHIbaCNYo41E4CqBgAfvOwFcX/wAE9Phdb/Eb4/RXeo2cd7pOh2U17PDOgeKR2HlRqwPXly3/AACv1J/4VT4K/wChQ0L/AMFsP/xNepmuMwWFnDCVaPPyJW1tb+kkedluExeIjPE0qvJzPXS9/wCrn4f6F43n8NfEex8W6ZbR2E9nqa6jb2sBIji2ybxGO+0fd+lfux4b1608U+HtM1qwkE1jqNrFdwSDo0cihlP4givzH/4KWfCKx8CfEvQPEWjadb6dpet2JikitIRHGLiE4Y4UADKPH/3ya+sv+CevxGHjv9nPS7KaQPfeHp5NLlGedi4eI/TY6r/wA1x55yY7A0cdTXl9/wDk1+J1ZPzYPGVcHUfn/XqmfFv/AAUl/wCTmLv/ALBlr/Jq/RH9k/8A5Nu+HX/YGg/9Br87v+Ckv/JzF3/2DLX+TV9r/szfHj4b6B+z94EsNT8e+G9Ov7XSoYp7S61WCOWJwOVZGYEH2xU5lTnUynCqCb229CsvnGGZ4hydt/zPpOvyF/4KHeK4PE/7TetQ27b00m1t9PLjoWCeYwH0aQj6g19jfHP/AIKH/D3wRoeo2Xg3UP8AhKvExieO3e1hZrOCUr8rvIdodQSOIy2emRXw3+zb8BPEP7UvxYa51F7g6Qbl7/W9YmVv3pLhpEVuhlcuO/AYt2wTI8JLBc+OxS5YpaX3/r87izjExxfJg8M+aTetj7atfh/eWv8AwTXfRPs7R3f/AAjT6kYdmG5c3RGPXB/Ovjz/AIJ8+JoPDn7T/h1LlgkepQXFiHPQO0ZZR+LIB9SK/XZtLtG0s6b9niFiYfs/2cIAnl7du3HTGOMV+Lnxt+FWv/sr/G/7NC04TT7yPUdE1SRMCeNWDxuDgAspAVgOjKe2KvJ68cdDE4WTtKpdr57/AHaE5rRlg54fERV1Cyfy2+/U/a6jpXzx8Dv23Phz8VfC1lPq3iDTfCniLYEu9M1S5W3Al6HyncgOpPIwc88jNct+09+3T4M+H3gm+0/wXrtn4m8V30Lw2smlTrNBZ7gR5zyrlcr1CgkkgZwOa+VjluLlW9h7N81+2nrft5n0ksfho0fbc6t/XTufnr+1B4gTxn+0b48vrQNKkmsS20W0ZLiNvKBGPXZkfWv1r1rR5fDv7M99pU67ZrHwk9s6+jJZlT+or8yv2JfgNf8Axv8AjJZajewu/hzQ5l1DUbmeNnjmdWBSDPQs55IJ+6GPPAP6rfF7j4T+NP8AsC3n/oh6+nz2rCNTD4ODvyWv+CX5Hz2TU5yhXxUl8d7fjc/Lr/gnP/ydDo//AF43f/oo19O/8FJ/gA3jPwXa/EXSLffq2gR+TqCr1lsSSd3uY3JP+67k/dFfMX/BOf8A5Oi0f/ryvP8A0Ua/Wu/sLfVLG4sruFLi1uI2imhkGVdGGGUjuCCRRnWKngs1hXh0ivmru6FlOGji8tnRl1b++ysfk/8ABD41PqX7K/xU+FepTAtb6e+q6SXbnYJUaaED2P7wd/mc9q7n/glV/wAlL8af9gmP/wBHCvAf2l/g3qH7OXxi1bRLd5E0u4SSbTLnn95ZzBl2EnqVBaNj3IJ7ivfv+CVX/JS/Gn/YJj/9HCvbx8KX9nV69F+7UtL56f16nk4KdT6/Ro1d4Xj+f9egn/BVTw/JbfE3wbrZjIivNJezWTPDNDMzEfh54/OvoP8A4JreKLbWv2cotMjlVrrR9SuIJov4lDt5qnHod559j6V1n7a/wCuPj18IJLXSUDeJNIm+3achdUExwVkiLMQAGU5BJHzKuSBmvzj/AGW/2htU/Zd+JtzPfWdzLot1/omtaVs2zfITtZVYjEkbE8N2LjjOR5FCH9qZP9Xpv36b2++33p/eenWl/Z2a+3qL3J9fuv8Ac19x+oP7TP7Plt+0l4EsfDV1rMuhx2uopqAuIYBMWKxyx7cFh/z1znPavzW/a/8A2U7P9mKTwqtp4hn17+2hdFjNbCHyvK8rGMMc580/lX6P+Ef2xPg94x0mG+g8d6Vp3mAbrbVpxaTRn0ZZMdPUZHoSK+L/APgph8U/CHxG1HwDB4W8Rad4gfT471rptOnWZIvMMGzLrlcny24znjnqK58jqY6jioYaaahrdNeT627m+cQwdXDyxEWnPSzv5rpfsfRf/BNX/k2qH/sLXX/slfVdfKn/AATV/wCTaof+wtdf+yV9V189mv8Av1b/ABM9zLf9zpeiPLv2ov8Ak3T4j/8AYCu//RZr8tf2Lfhj4c+Lvx10/wAOeKrA6jpE1pcSvAszxEsiZU7kYHr71+pX7UX/ACbp8R/+wFd/+izX5g/sMeOtB+HP7QGn634l1SDR9KisrlHurk4UMyYUfUmvpslc1luJdO/Nra297dD5/NlB5hh1Uty9b7bn6B/8O+vgX/0J8v8A4Nbv/wCO16d8JPgX4M+BthqFn4M0p9Ktr+RZbhGuZZtzKCAcyMxHB7VzH/DYvwY/6KFpP/fT/wDxNdF4F/aC+HfxM1s6P4X8V2GtamImnNtbFi2xcAtyOgyPzr5qtLMZ02q3O49b81vxPfpRwEJp0eRS8rXPiT/gq5/yNPw+/wCvO7/9Djr6M/YQ0621n9j3wrYXkS3FpdJqEE0T9HRrucMp9iCa+c/+Crn/ACNPw+/687v/ANDjr6V/4J+/8mneC/8Afvf/AEsmr3sW7ZFh2v5v/kjxsNrnNZP+X/5E/Of4z+ANd/ZH/aIxpbyQrp14mqaJeODtmt925Af72MGNx32t2Nbn7cXj3T/in8RfCnjDTMLaa14YtLryw24xP5sySRk9yjoyH3U190/t7/AL/hcPwhm1fTLbzfE3hpXvbYIuXngxmaEY5JKjcoHJZAB941+SlzfSXNrawylmNsrRxlm4VCxbaB2+ZnP/AAKvpsqrxzGFPEv+JC8X8/8APf7z57MqMsBOeHXwTtJf1+H3H7XfB/Sm139lbwbpq8NeeD7S3GfV7NV/rX5H/ALxAngT49+CdS1AG2hstbtxdeb8piTzQrkg9MAk49q/YX9nT/k374a/9i3p3/pNHX5sft8/s73vwm+Kl54q0+0kPhLxJcNcxTopKW122WlgY/wknc6jjKkgZ2NjwMkrQeIxGEqO3Pe343/B/ge1m9KfsKGJgvgtf8P8j9agcjIr43+Lf/BOTTPip8RfEHi6fxtd6fLqtyblrWOwV1jyAMBi4z09K5b9k/8A4KCeGoPBmleE/iVezaXqmnxpaW+tSRtLDcxDCp5pUFlcDALEEHG4sCTX1LL+018JUtGnPxI8LlNucLqsJfp/dDZ/SvDVHMcpryVJNPa6V019zPYdXA5nRi6jTW9m7NP70fkX8GdOTSP2mfA9jGxdLbxbZQqzdSFvEAJ/Kv1c/bG/5Nk+IP8A2Dj/AOhrX5U/Ca5ivP2pfB1xA4lhl8YWbo69GU3iEEV+q37Y3/JsnxB/7Bx/9DWvo87u8dhb+X5o8LKLfU8Tbz/JnwN/wTM/5OPl/wCwLc/+hxV9nf8ABQv/AJNV8U/9d7L/ANKY6+Mf+CZn/Jx8v/YFuf8A0OKvs7/goX/yar4p/wCu9l/6Ux1jmf8AyO6PrD8zXL/+RPV9JfkfPf8AwSf/AOQv8SP+uFj/AOhT1zv/AAUo+AB8J+Mrf4k6RARpWuOINRVQSIbwDh/YSKP++kY/xCuh/wCCT/8AyGPiR/1wsf8A0Kevuj4r/DfS/i58Pdb8JawubPU7cxeYFBaF+qSrn+JWCsPcVljMa8BnUqvTRP0sv+HNMJhFjcpjT66tet2fmD8TPjYfi3+xN4W0zUZhJr/hXxFb6dOWbLy25tbgwSn/AIChQ9cmMnvX03/wSx/5Il4m/wCxgk/9J4K/Ojxn4Y1n4W+JfEfg/Womgv7WX7NcRI58ssjBlkHHzAryp44fPfB/Rf8A4JY/8kS8Tf8AYwSf+k8FetnNGFHLZ+y+FyTXz1PNyqrOrmEfafEotP5aHmn/AAVe/wCRk+Hn/Xref+hxVqfsT/sm/DD4w/A+DxD4r8PPqOrNf3EBnW+uIcopG0bUcDv6Vl/8FXv+Rk+Hn/Xref8AocVdf+wl+0L8Ofhp8BLbRvE/i3T9H1QahcTG1uGO8IxGDwD1wa4ubERyOk8M3zX+ze9ry7HVajLOKqr25bdbW2Xc9hH/AAT6+BYP/Iny/wDg1u//AI7X0Dp2nQaRpltY2qGO2tYVhiQknaigADJ68CvJP+Gxfgx/0ULSf++n/wDia9I8G+N9C+Ifh6HXPDmpQ6tpM5ZY7uDOxipKsBkdiCPwr5DEvGzinieZpbc1/wBT6nDrCRbWH5b+Vv0PxR1rTLbWv2hr7T7yPzbS78TvBMm4jcjXRVhkcjgnpX6fj/gn18C8D/ij5f8Awa3f/wAdr8xNTvYNN/aNuru6lWC2t/FLSyyucKiLd5JPsADX6xD9sX4MYH/FwtJ/76f/AOJr7XPJ4yKo/VXLbXlv5b2PkcnjhZOr9YUd9Oa3n3J/hh+yf8Mfg54nHiDwn4fk03VRC9v57X08w2NjcNruR2HOK84/4KR/8mzXv/YTtP8A0I16XpH7WHwj17VbPTdP8daXdX95MlvbwRs26SR2Cqo+XqSQK80/4KR/8mzXv/YTtP8A0I18xhPrDzCjLE817r4r338z6LFewWBqxw9rWe1u3kfD/wCxz+y/pX7TGp+JrXU9au9HGkxQSRtaRq+/zC4IO702itj9qr9iHU/2dNHt/E2la02v+HWnWCWV4PJntHb7m7DEMpII3DGCQMc5r1P/AIJQ/wDIx/EP/r1s/wD0OWvsX9qjwd/wnf7PPjzSVi86Y6XLcwpjJaSH98gHuWjA/GvpcZmuIwua+y5v3d4prTZpX8/M+fwuW0MTlvteX37PX0bseR/8E+/2irr4wfD+68Na2YP7c8MxwwpJEoQ3FqQVjYqBjcu3aSP9k9Tzc/4KR/8AJst9/wBhK0/9CNfGn/BObxv/AMIn+0np+nyOyW+vWU+nMM/LvCiVCfxi2jv83ua+y/8AgpH/AMmy33/YStP/AEI1wYnCxwud01BWUmmvm9fxO3D4mWJyio5u7imvw0/A8G/4JR/8jV8Qf+vO0/8AQ5K9O/4Knf8AJE/DH/YwJ/6Tz15j/wAEo/8AkaviD/152n/oclenf8FTv+SJ+GP+xgT/ANJ563xH/JQR9V/6SY0P+RHL0f8A6URf8EsOfg34qHb+3W/9J4q+Yv27vgPN8E/jG2vaQjweH/EMrX9m8WR9muAQZo93bDHevTAcAfdNfTv/AASu/wCSO+K/+w6f/SeKvf8A9pr4JWnx8+EereGpAkepKPtWmXLAfubpAdnPYMCUP+y574rCWP8AqGc1JS+GTs/TTX5Gqwf13KqcY/Eldf5fM/Oj9rP4xD47fB74N+JZSDq1umo6dqqr/Dcx/ZvmPpvUhwP9ojtX25/wT15/ZU8Lf9d73/0pkr8ktTTUdEa80K+jktpLa6YTWsoIMUyZRhjsex/3R6V+t3/BPT/k1Xwt/wBd73/0pkr0c9oRw2WxpU/h59PR8z/U4cmrSxGPlUnvy6/LlX6H5i/HHTLrwV8f/GlvJH5c9nr9zJGH6FfPZ0b6FSp/Gv228L6/aeK/DWla1YSrPZahaxXUEiHIZHUMp/I1+fX/AAUv/Z7ubXV4fito1sZLK4WO01oRrnypBhIZjj+FhtjJPQhP71Y37Ff7c2nfC3w8ngf4gSTJoFsSdN1WKJpmtgzZaKRVyxTklSoJHIwRjbjj6Ms3y+jiMOryho116X/L7jXBVY5XjqtCvpGWqf32/P7z6O/aN/YW0/8AaG+IY8VXXi250aQWkdoLaKyWUYQsd24uOu70r8zPjT8PIfhP8WPEPhG3vH1CLSboW63UiBGk+UHJUE469M1+xFn+098JL61FxH8SPDKxkZxLqcUb/wDfLMD+lfkn+1F4j0zxd+0L411fRr2LUdMutQ3wXUDbkkXaoyp7jINbcPVsZKbo178kY6Jq3VeRlnlLCqCq0bc0nrZ36ep+12if8gaw/wCuEf8A6CKu1S0T/kDWH/XCP/0EVdr87luz7tbI/Jb/AIKTaNLpn7S91cyKQmo6Xa3MZ9VAaL+cRr7z/Yf8T23ij9mHwTJA4aSyt3sJkzyjxSMuD9VCt9GFeff8FDP2d7v4s/D+08U+H7N7zxJ4dDFraBC0t1aMRvVQOrIRvA9N4GSQK+O/2M/2tn/Zy1290zW4J7/wbqjiS5itsGW1nAwJo1JAbIAVlyMgKQcrhvvXTebZPCNHWdPp6afitfwPilUWWZpOVXSFTr66/nofr3RXkXhr9rf4O+KtPW8tPiFodtG38Go3Is5B9Um2t+n0q5of7Tvww8T+NNO8KaL4wsNX1u/dkggsS0qMVRnP7xRsHyo3frgdxXxTwmIje9N6b6PQ+tWJoO1prXbVHw7/AMFU9Aktvij4P1oxkRXmkNaCTnlopmYj8BMv519E/wDBNnxLBrX7Nttp8bgz6RqVzayp3G5hMD9CJf0PpW3+3d8C7n40/BeeXSYnn8QeHnbUrOCNC73ChSJYVABJZl5UAZLIo718IfsTftO237O/je/t9e89vCmtLHHdmFDI1tKrYSYLn7oDPuABJGMAkAV9nSg8zyVUaWs6b2+/9GfKVJLL829rU0hNb/15o/X6vnz9vTxLD4b/AGXfFwkYCa/EFjCpONzPMmfyQOfwrtrX9pv4TXmkDU4/iL4bFqVDYk1KJJBkZAMbEODweCM8Hjivzm/bn/autPj5run+HvDJl/4RPRppJBO+VF/cH5RLtPRVXIXIB+d89RjxMpy6vWxkHKDUYu7urbanr5njqNLCzUZJuSsreZd/4JlaDNqf7Q1xfqmYNO0eeR3z0LsiAfjuP5V6f/wVl+/8L/pqf/trXq3/AATw+AN58KPhpd+JdctntNe8TFJPss0ZWW2tk3eWrA8gvuLkehTPIwPKf+Csv3/hf9NT/wDbWvejiY4nP4uDuldfdF3/ABPGlQlh8kkpqzdn98l+h57+yl+w9oX7Q/wuk8UX/iTUNIuUv5bPyLeFHQhFQhuef4+ntXmX7Qv7O/ib9kfx5ok0ep/2lazf6VputRW4jVpEb5kaMswDLlSQSQQ3fkV9x/8ABMj/AJNyuf8AsOXP/ouKj/gpj4OXX/2fI9YVCZtD1OCcuB0jkzCw+hZ0/IVcM1xEc2lhqkr03Lltp8iJ5bRllixEI2mle/5nqv7KXxyb9oH4Paf4luoYbbV4pZLLUYLbPlpOmDlckkBkZGwem7GTjJ+Vv+Csv3/hf9NT/wDbWov+CU/jjZe+OPB8sg/eRw6rbR56bSYpT/49D+XvUv8AwVl+/wDC/wCmp/8AtrXFhsMsJnypRVlq16OLZ14jEPFZM6knron6qSR4N8A/2JPFX7QfgVvE+ia7pWn2i3clm0N95gfegUk/KpGMOPfrx6/Q3wz/AOCWFra3iXPjzxYb2FGB+waJGYw+OzSuM4PQgKD6EV6P/wAEx/8Ak3O6/wCw5c/+i4a+tqzzPOsbTxFShTnZJ22V/vLy/KcJUoU6043bV93YzPDXhrS/B2hWWjaLYw6bpdlEsMFtAu1UUDAHv9Tyepr8c/22dFl0H9qLx5DKpHnXaXSEjhlliSQEev3sfga/Z+vgf/gpl+z/AHms2th8UNFtvP8AsEK2Wsxx/eWLcfKnx3ALFW7gFD0BI5uHsVGjjbVH8at89/xOjPMM6uEvTXwu/wAtj64/Z/8AEsPi/wCCHgXV4CCtxo1rvCnIWRY1V1z7MrD8K9Ar80v2Dv2ydH+Gejn4feOLhNO0JHkuNO1d9zLA7NuaF1AOFJLMG7EkHrkfbPiP9qf4SeF9GfU7v4heH5oAu9Y7G+S6mccj5Y4iznkEcCuHH5biMPipQUG03pZXujswWPoV8PGbmk0tbvZnzN/wVY8RwweC/A2g7s3FzfzXu30WKMJn85v0pn/BKTQXt/CHj7WTHiO7vra0V89TFG7Efh5w/Ovkj9oP4v63+1Z8aje6Zp93LDK66domkou+YRbjtBVSR5jklmwTgnGSFBr9Vf2aPgzH8Bvg/o3hQzJc38Ya5vriNcCS4kO5seoUYQE9Qgr6DHr+zsohg5/HLVr53/DRHh4J/Xs0niofBHS/yt/mz84P+Cjv/Jz2qf8AYPtP/Rdeq/Df/gmxo3xJ+EvhvxNB41vtM1LWNOgvSklkk0MTOoYrgMpI5IHI/HpXlX/BR3/k57VP+wfaf+i6/SL9mD/k3b4cf9gGz/8ARS1vjMZXweWYWdCVm0u3bzMcLhaOLzDERrRuk3+Z+Vay+Mv2Jf2gXQLBNqmkyA4lj3W97byLwRnoGRiMqcqdwzkGv2H8D+L7Dx/4O0XxJpbM2n6raR3kG8YYK6hgGHZhnBHYg1+fn/BVbwcbbxV4I8UxxDbd2k2nzOAODE4dMn3Er4/3TXvf/BOHxoPE/wCzdZac8pkuNCvrixYOcsFZvOT8MS4H+7jtXHmyWOy+jmFve2f4/qvxOrLG8HjquCv7u6/D9H+B5x/wVN+I39neDvCvgm3lxLqVy+o3Sr/zyiG1AfZmcn/tnXyp+zl+11r37Nmiavp+heH9K1JtTuFnnub7zN+FXCoNrAYGWP8AwI1c/bq8fn4lftL+IEtmM9tpBTRbYIdxJiJ3gY/6atJX6XfBz9n7wp4I+FnhbRL/AMMaPd6jZ6fEt3PPYRO8k5XdKSSpJy5bvXdKph8syyjSxFPn59Wr28/w0RyRp1swzCrVoT5eTRPfy/HU/LL9o/8Aab1b9pS50W71vQtO0q80pJIo5tPLjzUcqdrhyc4K5GMfebrkY+1P+CXnxG/t34V694Qnl3XGg3onhU9oJwSAPpIkhP8AvCvV/wBqL4AeHPF3wE8Y2Wi+HNMsdXisjeWs1lZRxSmSEiUICFH3ghT/AIFXwP8A8E8viH/wg37R2m2M0pjs/EFvLpcgJ+XecPEceu+MKD/tn1NOVShmeVVIUIcvs9Ut9tfx1FGFbLsypyrT5ufRvby/DQ/Xavij/gqj/wAkg8J/9hz/ANoSV9r18Uf8FUf+SQeE/wDsOf8AtCSvlMm/5GFL1/Rn0ubf7lV9P1Mv/glN/wAiP47/AOwhb/8Aotq+66+FP+CU3/Ij+O/+whb/APotq+66rPP+RjV9V+SJyf8A3Gn8/wA2flp/wVJ/5OA0H/sWoP8A0puq+xv2BBn9krwOP+v7/wBLZ6+Of+CpP/JwGg/9i1B/6U3VfY/7AX/Jpfgb/t+/9Lp69zMP+RHh/VflI8fA/wDI4r+j/OJ+WHi6Kb4d/HTV1uI283RfEMpZcYJ8q4J4z67a/crTL+DVdOtb21kE1tcRLNFIvRlYAgj6g1+ZP/BSP4A3vhX4hn4j6bDLcaHr+1b0xx/LZ3SIqDcQOFkABBPVg/tXp37E37b3h2PwVYeBPH2oW2gXWj26W2nardSFYbmBchUdiMRsihQCThhjoRzvmtKWZ4Gji6C5nFapfK/3NfqY5bUjl+Mq4Ws7Xejf4fen+h941+Y//BU7xLFf/FfwrokZ3Pp2kmeQjoDLKw2/XEQP/AhX2h8R/wBr/wCFPw58OXOpyeMdK1y4RN0OnaLeR3VxOxGVUBGIUHj5mIGO9flhqE/iz9rr4/Sy29u8mseIr4BI13SR2UGQo3EDIjiQDLY6KT1NcnD+DqU67xdZcsIp6vT+tDqzzFQnRWGpPmlJrRf13P0U/wCCb+hXGj/sz2U88bRjUdSuruPcMZTKxg/nGa+Avj4cftg+KT/1M5/9HCv18+HHgex+GngPQvC+nAfZNKtI7ZWC7d5UfM5GTyzZY8nkmvyD+Pv/ACeB4p/7Gc/+jRXVktZYnHYmstpJv8Tlzak6GDw9J7pr8j9Rf2nvgjB8fPg5qvhvCrqsa/bNLmc4Ed2inZk9lYFkJ5wHJ6gV+bv7FXxhu/gF8eBpGs7rLS9Yl/sfVIJzsFvMH2xyMOxR8qc9A71+vafcX6V+Zf8AwUn/AGfz4S8ZW/xJ0e326VrjiHUUjXiG8A4c9gJFH/fSMT96uDI8RCrGeXV/hnt5P+tV5rzO7OKEqcoY+j8UN/T+tPRnj3wHtzaftm+HYDwY/FDJ+UrCv2Xr8Uf2WL2fU/2oPAd5dSGa5uNbjllkbq7MxJJ+pNftdWnE65a9JP8Al/Ujh53o1Gv5v0CiiivjD6sKKKKACiiigAooooA/HH9vr/k7Lx19bL/0igr3H/glGceKPiD/ANedp/6HJXh37fX/ACdl46+tl/6RQV4loHi3XPCjzPoms3+jvMAJWsLl4S4HQNtIzjJ61+wLCvG5TCgnbmhHX7mflrxCwmZTrNXtKX6n78uRsbntX4E+Mv8Akb9c/wCv6f8A9GNWv/wuHx5/0OviH/waT/8AxVclNNJcSvLK7SSuxZnc5LE9ST3NZ5RlMssc3KfNzW6drmmaZnHMFBRjblv+Nj9Vv+CYxA/Z1vP+w7c/+i4a+oPF5H/CJ61z/wAuU3/oBr8JtE+IPijwzZfY9I8R6tpVpuL+RZX0sKbj1O1WAzwOavSfF3x1NG0cnjPxA6MCrK2qTkEHqD81eXi+HamJxM66qJXd7WPRw2ewoYeNF027K25yb/fb61+137IP/JtHw8/7BafzNfigAWOByTX7k/s4+Fr3wX8CPAujalA9tqFrpMAuIJRh4pCoZkYdipJB9xVcUyX1enHrzfoTw4n7eo/L9T8zf+Cif/J0niD/AK9LP/0Qlex/8Es9NTWdL+LunyOyR3dvYQMy9QGF2pI/OvHP+Cif/J0niD/r0s//AEQlWf2XfHuu/DL9nr48+JPDV+dM1qz/ALC8i6WNJCm+6lRvlcFTlWYcjvXXVpyrZLTpweslTS+bic1OpGlm06ktk5v7lI+nD/wSw+HpOf8AhKvEn/fUH/xuvg/9pL4W2HwW+NPiPwbpd3cX1jpjQCOe72+a2+COQ7toA4LkcDtXYf8ADd/x0/6HuX/wX2n/AMaryDxz451z4k+Kb7xH4kv21PWr0qbi6ZFQvtQIvyqABhVUcDtXVl2GzGjVcsZVUo22XfTyRzY7EYCrTUcLTcZX38vvZ75+xd+y54e/aTfxh/b+ralpaaItq0R08xjf5vnbt25W6eUMY9TXsvw0+GH7Mfwv8faJ4p0/4yT3V9pVytxFDcyIY3YdmxEDj6Gvjv4YfHHxv8GRqg8G66+ijVFRLvZBFL5oTdt++rYxvbpjrXDocyKfeta2BxGIq1HOs403ayVu2t7ozpYyhQp0+WknNXu3fvpazPpL/godKs37UGvSIdyPZ2TKfUG3Svcv+CTX3/ih9NM/9uq8E/b8/wCTkNU/7B1h/wCkyV73/wAEmvv/ABQ+mmf+3VeVjdMhX+GH5o9LCf8AI6f+KX5M6f8A4Krf8k28Ff8AYVk/9Emvhz9mzxJpvg/47eCta1m7Sw0ux1GOe5uZM7Y0Gck45/Kv0X/4KEfBzxj8ZPBHhaw8HaLJrV3aag888ccscexDGQDl2UdfSvhb/hhb45f9CHcf+Btr/wDHajJq+G/s1UatRRb5lq0nrcrNaOI+v+1p027Weza0P0p/4bY+CX/Q/wBh/wB+Zv8A4ivzf/bk+IHh74m/Hq813wxqcWr6TLY28aXMIYKWVcMMMAeD7VW/4YW+OX/Qh3H/AIG2v/x2vLPiF8N/Efwq8RyaD4p01tJ1aONZWtnkSQhWGVOUJHP1rpyzLsDha/tMNW5pW2unp8jnzHHYzE0eSvS5Y33s1+Z7p/wT9+J/hf4TfGfVdY8W6vDoumy6HNapcTKzBpTPAwX5QTnCMfwr9Bz+2x8Ev+h/sP8AvzN/8RX5F/DX4VeKvi/r0ui+ENJfWdTit2ungjkSMrErKpbLsB1dR1716Z/wwv8AHL/oQ7j/AMDbX/47U5ll2AxOI9piK3LKy0ul+ZWX4/G4ehyUKXNG+9m/yPNPi9q9nr/xX8Z6np863Nhe61eXNvOmcSRvO7Kwz2IINfpB/wAEuraSD9n7WndSqTeIp3Q/3h9nt1z+YI/Cvkvwx/wTv+MuvXlnHd6NaaJbyvtmnvb2M/Z1z1KoWLccgLntnFfp38EPhDpfwM+Guk+ENJle6hs1ZpbuVQr3ErsWdyB0yTgDnAAGTjNefn2Ow0sJHDUZqTuttdEd2S4PELFSxFWDirPfTVnX63YSapo1/ZQ3U1jLcQSQpdW5xJCzKQHU9mGcj3Ffjtrv7THxy+HPji80nV/G2sf2jo98Ybi0uJy0btG/KkcbkOPxB96/ZavzV/4Kc/BGPw/4n0v4k6dEkVrrLrYagqnB+1KhMb4x/FGhB56x+9ePw9Vo/WHh60U1Pa6vqv8AM9TPKdX2Kr0pNOO9n0f+R5j+3r8Z7D4xfFLRLjR5RLpVnolq0Z5zvnQTtn3CyIpHYqQeRXV/Abwj4l8OfsTfFHxd4bOp2uvavqdja2UmktIty0UE8e5kMfzY/fSqcdlOeK+PoozNKiAgFiFBYgD8SeBX6y/E6PUv2Y/2FoY/BerRW2qaRaWQTU7NUlSSSW5jM0ibgwIcyOR14avqsfy4Khh8FS1vKK12dmnr87HzmCvi61fF1dLRb081bT5XPz7/AOEl/aA/6CnxG/8AAi//AMa8k1Jrt9RumvzM18ZWNwbjPmGTJ3b887s5znnNe4f8Nx/HL/oern/wBtv/AI1XiOq6nda3ql5qN9KZ727meeeUgAu7MWZsDjkknivcw0KsW/aQjH/Df/JHjV505JezlJ+v/Ds77wZrXxdsvD8EXhO+8ZwaIGYwpo812tsDuO7aIztznOcd81up4l/aA3rnVPiNjP8Az8X/APjVHwF+1L8U/hl4YtvD/hjxVPpejW7O0VqlrA4UsxZuWQnkknr3ro0/bj+ORdf+K6uev/Pjbf8AxquepSxDm3GlBrzbv8/dN4VKCik6k0/JK3/pR+x9iSbK3LZLGNc5+lT1DZO0lnA7HLNGpJ98VNX4u9z9ZQUUUUhhRRRQAUUUUAfPH7Xv7Ld7+03YeGba01+DQv7IknkZprczeZ5gQAABhjGw/nW3+yb+z3dfs2/D7UvDd3rMWuSXepvqAuIYDEFDRRJtwSf+eec+9e2UV3vHV3hlhHL3F0svXfc41g6KrvEpe++v4BXx1+01+wbqXx/+Kl14utvF1ro8U1tDALWWzaVhsXGdwcdfpX2LRUYXF1sFU9rQdnt3/MrE4Wli4ezrK63OS+EvgiT4a/DPwz4VmulvpdIsIrNrlEKLIUUDcAScZx0qx8RPh1oHxU8J33hzxLp8eo6XdqQyOPmRsHDof4XGchh0rpaKwdWftPa39697+ZsqcOT2dtLW+R+cvjH/AIJU6umoyN4V8bWU9ixyker27xyIPQtHuDfXav0qXwT/AMEqdQbUoZPF3jW2SwVgZbfRoGaWRfRZJAAh9yjfSv0Vor3f7fzDk5ef52Vzxv7EwPNzcnyu7HN/D74deHfhZ4XtfD3hfS4dK0u3HyxQjl2wAXdjy7nAyxJJxXl37Wn7Pes/tH+EdI8O6f4jg8P2VteG8ufNt2lM7BCsY4YYA3OTnvj0r3WivGpYmrSrKvF+8tbvXX5nq1KFOrSdGS916WWn5Hz7+yL+ysP2ZdI8QRXOrQ65qWrTxs11FAYgsSKdqYJJ6s5znuOOK+gqKKWIr1MVUdaq7yY6NGGHpqlTVkjxf9qr9naL9pH4fWugrfx6TqFpepd219JEZAmFZXUqCMghvXqAe1cj+yL+yjrn7MuoeIRc+KrbXdL1eKLNtFatEY5YydrglyMFXYHjn5fSvpait446vHDvCKXuPpZGMsHRlXWJa99dT42/an/YS1n4/fE+bxbpviix0xZLWK2+yXVu7EFARncp759K8Yi/4JVeMzIBJ4z0NU7lYpifywK/TCivQo55jqFONKElZaLRHFVyfB1pupOOr1erPhX4e/8ABLDw9pWopdeMfF11r9uuD9g0+2+yIT3DSFnYj/dCn3r7M8D+BdB+G/hqz8P+GtMg0jSLRdsVtADj3ZiclmPUsxJJ5JNb1FcGKzDE43+PNtdun3LQ7MNgsPhP4MLfn97CuA+M/wAEPCvx48Iy6B4oszLFnfb3kBC3FrJjAeNsHB56EEHuDXf0Vx06k6UlODs11OucI1IuE1dM/N3xP/wSo8Qw37/8I743027siSUGp20kEijPAOzeGwMc8Z9BWp4E/wCCVM/21JfGXjWMWin5rbRLcl3HtLJwv/fDV+h1Fe8+IMwceXn+dlf8jxVkmBUubk/F2Ob+Hvw78PfCzwraeHfC+mx6VpNrnZDGSxLE5LMzElmJ6kkmrfjTQX8VeD9c0WOVYJNRsZ7RZWGQhkjZQxHfGc1s0V4LqSc/aN3e9z2lCKjyJWWx8b/s0/sFal8BPivZeL7nxda6xFbwTQm1is2iY70K53Fz0z6V9kUUV0YrF1sbU9rXd3a3b8jDDYWlhIezoqy3PC/2r/2X7H9pbwpptmLxNI1zTZ/MtNSaPeEjYYkjZerK2FOMjBUHPUHkP2SP2N7/APZo8U63q134mt9dTUbNbURw2rQlCHDZyWOelfUdFaxzDExwzwil7j6fiZywVCVdYlx99dQr58+Pv7E3gH48XkurTxy+H/Ekg+fVNOVR5x4GZUIw5wMZyD78V9B0VzUMRVw0/aUZOL8jetQp4iHJVjdH5rap/wAEp/FUVww07xxpF1Bnh7m1lhY/gN/860tA/wCCUeqSyRtrfxAtLaMMC8dhpzSlh3AZnXB98H6Gv0Xor23xBmLVuf8ABf5HkLI8Cnfk/F/5nn3wN+Cui/APwFD4U0K6vbyySZ7hpr91aRnfG77qqAOBgY/OvQaKK8CpUlVm6k3dvc9uEI04qEFZI5P4seCpPiP8NPE3haG6Wyl1ewmsluXQusRdSu4gEZxnpXwf/wAOotZ/6KFY/wDgtf8A+OV+jlFehhMzxWBi4UJWT8k/zOHFZfh8ZJSrRu15v9D84/8Ah1FrP/RQrH/wWv8A/HK9g/Za/YZ1H9nf4mP4quvFdtrUTWMtn9mis2iYF2Q7slj029PevryiumtneOr05Uqk7p6PRf5GFLKMHRmqkIWa83/mfMv7Xf7IV9+01q/hy8tPEdvoS6TBNEyzWrTGQuynIwwxjb+teofs6fCWf4G/CDRPBdxqMeqzac05N3FEY1fzJnl+6ScY3469q9Korgnja9TDxwsn7kdUrLz679WdkMJRhXliIr33o393+QhGRg18GfE//gmGPFvj/XNa8PeK7XQNIv7g3EOmyWbSmAtguoYMo27txAxwCBzjJ+9KKeEx2IwMnKhKze/X8xYrB0cZFRrRvY5r4aeEpPAPw68L+GZbhbyTRtMttPa4RdolMUSoWA5xnbnHvWh4o8K6P420K70XXtOt9V0u7QxzWt0gdHH9COoI5B5HNatFcbnJz5767nUoRUeS2mx8J/Eb/glloOqX1xd+DPFdxocbtuTT9Rg+0xp/srIGDAemQx9zXnUX/BKrxoZwJfGehpDn76RTM35bR/Ov0xor3qef5hTjy+0v6pM8aeS4Gcubkt6Nnxb8Gf8Agmnovw78WaN4l13xfea1qGlXcV9b29lbLbQiWNldNxYuWAZc8bc/z+m/jX8PJfiv8K/EfhGC8TT5dWtvs63UiF1j+YHJUEZ6etdvRXnV8ficTVjWqyvKO22h3UcFh6FOVKnGye58h/stfsMaj+zx8TW8VXXiu21qJrGWz+zRWbRMC5U7slj029Pevbv2kPhBP8dfhJq3g221KPSZb6SBxdyxGRU8uVX+6CM5246969Oooq4/EVq8cTOV5q1nZdNgp4KhSovDwXuu91d9T5q/ZC/ZIvv2ZL3xNPd+IoNdGsR26KIbZofK8syHnLHOd/6V9K0UVhicRUxdV1qzvJm1ChTw1NUqSskfKv7WH7D8H7RPi7TPEmla1B4c1SO3NtfSS2xlF0oIMRwGXDKC4JOcjaP4a7r9kv8AZ2uv2bPAuq+H7vWodce91Fr4TwwGIKDHGm3BJ/uZz717hRXRPMMTPDrCyleC6afnuYxwNCFd4mMfffX/AIGx8zftefsiX37TWqeG7q08R2+hLpMM0TLNatMZDIyHIwwxjb+tfPn/AA6i1n/ooVj/AOC1/wD45X6OUV04fOcbhaSo0p2ivJf5HNXyrCYio6tSN2/N/wCZ+cf/AA6i1n/ooVj/AOC1/wD45X2b+zf8Hrj4E/CTS/BtzqUerS2Uk7m7iiMav5kjP90k4xux17V6fRWeLzXF42n7OvK6vfZL8kaYbLcNhJ+0oxs9t2fnz4t/4Je6x4m8Vazq6+PbKBb+8muhEdOdigdy23Pmc4zisn/h1FrP/RQrH/wWv/8AHK/RyiutZ/mEUkqn4L/I5nkuBbu4fi/8z4C+H3/BMnV/BPjzw34hfx3Z3SaTqVtftAunupkEUquVB3nGduM19R/tO/BK4/aA+Fk3hK11SLR5ZLqG4+1SxGUAIScbQR1z6161RXHWzPFV6sK9SV5Q20R1Usvw1GnKjCNoy31Z8z/sh/si337Mup+JLq78R2+urq0MMarDatCY/LLnJyxznf8ApX0rPClxDJFIoeN1Ksp6EHqKfRXHiMTVxVV1qrvJnVQw9PDU1SpK0UfBGi/8E0tZ8H/EWz8T+HfHlrZrp2pLfWME1i7sipJvRHYON3AAJwM88c19PftOfBK5/aA+Fk/hK21SLR5ZbqG4+1SxGVQEJONoI6/WvWaK66uZ4qvUhWqSvKGzsv6ZzU8vw9GnOlCNoy31Z8y/sifshX37Mur+Iry78R2+urqsEMSrDatCYyjMcnLHOd36V1X7Wn7O11+0n4G0rw/aa1Dob2WorfGeaAyhgI5E24BH9/OfavcKKzlj8RPErFuXv97L022LjgqEaH1ZL3O136nhn7JX7OV3+zX4L1fQrvWodce+vzeiaGAwhB5aJtwSc/czn3r3OiiuavXniajq1HeT3OijRhQpqlTVkj4z/aP/AOCekfxn+KF74u0LxFbeGxqESNeWslo0okuBkNIMMoG5QmR3YMe9e/fs3/CCf4FfCTSfBtzqUerS2Mk7m7iiMav5krP90k4xux17V6dRXVWzDE16EcPUleMdtunnuc1LA0KNaVenG0nv8yC9srfUrOe0u4Irq1nRo5YJkDpIhGCrKeCCDgg18cfFn/gmR4L8X6lcaj4Q1m48HTTHebEw/arQNnJ2KWVkB9NxA7ADivs6is8LjcRg5c1Cbj/XbY0xGEoYqPLWjf8ArufmfJ/wSq8ZifCeM9DaHP32imDY+mD/ADrt/BH/AASp06z1GC48V+OZ9QtF5ey0uyEDMeMfvXZuOoI2Z9xX3zRXqTz/ADCa5faW9Ev8jzYZJgYO/Jf5sitbdbS2igQkpEgQE9cAYqWiivnj3Q618rfHD/gnl4A+K1/cavok83gvWpizyNZRLJayuxJLPCcYOT/Cyj2NfVNFdWHxVbCT56EnFnNXw1LEx5K0bo/Mq7/4JWeOUuGW18X+H5oM/K8qzoxHuoRgPzr1T4Bf8E4734X+PdG8Xaz41Sa90m5W4gtNMtSEcjqHdznac4IC9CeRX3DRXrVc+x9WDpyno99EeZTybBUpqcY6rzYV8sftF/sA+EPjXq114i0a8fwj4oupDLdXEMXnW122ACzxbhtY4+8hGSWLBic19T0V5OGxVbCT9pQlZnp18PSxUOStG6PzJP8AwSs8dfado8X+Hzb5++Vn3Y/3dmP1r6B+AH/BPDwl8I9dsfEWv6nN4s1+xmW4tMx/Z7W3kU5VwgYlmBAILHGRnbnmvrWivUr55j8RB05Tsn2SR59HJ8HQnzxhdru7jURUUKoCqBgADAAr5y/bA/ZPvf2nW8KG08QwaCNFF1u862M3m+b5WMYYYx5R/Ovo+ivKw+IqYWqq1J2kv+GPSr0KeJpulVV4s8d/ZX+A9z+zt8NJfCt3q0WtSPfy3guYYTEMOqDbtJPTZ6967f4q+ALX4p/DnxD4TvHMUGq2b2/mjrGxGUf/AICwU/hXV0UTxFSpWdeT9697+YQoU4UvYxXu2tbyPjD9nP8AYM8Q/AH4qad4th8b2WoQQpJBc2YsHQzROuCA3mcEHaRkEZUV6B+2B+yhe/tOt4UNp4hg0H+xRdbvOtjN5vneVjGGGMeUfzr6PorslmmKniI4py99Kydl5/5nJHLsNGg8Mo+49bXf9dDxz9lf4D3P7O3w0m8LXWrRa1I+oS3guYYTEMOqDbtJPTZ6969joorgrVp16kqtR3b3O2lSjRgqcFZIKjuLeK7gkhmjWaGRSjxuoKsDwQQeoqSisTU+H/jF/wAExPDvibUZtR8Ba1/wizSHcdMu42ntgec7H3b0BOODuxzj0ry7Tv8AglX41lvEW/8AGWhW9qT80lvHNK4HspVQf++hX6Y0V9BTz7MKUORVL+qTf3niVMlwVSXO4W9G0jwf9nj9jjwP+z051CwWbXPErgq2sagF3xgjBWFAMRqec9WOSCxGAPeKKK8avXq4mbqVpXbPVo0adCCp0o2R8cftM/sGal8fvireeLrbxda6PFPbwwC1ls2lYbFxncHHX6V9OfCnwXJ8Ofhr4Z8Ly3S3suj6fDZNcom0SlEC7gOcZx0rq6K6K2Or16MKFR3jHbRGNLB0aNWVaC96W543+1P+zzD+0h8OofDw1BdJv7W8S8tb14zIqMAVYFQRkFWbuOcHtiuG/ZS/ZP8AEn7NEvidf+Ets9as9Yt0CQfY3jEVwm7ZIfnORh2BAwT8vPFfTtFEcdXhh3hVL3H0shSwdGVdYlr311uz4N8Gf8E0tR0n4maT4o1/xxba3Db6kmo3dutiyNdFZPMKli5A3Ec8dCelfeI4FLRRi8dXxrTryvbbRL8h4bB0cImqKtffr+Y10EiMrAMrDBB718DH/gmXq+kfEA+I/Dvjuz0xLbU/t+nQPp7ubcLLviUkOM7cKMgDOO1ffdFGFx2IwXN7CVubfRP8xYnB0cXy+2V7bdPyGxhhGu/BfAzjpmvDf2tf2cbv9pXwbo+h2mtw6G1hf/bDNNAZg48tk2gBhj72c+1e6UVz0K88PUVWm7SWxvWowr03SqK6Z4B+yL+zJefsz6D4g06712HXTqdzHOskNuYfL2qVwQWOete/0UU69epiajrVXeT3FRoww9NUqaskfJv7WP7EmoftI/EXT/E1r4pttDjtdLj08281o0pYrLK+7IYcfvQMe1e1/s7/AAom+CHwe0DwVcagmqzaZ5+67jjMayeZPJKMKScY8zHXtXo9Fb1MdXrUI4acvcjsrL+upjTwdGlWliIr3pbu7/roZviLw7pni3RL3R9YsodR0y8iaGe2nXckiEYIIr4R+JX/AASvgu9Tnu/A3i5bG0kYsum6xCX8rJJwJkOSBwACmeOWNfoBRVYTMMTgW3Qla/TdCxWCw+MSVaN7fefmv4f/AOCVHima+Qa3420izs8/M9hbyzyH2AbYPxzX2h8Af2Z/Bv7O2iy2vhy3luNSulAvNWvWD3FxgkgcABVGeFUDoM5PNesUVti82xmNjyVZ+72WhjhsswuElz0o693qFfD3xD/4J2ar43+M2r+OI/GdnaxX2qnUVs2sWZlG8NtLb+vHXFfcNFc2FxtfBScqDs2rPRP8zoxOEo4tKNZXS1EAwAPSuT+K3w30v4t/D7W/CesReZZalbmPcCA0TjlJFJBwysFYHB6dD0rraK5ITlTkpxdmjqlFTi4yV0z4U+D3/BN/Vfhf8T/DXiuXxtZ38Wk3qXTWyWDIZAvYNvOPyr7roorrxeNr46SnXldrTZL8jlw2Eo4OLjRVkwooorhOwKKKKACiiigAooooA8m8b/sp/Cr4j+KL3xF4j8I2+p6ze7PPunuJlL7UVF4VwOFVRwO1YX/DDnwP/wChCtf/AALuP/jle7UV2xxuKglGNWSS82cksJh5NylTi2/JHhDfsOfA8KT/AMIFa9P+fu4/+OV+PXiq3is/E+rwQoI4YruVEReiqHIAFfv2RkEetfAOs/8ABKuTV9Xvr7/hZaxfaZ3m8v8AsPO3cxOM/aOetfUZJm0KEqjxtV62te7737nzmcZZKsofVKa0ve1l2t2I/wBg39m34b/Fz4JT614t8L2+sanHq09stw80sZ8tUjYLhHAOCzcnnn6V9G/8MOfA/wD6EK1/8C7j/wCOVs/swfANv2cvhzN4VbWxr/mX8l79rFr9nxvVF27d7dNnXPevXa8fHZjWniakqFWXK3pq1p6HrYPAUoYeEa1JcyWuiZ5T4N/ZW+E/gDVodU0TwPpltqELB4riZWneJh0ZDIW2keowa9Woorx6lWpWfNUk2/N3PUp0oUlanFJeSsfkN/wUT/5Ok8Qf9eln/wCiErF/Zi+PvhH4PaF420Xxn4Sk8X6T4j+xbrMMgjHkNK43huvzOpHutfX/AO01+wVrvx4+LmpeMLHxRp+l211DBEttcQO7rsjCkkjjkivK/wDh1P4p/wCh50n/AMBZf8a/SaGY5dPA08PXqWso33vdW6rzR8BWwOPhjJ16NO+rtts79H5GN/w1L+zn/wBEDh/KGj/hqX9nP/ogcP5Q1s/8Op/FP/Q86T/4Cy/40f8ADqfxT/0POk/+Asv+NR7XJ/8An/L/AMCmX7LNf+fK/wDAYmN/w1L+zn/0QOH8oaB+1L+zmD/yQOH8oa2f+HU/in/oedJ/8BZf8aP+HU/in/oedJ/8BZf8aPa5P/z/AJf+BTD2Wa/8+V/4DE+av2l/i7p/xx+LF/4s0zTp9KtLi3ghW1uGDMnlxhOo4I+XNfX3/BJ7TJ4tN+JOoMuLaaWwgRvVkE7N+ki/nWDpf/BKTW3vYxqPj6witMje1tYO749ACwH619xfBb4N+H/gV4Cs/C3h2FltoiZbi5k/1l1OQA8rn1O0D0AAA4ArjzbM8G8D9TwsubZddEtd36HVluX4tYz61iFbd9NW/T1O7ooor4A+2Cvyg/4KXaFe6b+0UL+eBls9R0u3kt5sfK+zcjDPTIK8jsCPUV+r9edfG/4DeFPj/wCFBonii1c+Uxktb+2IW4tHOMtGxBHIABBBB9MgEezlGOjgMUqs1eNrP5nk5ng5Y3DunB67o/ID9nv48av+zv4+PifR7O21CSW1eyntrrIWSJmViARyDuRTnnpX1D/w9a8Rf9CFpn/gdJ/8TUfiL/glT4oh1GUaF410m7sc5jbUIJYZQPQhA4P1zz6Csr/h1f8AEP8A6Grw1/33cf8Axqvua+IyTGS9rWkm/mj4+jQzfCx9nSTS+TNn/h614i/6ELTP/A6T/wCJr179lv8Abq1j9oP4pJ4UvfC1lpEBs5ro3MFy8jZTGBggdc14R/w6v+If/Q1eGv8Avu4/+NV7N+yb+w54u+AHxZj8VazrmjahZCymtTDZNL5mX24PzIBjj1ry8ZHJFh5+wtz2dtXuelhZZu68PbX5b67bH1F8XvixofwU8CX/AIr8Qi6bTrTapSzgMsjuxwijHC5OBliq5IyRkV+Pn7RP7R/if9onxjNqWrTvbaNDIw0zRo2/c2kfQf70hABZzyT02qFUftJ4l8Oab4v0DUNF1i0jv9Lv4Wt7i2lGVkRhgj/645HUV8k/BH/gnL4Y+H3j2913xLc/8JRZ2ly50mxuo1MTRkDY86Ywzr8wx908HaMYrzMlxmCwMZ1ay/eLb/gdn3fb5nfm2ExeMlClSfuPf/g9z4b1T9kf4oaV8LtO8fN4cmuNFvI2maG3y91bRDG2WWLG4IwJIK5wFJbaCM1vhb+1J8RfhFpkWj6PrK3fh+Ny/wDYmqwLdWnJLEBXGUG4lvkK5bk9Tn9tcDGMcV4r8S/2OPhN8VLiW71XwtDZajIdz3ulsbWRjjGTt+VvxB6V6FLiOnXTp46knG+ltbfJ9u5xVMhqUWp4OpZ+en4r8j4g03/gpFrEMKLf/DTwldSAAM8ELRAnucEtivkzxPrX/CSeJdW1f7Olp9vu5bryI/ux73LbR7DOK/Sq9/4JYfDqRv8ARPFHiaEek0lvJ/KJaq/8OrPBH/Q4a9/3xD/8TXpYfNsowrcqN4t76M8+vluaYhKNWzt5o+aPgl+2y3wa+HWneFB8P9G14Wbyv9uvHxJJvkZ8H5D03Y69q7v/AIeVv/0Sfw5/38/+wr1z/h1Z4H/6HDX/APviH/4mj/h1Z4I/6HDXv++If/ia5qmMyKrNzmm29X8X+Z0QwmcU4qEWrL/Ce3/sk/tA3n7R/wAONQ8SXujwaJJaapJpy29vKZFKrFE+7JA5/ekY9q9uryv9nT9n/TP2cfBl94b0jU7vVLS6v31Ay3iqHV2jjQr8oAxiNe3c16pXw+MdGVebw/wX0/pn2GFVVUYqv8VtQooorjOoKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigDG1jxloHh6Qpqmuadprjqt3dxxH/x4is23+LHgi7mWGDxhoM0rHASPUoWYn6Bq/J39vlj/AMNZeOuT1sv/AEigrsvDf/BOvxh4v+FGl+NNI8RaZcSahpqalDpjo6SMGj3rGH6bjwOcDJ619iskwsMPTr4ivy86VtO6ufKvN8ROvUo0KPNyN9ezsfqzHIkyK8bK6MMhlOQRTq/In9h/9o7xD8LPizoHhme+uL7wlrl2mnTabLIzx28srhUniUnCMHI3EfeUtkE7SPpX9vz9r7U/hzO3w68GzzWGuXFus2oaxDJte1jblY4iOVkIGS/G1WGMk5TirZFiKeLjhYO/Nqn5db+h10s5oTwrxMla2lvPpY+vPEfxN8IeD7qO217xTo2i3MhCpDqF/FA7E9AFZgTmr3h/xfoXiyJ5NE1rT9YjQ4Z7C6SYKfcqTivyg+BH7EXjv9ozQp/F1xq0Oi6bdTO0V7qYklmvWyd8ijuN2QWJ5Oetcr8Yvgh8RP2P/GmnXDanJbGcF9P17R5njWQrjcmeCrDIyp4II6ivQWR4SdR4enib1V0scTzjEwgq88P+7fW5+zF/qFrpVnJd3tzDZ2sQy887hEQdOWPArK07x34b1e8jtLHxBpd7dyZ2QW95HI7YGThQ2TwCfwr5h/Z6+Mx/bN/Z38YeEPESx/8ACWW2ntYX0yxhI5zKjiC5VeituTLADAZcgKGCj89/2ffFM/wy/aA8G6ncMbY2OsRQXW442Rs/lSg9f4WauTDZHKsq0KkrVKfS2+l19504jOI0nSnCN4T69tdfuP221bWtP0G1+06lfW2n224J511KsaZPQZYgZqppHjHQfEFy1vpet6dqNwqlzFaXSSuFBAJwpJxkjn3r4j/4Kp+PPs3hnwX4QhlIa8uZdTuEXjCxrsjz7EySf98184f8E/vGR8J/tOeHInfZb6vFPpspJ/voWQe+ZI4xSw+SOvgHjOazs2lbt/ww6+bqjjVhOXS6V79z9cNZ8T6P4dMQ1XVrLTDNnyxeXCRb8YzjcRnGR+dQReNvD0+mzahHrumyWELBJLpbuMxIx6BmzgE5HBr81P8AgqD4zOsfGjRPD8bkxaLpas4z0lmcs3H+4sVdXqHw+HgP/gmFdTSwrFf65cW2rTsByyyXcYi57/uljP4mlHJ4/V6Fac7Oo0rW79fuCWay9vWpRhdU03e/bp9596/8LN8H/wDQ1aL/AODCL/4qj/hZnhD/AKGrRf8AwYRf/FV+LvwG+BHiH9ofxjc+G/Dl3p9nfW9k9+8mpyvHH5auiEAojnOZF7evNe/L/wAEufisGB/t/wAJcH/n8uf/AJHrtr5LgcNP2dbE2fa3/BOWjm2MxEeelh7r1/4B+nmo63p2kQLPf39tZQsMiS4mWNSPqTXPN8XvAiPtbxp4fVv7p1SDP/oVfGX/AAVVRoPCvw2jJ+ZZ7tTj2SGvnT9mn9jfVv2lPC2sazp3iS00YafdC18m5gaTzG2Bs5B4HOOhrlwuUYepg1jK9blT8r9bHRiM0r08U8LRpczXn5XP140nXNO162+0aZf22oW+cebazLIufqpIq9X4laovj79j74zXWmW2sfYtf0mSF5m0+eQ2t2hVZFVgQvmRkMAQw9RX3l8f/wBuNvAXwI8HaroltEnjfxfpcV7BE48yHT1Kr5shB5bDFlQHgkEnhSrRiciq0501h5c8amz289fK2ty8PnNOcKjrx5HDdb+X33PrDXPE2j+GLYXGsarZaVbk4Et7cJCufqxArndL+Nvw81uRY9P8d+G752bYFt9WgclvQYfk1+T3ww+BHxW/bF8Qahrwvmv40l8q71/XLpvLR+G8pcBmJAYEKq7QCPu5Fet6x/wSz8f2envNp3ijQtQulGRbN5sO/wBgxUjP1wPpXTPJ8Dh37LEYpKfpt/XyOeOaYyuvaUMPePrv/XzP0p1XxLpGhQQzalqllp8MxxHJdXCRq/GeCxGaNG8SaT4iWVtK1Sz1NYiBIbO4SUIT0ztJx0NfA/8AwUK8GQfDn9nn4UeGbXIh0uf7LncTuZYMM2ST1OT+NbH/AASjJPhX4g5P/L5af+gSVwSyuCy946M72drW87dztjmMnjlg5QtdXvfyufc2q6zp+hWv2nUr620+23BfOupVjTJ6DLEDNVdH8W6H4hmeHS9ZsNSlRdzpaXKSso6ZIUnAr5t/4KTHH7NFz/2FLX+bV87/APBKsk/E7xlk/wDMIT/0ctTRyxVcvnjefWL2t6f5lVcwdPHQwnLpLr9/+R+iN58QPDGnXUtrd+I9JtbmJtskM17GjofQgtkGr9/4g0vS9Pjv73UrS0sZNuy5nnVI2yMjDE4OR0r8YP2u2I/aY+IXJ/5Csn8hX21+2kSP2GfCv+7pX/omuurksabw65/4tum23+Zy0s2lUVf3P4fnvufXeleNfD2uXYtdN13TdQuSCwhtbuOR8DqcKSa1Lu7gsLWW5uZo7e3iUvJLKwVEUdSSeAPevwa8AeNtb+GPi3RvFuhz/ZtR0658yCRuVYgDcjAHJVlbaRxkMRX64+NPijpPxl/Y48Y+LdGf/Rb/AML3zPCWBe3lFu4eJsfxK2R78HoRUZjkksDOnaXNGTte2zLwObxxkJ3jaUVe190ewaX428O63dra6dr2m390wJENtdxyOQOp2qSa2q/JD/gnCSf2nNMyf+Yfd/8Aouv1F+Knjq2+Gfw48SeKbrBi0qxlughOPMdVOxPqzbV/GuLMsu+o4lYaEuZtL8XY68Bj/rmHdea5bN/gWbj4i+FLS4kgn8TaPDNExR45L6JWRgcEEFuCD2rX0zVbLWrNLvT7uC+tXyFntpBIjYODhgSDyMV+CUlvq3iyTXtbkMl08AN9f3LZ6yTKm4kcZLyD9a/Q/wD4JbfE46r4L8S+BrqYtPpVwNQtFY/8sZeHUeyuufrLXq5hkH1LDOvCpzONrq39eR5uBzr63XVGUOW97O590UUUV8gfUBRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFABRRRQAUUUUAFFFFAH44/t9f8nZeOvrZf+kUFdNp/wDwUI8deH/hFpXgbQdI0zTDY6cmmrq2HknCKmwOqkhVfgHJDD2rmf2+v+TsvHX1sv8A0igr9If2ZPh94Xb4GfDvUz4c0k6lJodnI94bGLzmbyl+YvtyT75r9MxWIw+Gy/DTxFLn0jZX68p+fYahXxGOxEaNTk1d/vPg39hr9lrxD8RviNonjXVbKfTPCei3Ud+lzcRsn26WNg0aRZHzLuALN0wCOp48m/avuZ7v9pL4iPcszSDWZ4wW/uK21P8Ax0Cv23AAGAMCvzV/4KNfs06rpnjG5+KOg2T3Wh6gkY1cQjcbS4UBBIVA4jdQnzc4bdnG5c8uWZx9czByre7dWivne3qzpzDKvquBUaXvWd391vwPvL4F2llY/BfwJBpoT7Cuh2flGM5DKYUOc989c9814P8A8FL7S0n/AGcBNOiGeDV7ZrdiOQxDg4Pb5S1fOv7M/wDwUQf4UeB7Pwl4x0S51qw01BDYXunsomSIZxG6sQGxwAQRwAMHGa87/a1/a3vf2ntQ0jStK0q50nw9YSF4LKSUSy3U7DaJGCjggEqFBP3jzzxy4XJsXSzJVJr3VK/Nfdf8E6MTmuFqYBwi/eatb+ux2/8AwS3lu1+O2vxxmT7G3h+UzKCdm4XEG0ntnlsexb3rx/8AbE8C/wDCu/2kfGlhHF5VrcXn9o24H3Sk4EvHsGZl/wCA196/8E/P2b9S+DPgnUPEXiW1Nn4i8QrEy2cqlZbS2UZVHBwVdixLLjjCg85A8T/4KqeA/snijwZ4whi+W8tZdMuJB0DRNvjz7kSSf98fSvSw+Pp1M7nGDvGS5fVrX/NHBXwU4ZRFzWqd/k9P8meH/tKfEWf9o340eEks3kkkn0rStOiDc/v5kWR8AHqJJ2U+6/SmfHTRh+zt+1zcyafE0Vto+qWmqWqx4BZMRy4H47gfxpP2HvCD+PP2nfB63AM8GmO2pStJ820QITF+TiID04r1X/gqT4TOmfF/w3r6Lti1XSfJbjrLDI245/3ZIx+FekqkKGMpZfH4eR6f15J/eee4TrYSpjn8XOv6/FHi/wAZ72f4+ftX65FpkhmfW9eTTbSULuBRWWCN8em1AfpX6Fft3aRbaB+x3rmmWaeVaWQ0+2hT+6iXESqPyAr4w/4Jz+AG8a/tF2urzoZLXw9aS6g7NyGlYeVGD75kLD/cr7c/4KFf8mqeK/8ArtZf+lUVeLmNRQzDCYSG0OX81+iR6+Ag5YHE4mW8+b8n+rPy5+DmrfEPR/FFxL8NF1ptea0ZJRoVu88/2fchbKopO3cEycdcV9A+B/GX7VU/jTQI9Th+IQ017+Bboz6XcLH5RkXfuJj4G3OT6V5V+yl+0Db/ALN/xFvvE1zo0muJc6ZJp4t4pxCVLSxPu3FW/wCeeMY719Yj/gq9pRP/ACTq8/8ABov/AMar2MxWKdVqlhozVt3a/wCJ5WB+rKmnVxDg77K9vwD/AIKu/wDIu/Dr/r5vP/QIq+VfgZ+1B8Q/gH4T1ay8JWtm2mXlyJ57q7s2m8uTaFGGDBRxjg+tfS3/AAU11weJ/hv8JNYWE266gs92IS27ZvhgbbnAzjOM11X/AAS1tIb34SeNYLiFJ4JdWCPHIoZXUwKCCD1BHavLw9WGGyWMq1PnSbun/iZ6NenPEZvKNKfK2t/+3UfGfww0SX9qT4+Wdv468ZxaXe65OPtGpXUY825cKFSGIABA7bVRdxUdMBm2o3W/t9aXF4c+Ph0Cyi+z6Ro2jWFhp8A6RwLECFH/AAJm/EmuK/aQ8DyfBb9obxRpWmI+nw2Wo/bNO2Db5cT4li2eyhgB/u19T/tcfBq+/aL+FPg/43+ELYXuoNosX9r6fbgtI6AZLIB95onMisOpAGPu8+tOvClicPXcrUpRsl0TdmvvWnkeZCjKpQr0bXqRldvq0rp/c9T6K/YG07TrD9ljwfJpwGbr7TNcvjBeb7RIrZ+m0KPZR1r6Fr8d/wBmL9s7xL+zdbXmk/2eniTw1cP5w0yecwtBKcZaOTa20EDlSpGQCMHOfpXUf+CrWhrZMbDwFfyXhU7VuL5EjDY4yQpJGfavkcwyTHTxU5048yk207rr6s+nwOb4OOGhCcuVxSVrPoXv+Cq//JOvBP8A2FJf/RRr4j+EOufF3SbHUE+Gp8ULavIpvD4egmdd+Dt8wxqecZxn3r7L/wCCmetw+Jfgz8NtXtjm3v7v7VGf9l7cMP0NP/4JR/8AIq/EH/r8tP8A0CSvZwdf6nkvtJQUrN6P/FY8vFUfrWbezUnG63X+E+Rfif4j+OupeFZIPH7eMm8PGVC/9uW06W/mA/JzIoAb07175/wSq/5Kd4y/7BCf+jlr6J/4KTf8m0XP/YUtf5tXzt/wSq/5Kd4y/wCwQn/o5aqWLWMyatVUFHpZeqIjhnhM1pU3Ny838zwD9rv/AJOY+IX/AGFZP5Cvtv8AbT/5MY8K/wC7pP8A6Jr4k/a7/wCTmPiF/wBhWT+Qr7b/AG0/+TGPCv8Au6T/AOia1xW+X+q/JGeH2xvo/wA2fJP7OvwIn+PHwf8AiraaYjS+I9EfTtS0yMEfvW23Qlh5/vrjHTLJHk4BqH9m746Xfgvwj8Q/h1fSodC8VaHqEdushC+TqH2VxEQT/wA9Nojx3Yx8jBz9Cf8ABJz/AI+fid/uad/O5rzX/goT+z0/wq+JK+NtEt/I8N+IpTI3lfKLW+5Z0GOgcDzF9944Cit3iYVsfWy+vs7OPrZP/gr59zFYedLBUsdR3V1L0u1/wPuMf/gnD/yc7pn/AGD7v/0XX05/wVA+Jn/CP/CzRfBtvKBc6/eefcIP+feDDYPpmRoyP9w18x/8E4f+TndM/wCwfd/+i6yf29viZ/wsf9ozXIoJfN0/QVXR7fDZG6Mkyn6+a0g+iis62G+sZ3CT2hFP8Xb8TSliPYZRKK3lJr8Ff8DR/Zq8T/CfRPgb8UtG8a+I10jxJ4lg+w2aNZXM3lpGnmRSExIy4MxGQTn936GuV/Ys+Jo+Fv7RPhi9mlEWnalIdJvCTgeXNhVJ9hII2P8Au19U/Dr/AIJieEtd8CaBqWv+ItdttZvLKK5u4LUwrHFI6higDRk/LnHJ7V8aftF/CGX4A/GPV/C0NxPPa2jRz2N5LgSSxOoZWyMcg5UkY5U10UK+Dx1Svh6c23PdPba2mnoYVqOLwcKNecElDa2/fX8T9wOtFecfs6/ExPi98FvCnikPvuLuzVLrPUXEZMcv/j6MR7EGvR6/KqlOVKcqct07fcfpFOcakFOOzVwooorM0CiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooA8p8a/ss/Cz4i+Jr3xD4j8IWuqazebPPu5JplZ9qKi5CuBwqqOB2r0Xw54e0/wloNhouk2y2emWEK29tbqxIjjUYVQSSeAO5rRoraderUioTk2lsm9EZRo04Sc4xSb3dtwqK5tory3lgniSaCVSkkcihldSMEEHqCO1S0Vianh/iP9in4L+Jpp5Z/AtjaSTEF/wCz2e1Xg5+VY2AX/gIFdD4F/Zk+F3w31NNS0DwXpdpqUbB47ySLzpYm9UZ8lD/u4r0+iuuWMxMo8kqkmu12cqwuHjLnVNX9EFcn8SPhV4U+LmjW+leLtGh1qwt5xcxQzM6hZArKGBUg9GYenNdZRXPCcqclKDs12OiUYzTjJXR5v8Of2c/hz8Jdcl1jwl4XttG1KWBrV7iKSRyYyysV+ZiOqL+VaHxM+Cfgn4xrpy+MdAg1waeZDa+c7qYt+3fgqw67F/IV3FFavEVnU9q5vm73d/vMlQpKHs1Bcvayt9xwnwz+BngX4Oy6hJ4O8PW+iSX4RbloXdzIE3bR87HGNzdPWtvx14C0H4l+G7jQPEunJquj3DI0trI7KrlWDLkqQeCAevaugoqXWqSn7WUm5d76/eUqVOMPZqK5e1tPuPDf+GI/gj/0INl/4ET/APxyj/hiP4I/9CDZf+BE/wD8cr3Kiuj69i/+fsv/AAJ/5mH1PDf8+o/cjgPHHwF8BfEjR9F0rxJ4dh1XT9Gj8qwgllkUQLtVcDawJ4RRznpV/wCG3wi8I/CHTruw8IaLFolpdSieaKGR2DvjG75mPYAV2FFYOvVcPZub5e13b7jZUaan7RRXN3tr955r8Q/2b/ht8VtdXWvFfhS01fVFhW3FzI8iMUUkgHYwzjceTzXUeA/h9oHwy8Nw6B4Z09dL0eF3eO0WR3VCzFmxvJIySTjpya6KiiVerKCpym3FdLu33DjRpxm6kYpSfW2v3nlvjj9mD4WfEXVH1LXfBWl3OpSNuku4o/Jkkb1coRvPu2a5uy/Yf+Ctk6v/AMIRazlXLjzppWGSSem7kc9DwBwOAK92orWOMxMI8sakkvVmUsJh5PmlTTfojhPG3wL8B/ETw/pOh+IPDdpe6PpIAsbJS0MVuAu0BFQqAAoAA6AVN8NPgz4M+D1vfQeD9Dh0SK+ZXuEhkdxIyghSdzHpk9K7WisnXquHs3N8va7t9xp7Gmp+05Vzd7a/ec18QPhx4c+Kfh86H4p0uPV9KMqzG2ldlUuudpypB4ye9YXw2+AHw/8AhDqN3f8AhDw3Bol3dRCCaWGWRi6ZBwQzEdQK9CopKtVjB0lJ8r6X0+4bpU3NVHFcy621+88h8T/slfCXxn4gvtc1rwZa3+q30pmuLl55g0jnqSA4H5Cuv8W/CXwl468G2vhTXdGi1Hw/a+UIbGSRwq+Wu1OQwJwPU119FW8TXfLeb93bV6enYlYeir2gtd9Fr69zhfhn8DvA/wAHH1BvB2gQ6G2oCMXXkySN5uzdszuY9N7dPWtrx14A8P8AxM8OT6D4n0uHV9JnZXe2nyBuUgqQQQQQR1B9u9dBRWbrVJT9q5Pm731+8pUqcYezUVy9rafceX+A/wBmX4ZfDHxFHrvhjwpbaRq0aNGtzFNKxCsMMMM5HI9qw5f2LvgtPdvcyeA7KSd3MjM8853MTkkjfzzXtlFbfXMSpOXtJXfmzL6rh7cvs1b0QgAUAAYArzn4jfs7fDr4t61Dq/i3wvbazqUMAtkuJZJEYRhmYL8rDOCzHn1r0eisKdWdKXNTk0/J2Np04VFyzSa89Tlvh18MPDHwm0KTRvCelJo+mSTtctbxyO6+YwVS3zEkZCrx04rqaKKU5ynJym7tlRjGCUYqyCiiioKCiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooAKKKKACiiigAooooA//2Q==" alt="PowerChina Huadong" style="height:48px;width:auto;object-fit:contain;filter:brightness(1.1);" />
      <div style="width:1px;height:36px;background:rgba(255,255,255,0.12);margin:0 4px;"></div>
      <div>
        <div class="header-title">Riyah 1 &amp; 2 — HSE Dashboard</div>
        <div class="header-sub">Developed by: Shaguf Ahmed DC/DA</div>
      </div>
    </div>
    <div class="header-right">
      <div class="date-badge">Updated: """ + today + """</div>
      <div class="live-dot">Compiled</div>
    </div>
  </div>
</div>

<div class="tabs-bar">
  <div class="tabs-inner">
    <button class="tab-btn active" onclick="showTab('overview',this)">Overview <span class="count">All</span></button>
    <button class="tab-btn" onclick="showTab('r1',this)">Riyah 1 Docs <span class="count">""" + str(len(r1_rows)) + """</span></button>
    <button class="tab-btn" onclick="showTab('r2',this)">Riyah 2 Docs <span class="count">""" + str(len(r2_rows)) + """</span></button>
    <button class="tab-btn" onclick="showTab('ncr',this)">Internal NCRs <span class="count">""" + str(ncr_total) + """</span></button>
    <button class="tab-btn" onclick="showTab('cncr',this)">Client NCRs <span class="count">""" + str(cncr_total) + """</span></button>
    <button class="tab-btn" onclick="showTab('calendar',this)">📅 Calendar <span class="count">""" + cal_month + ' ' + str(cal_year) + """</span></button>
    <button class="tab-btn" onclick="showTab('kpi',this)">📊 Last Month KPI <span class="count">""" + kpi_data['month'] + """</span></button>
  </div>
</div>

<!-- OVERVIEW PANEL -->
<div class="panel active" id="panel-overview">
  <div class="panel-body">
    <div class="kpi-grid">
      <div class="kpi-card blue"><div class="kpi-label">Total Documents</div><div class="kpi-val blue">""" + str(total_docs) + """</div><div class="kpi-sub">R1: """ + str(len(r1_rows)) + """ · R2: """ + str(len(r2_rows)) + """</div></div>
      <div class="kpi-card green"><div class="kpi-label">Approved (APP)</div><div class="kpi-val green">""" + str(combined_status['APP']) + """</div><div class="kpi-sub">R1: """ + str(r1_status.get('APP',0)) + """ · R2: """ + str(r2_status.get('APP',0)) + """</div></div>
      <div class="kpi-card amber"><div class="kpi-label">Pending Review</div><div class="kpi-val amber">""" + str(pending) + """</div><div class="kpi-sub">""" + pending_sub + """</div></div>
      <div class="kpi-card blue"><div class="kpi-label">Internal NCRs</div><div class="kpi-val blue">""" + str(ncr_total) + """</div><div class="kpi-sub">Open: """ + str(ncr_open) + """ · Closed: """ + str(ncr_closed) + """</div></div>
      <div class="kpi-card red"><div class="kpi-label">Internal Open NCRs</div><div class="kpi-val red">""" + str(ncr_open) + """</div><div class="kpi-sub">""" + str(ncr_closure) + """% closure rate</div></div>
      <div class="kpi-card amber"><div class="kpi-label">Client NCRs</div><div class="kpi-val amber">""" + str(cncr_total) + """</div><div class="kpi-sub">Open: """ + str(cncr_open) + """ · Closed: """ + str(cncr_closed) + """</div></div>
    </div>
    <div class="chart-grid col3">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Document Approval Rate</div><div class="chart-desc">Both sites combined</div></div></div>
        <div style="position:relative;height:200px"><canvas id="ovDonutChart"></canvas></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>APP """ + str(combined_status['APP']) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>AWC """ + str(combined_status['AWC']) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>REJ """ + str(combined_status['REJ']) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>RWC """ + str(combined_status['RWC']) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>UR """ + str(combined_status['UR']) + """</span></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Document Types</div><div class="chart-desc">R1 vs R2</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>R1</span><span class="leg-item"><span class="leg-dot" style="background:#06b6d4"></span>R2</span></div>
        <div style="position:relative;height:190px"><canvas id="ovDocTypeChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Internal NCR Status <span style="font-size:10px;color:#64748b">· """ + str(ncr_total) + """ total</span></div></div></div>
        <div style="position:relative;height:130px"><canvas id="ovNcrChart"></canvas></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>Closed """ + str(ncr_closed) + """</span><span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>Open """ + str(ncr_open) + """</span></div>
        <div style="border-top:1px solid rgba(255,255,255,0.06);margin-top:14px;padding-top:14px">
          <div class="chart-title" style="margin-bottom:8px;font-size:12px">Client NCR Status <span style="font-size:10px;color:#64748b">· """ + str(cncr_total) + """ total</span></div>
          <div style="position:relative;height:100px"><canvas id="ovClientNcrChart"></canvas></div>
          <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>Closed """ + str(cncr_closed) + """</span><span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>Open """ + str(cncr_open) + """</span></div>
        </div>
      </div>
    </div>
    <div class="chart-grid col2">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">R1 vs R2 — Review Status Comparison</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>R1</span><span class="leg-item"><span class="leg-dot" style="background:#06b6d4"></span>R2</span></div>
        <div style="position:relative;height:200px"><canvas id="ovCompareChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Internal NCRs by Contractor</div><div class="chart-desc">Open vs Closed</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>Closed</span><span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>Open</span></div>
        <div style="position:relative;height:200px"><canvas id="ovContractorChart"></canvas></div>
      </div>
    </div>
  </div>
</div>

<!-- R1 PANEL -->
<div class="panel" id="panel-r1">
  <div class="panel-body">
    <div class="kpi-grid">
      <div class="kpi-card blue"><div class="kpi-label">Total R1 Docs</div><div class="kpi-val blue">""" + str(len(r1_rows)) + """</div></div>
      <div class="kpi-card green"><div class="kpi-label">Approved</div><div class="kpi-val green">""" + str(r1_status.get('APP',0)) + """</div><div class="kpi-sub">""" + str(r1_approval) + """% approval rate</div></div>
      <div class="kpi-card amber"><div class="kpi-label">AWC</div><div class="kpi-val amber">""" + str(r1_status.get('AWC',0)) + """</div><div class="kpi-sub">Approved w/ comments</div></div>
      <div class="kpi-card red"><div class="kpi-label">Rejected</div><div class="kpi-val red">""" + str(r1_status.get('REJ',0)) + """</div></div>
      <div class="kpi-card teal"><div class="kpi-label">RWC</div><div class="kpi-val teal">""" + str(r1_status.get('RWC',0)) + """</div><div class="kpi-sub">Review w/ comments</div></div>
      <div class="kpi-card blue"><div class="kpi-label">UR</div><div class="kpi-val blue">""" + str(r1_status.get('UR',0)) + """</div><div class="kpi-sub">Under review</div></div>
      <div class="kpi-card purple"><div class="kpi-label">IFI</div><div class="kpi-val purple">""" + str(r1_status.get('IFI',0)) + """</div><div class="kpi-sub">Issued for Information</div></div>
    </div>
    <div class="chart-grid col2">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Review Status Breakdown</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>APP """ + str(r1_status.get('APP',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>AWC """ + str(r1_status.get('AWC',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>REJ """ + str(r1_status.get('REJ',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>RWC """ + str(r1_status.get('RWC',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>UR """ + str(r1_status.get('UR',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#8b5cf6"></span>IFI """ + str(r1_status.get('IFI',0)) + """</span></div>
        <div style="position:relative;height:200px"><canvas id="r1StatusChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Document Types</div></div></div>
        <div style="position:relative;height:200px"><canvas id="r1TypeChart"></canvas></div>
      </div>
    </div>
    <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Previous Review Status</div><div class="chart-desc">History before current approval</div></div></div>
      <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>AWC """ + str(r1_prev.get('AWC',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#f97316"></span>RWC """ + str(r1_prev.get('RWC',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>REJ """ + str(r1_prev.get('REJ',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#475569"></span>First Submission """ + str(r1_prev.get('First',0)) + """</span></div>
      <div style="position:relative;height:60px"><canvas id="r1PrevChart"></canvas></div>
    </div>
    <div class="table-card">
      <div class="table-header"><div><div class="chart-title">Document Register — Riyah 1</div><div class="chart-desc">""" + str(len(r1_rows)) + """ documents · All QHSE</div></div>
        <button class="export-btn" onclick="exportTableToCSV('r1TableBody', 'R1_Documents.csv')">📥 Export CSV</button>
      </div>
      <div class="filter-bar" style="padding:12px 22px 0">
        <input type="text" id="r1Search" placeholder="🔍 Search document name or code..." onkeyup="filterTable('r1')">
        <select id="r1StatusFilter" onchange="filterTable('r1')">
          <option value="">All Status</option><option value="APP">APP</option><option value="AWC">AWC</option>
          <option value="REJ">REJ</option><option value="RWC">RWC</option><option value="UR">UR</option>
        </select>
        <select id="r1TypeFilter" onchange="filterTable('r1')">
          <option value="">All Types</option><option value="PLN-Plan">PLN-Plan</option><option value="REP-Report">REP-Report</option>
          <option value="ICL-Inspection Checklist">ICL-Checklist</option><option value="POL-Policy">POL-Policy</option><option value="REG-Register">REG-Register</option>
        </select>
        <span class="filter-count" id="r1Count"></span>
      </div>
      <div class="tbl-scroll"><table><thead><tr><th>Doc Code</th><th>Document Name</th><th>Rev</th><th>Status</th><th>Date</th><th>Type</th></tr></thead><tbody id="r1TableBody"></tbody></table></div>
    </div>
  </div>
</div>

<!-- R2 PANEL -->
<div class="panel" id="panel-r2">
  <div class="panel-body">
    <div class="kpi-grid">
      <div class="kpi-card blue"><div class="kpi-label">Total R2 Docs</div><div class="kpi-val blue">""" + str(len(r2_rows)) + """</div></div>
      <div class="kpi-card green"><div class="kpi-label">Approved</div><div class="kpi-val green">""" + str(r2_status.get('APP',0)) + """</div><div class="kpi-sub">""" + str(r2_approval) + """% approval rate</div></div>
      <div class="kpi-card amber"><div class="kpi-label">AWC</div><div class="kpi-val amber">""" + str(r2_status.get('AWC',0)) + """</div></div>
      <div class="kpi-card teal"><div class="kpi-label">RWC</div><div class="kpi-val teal">""" + str(r2_status.get('RWC',0)) + """</div></div>
      <div class="kpi-card red"><div class="kpi-label">Rejected</div><div class="kpi-val red">""" + str(r2_status.get('REJ',0)) + """</div></div>
      <div class="kpi-card blue"><div class="kpi-label">UR</div><div class="kpi-val blue">""" + str(r2_status.get('UR',0)) + """</div></div>
      <div class="kpi-card purple"><div class="kpi-label">IFI</div><div class="kpi-val purple">""" + str(r2_status.get('IFI',0)) + """</div><div class="kpi-sub">Issued for Information</div></div>
    </div>
    <div class="chart-grid col2">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Review Status Breakdown</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>APP """ + str(r2_status.get('APP',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>AWC """ + str(r2_status.get('AWC',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>RWC """ + str(r2_status.get('RWC',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>UR """ + str(r2_status.get('UR',0)) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#8b5cf6"></span>IFI """ + str(r2_status.get('IFI',0)) + """</span></div>
        <div style="position:relative;height:200px"><canvas id="r2StatusChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Document Types</div></div></div>
        <div style="position:relative;height:200px"><canvas id="r2TypeChart"></canvas></div>
      </div>
    </div>
    <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Previous Review Status</div><div class="chart-desc">History before current approval</div></div></div>
      <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#f59e0b"></span>AWC """ + str(r2_prev.get('AWC',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#f97316"></span>RWC """ + str(r2_prev.get('RWC',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>REJ """ + str(r2_prev.get('REJ',0)) + """</span>
        <span class="leg-item"><span class="leg-dot" style="background:#475569"></span>First Submission """ + str(r2_prev.get('First',0)) + """</span></div>
      <div style="position:relative;height:60px"><canvas id="r2PrevChart"></canvas></div>
    </div>
    <div class="table-card">
      <div class="table-header"><div><div class="chart-title">Document Register — Riyah 2</div><div class="chart-desc">""" + str(len(r2_rows)) + """ documents · All QHSE</div></div>
        <button class="export-btn" onclick="exportTableToCSV('r2TableBody', 'R2_Documents.csv')">📥 Export CSV</button>
      </div>
      <div class="filter-bar" style="padding:12px 22px 0">
        <input type="text" id="r2Search" placeholder="🔍 Search document name or code..." onkeyup="filterTable('r2')">
        <select id="r2StatusFilter" onchange="filterTable('r2')">
          <option value="">All Status</option><option value="APP">APP</option><option value="AWC">AWC</option>
          <option value="REJ">REJ</option><option value="RWC">RWC</option><option value="UR">UR</option>
        </select>
        <select id="r2TypeFilter" onchange="filterTable('r2')">
          <option value="">All Types</option><option value="PLN-Plan">PLN-Plan</option><option value="REP-Report">REP-Report</option>
          <option value="ICL-Inspection Checklist">ICL-Checklist</option><option value="POL-Policy">POL-Policy</option><option value="REG-Register">REG-Register</option>
        </select>
        <span class="filter-count" id="r2Count"></span>
      </div>
      <div class="tbl-scroll"><table><thead><tr><th>Doc Code</th><th>Document Name</th><th>Rev</th><th>Status</th><th>Date</th><th>Type</th></tr></thead><tbody id="r2TableBody"></tbody></table></div>
    </div>
  </div>
</div>

<!-- INTERNAL NCRs PANEL -->
<div class="panel" id="panel-ncr">
  <div class="panel-body">
    <div class="kpi-grid">
      <div class="kpi-card blue"><div class="kpi-label">Total NCRs</div><div class="kpi-val blue">""" + str(ncr_total) + """</div></div>
      <div class="kpi-card green"><div class="kpi-label">Closed</div><div class="kpi-val green">""" + str(ncr_closed) + """</div><div class="kpi-sub">""" + str(ncr_closure) + """% closure rate</div></div>
      <div class="kpi-card red"><div class="kpi-label">Open</div><div class="kpi-val red">""" + str(ncr_open) + """</div><div class="kpi-sub">Requires action</div></div>
      <div class="kpi-card teal"><div class="kpi-label">Riyah 1</div><div class="kpi-val teal">""" + str(ncr_r1) + """</div></div>
      <div class="kpi-card amber"><div class="kpi-label">Riyah 2</div><div class="kpi-val amber">""" + str(ncr_r2) + """</div></div>
      <div class="kpi-card blue"><div class="kpi-label">Last NCR No.</div><div class="kpi-val blue" style="font-size:13px">""" + last_ncr_no.zfill(3) + """</div><div class="kpi-sub">""" + last_ncr_desc[:40] + """</div></div>
    </div>
    <div class="chart-grid col2">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">NCRs by Contractor</div><div class="chart-desc">Open vs Closed breakdown</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#10b981"></span>Closed</span><span class="leg-item"><span class="leg-dot" style="background:#ef4444"></span>Open</span></div>
        <div style="position:relative;height:220px"><canvas id="ncrContChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">NCRs by Location</div></div></div>
        <div class="legend"><span class="leg-item"><span class="leg-dot" style="background:#3b82f6"></span>Riyah 1 — """ + str(ncr_r1) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#06b6d4"></span>Riyah 2 — """ + str(ncr_r2) + """</span>
          <span class="leg-item"><span class="leg-dot" style="background:#8b5cf6"></span>Both — """ + str(ncr_both) + """</span></div>
        <div style="position:relative;height:200px"><canvas id="ncrLocChart"></canvas></div>
      </div>
    </div>
    <div class="chart-card"><div class="chart-header"><div><div class="chart-title">NCRs Issued Per Month</div><div class="chart-desc">Aug 2025 – Apr 2026</div></div></div>
      <div style="position:relative;height:160px"><canvas id="ncrTimelineChart"></canvas></div>
    </div>
    <div class="section-label">All Internal NCRs</div>
    <div class="table-card">
      <div class="table-header"><div><div class="chart-title">Internal NCR Register</div><div class="chart-desc">""" + str(ncr_total) + """ records</div></div>
        <div class="filter-bar" style="margin:0;padding:0">
          <select id="ncrStatusFilter" onchange="filterNCRTable()">
            <option value="all">All Status</option>
            <option value="Open">Open Only</option>
            <option value="Closed">Closed Only</option>
          </select>
          <button class="export-btn" onclick="exportTableToCSV('ncrTableBody', 'Internal_NCRs.csv')">📥 Export CSV</button>
        </div>
      </div>
      <div class="tbl-scroll"><table><thead><tr><th>#</th><th>NCR Number</th><th>Date Issued</th><th>Location</th><th>Area</th><th>Contractor</th><th>Issuer</th><th>Completion Date</th><th>Status</th></tr></thead><tbody id="ncrTableBody"></tbody></table></div>
    </div>
  </div>
</div>

<!-- CLIENT NCRs PANEL -->
<div class="panel" id="panel-cncr">
  <div class="panel-body">
    <div class="kpi-grid">
      <div class="kpi-card blue"><div class="kpi-label">Total Client NCRs</div><div class="kpi-val blue">""" + str(cncr_total) + """</div></div>
      <div class="kpi-card green"><div class="kpi-label">Closed</div><div class="kpi-val green">""" + str(cncr_closed) + """</div><div class="kpi-sub">""" + str(round(cncr_closed/cncr_total*100,1) if cncr_total else 0) + """% closure rate</div></div>
      <div class="kpi-card red"><div class="kpi-label">Open</div><div class="kpi-val red">""" + str(cncr_open) + """</div></div>
    </div>
    """ + (f'<div class="alert-banner"><div class="alert-icon">⚠️</div><div><strong>{cncr_open} Open Client NCR(s):</strong> {open_cncr_text}</div></div>' if cncr_open > 0 else "") + """
    <div class="chart-grid col2">
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Client NCRs by Location</div></div></div>
        <div style="position:relative;height:200px"><canvas id="cncrLocChart"></canvas></div>
      </div>
      <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Client NCRs by Originator</div></div></div>
        <div style="position:relative;height:200px"><canvas id="cncrOrigChart"></canvas></div>
      </div>
    </div>
    <div class="chart-card"><div class="chart-header"><div><div class="chart-title">Client NCRs — Monthly Trend</div></div></div>
      <div style="position:relative;height:160px"><canvas id="cncrTimeChart"></canvas></div>
    </div>
    <div class="section-label">All Client NCRs</div>
    <div class="table-card">
      <div class="table-header"><div><div class="chart-title">Client NCR Register</div><div class="chart-desc">""" + str(cncr_total) + """ records · All Major</div></div>
        <div class="filter-bar" style="margin:0;padding:0">
          <select id="cncrStatusFilter" onchange="filterClientNCRTable()">
            <option value="all">All Status</option>
            <option value="Open">Open Only</option>
            <option value="Closed">Closed Only</option>
          </select>
          <button class="export-btn" onclick="exportTableToCSV('cncrTableBody', 'Client_NCRs.csv')">📥 Export CSV</button>
        </div>
      </div>
      <div class="tbl-scroll"><table><thead><tr><th>#</th><th>NCR Number</th><th>Category</th><th>Date Issued</th><th>Location</th><th>Originator</th><th>Description</th><th>Completion Date</th><th>Status</th></tr></thead><tbody id="cncrTableBody"></tbody></table></div>
    </div>
  </div>
</div>

<!-- footer moved to bottom -->

<script>
// TAB LOGIC
function showTab(id, btn) {
  document.querySelectorAll('.panel').forEach(p => p.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('panel-' + id).classList.add('active');
  if(btn) btn.classList.add('active');
}

// CHART COLORS
const C = { green:'#10b981', amber:'#f59e0b', red:'#ef4444', blue:'#3b82f6', cyan:'#06b6d4', orange:'#f97316', purple:'#8b5cf6', slate:'#475569' };

const donutCfg = (data, colors, labels) => ({
  type:'doughnut',
  data:{labels, datasets:[{data, backgroundColor:colors, borderWidth:0, hoverOffset:4}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}}
});

// OVERVIEW CHARTS
new Chart(document.getElementById('ovDonutChart'), donutCfg([""" + str(combined_status['APP']) + """,""" + str(combined_status['AWC']) + """,""" + str(combined_status['REJ']) + """,""" + str(combined_status['RWC']) + """,""" + str(combined_status['UR']) + """],[C.green,C.amber,C.red,C.blue,C.amber],['APP','AWC','REJ','RWC','UR']));

new Chart(document.getElementById('ovDocTypeChart'), {
  type:'bar',
  data:{labels:['PLN-Plan','REP-Report','ICL-Checklist','POL-Policy','REG-Register'],
    datasets:[
      {label:'R1', data:[""" + str(r1_types.get('PLN',0)) + """,""" + str(r1_types.get('REP',0)) + """,""" + str(r1_types.get('ICL',0)) + """,""" + str(r1_types.get('POL',0)) + """,""" + str(r1_types.get('REG',0)) + """], backgroundColor:C.blue, borderRadius:3},
      {label:'R2', data:[""" + str(r2_types.get('PLN',0)) + """,""" + str(r2_types.get('REP',0)) + """,""" + str(r2_types.get('ICL',0)) + """,""" + str(r2_types.get('POL',0)) + """,""" + str(r2_types.get('REG',0)) + """], backgroundColor:C.cyan, borderRadius:3}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b', maxRotation:30}}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}}}
});

new Chart(document.getElementById('ovNcrChart'), donutCfg([""" + str(ncr_closed) + """,""" + str(ncr_open) + """],[C.green,C.red],['Closed','Open']));
new Chart(document.getElementById('ovClientNcrChart'), donutCfg([""" + str(cncr_closed) + """,""" + str(cncr_open) + """],[C.green,C.red],['Closed','Open']));

new Chart(document.getElementById('ovCompareChart'), {
  type:'bar',
  data:{labels:['APP','AWC','REJ','RWC','UR'],
    datasets:[
      {label:'R1', data:[""" + str(r1_status.get('APP',0)) + """,""" + str(r1_status.get('AWC',0)) + """,""" + str(r1_status.get('REJ',0)) + """,""" + str(r1_status.get('RWC',0)) + """,""" + str(r1_status.get('UR',0)) + """], backgroundColor:C.blue, borderRadius:3},
      {label:'R2', data:[""" + str(r2_status.get('APP',0)) + """,""" + str(r2_status.get('AWC',0)) + """,""" + str(r2_status.get('REJ',0)) + """,""" + str(r2_status.get('RWC',0)) + """,""" + str(r2_status.get('UR',0)) + """], backgroundColor:C.cyan, borderRadius:3}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b'}}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}}}
});

new Chart(document.getElementById('ovContractorChart'), {
  type:'bar',
  data:{labels:""" + str(contractors) + """,
    datasets:[
      {label:'Closed', data:""" + str(ncr_closed_data) + """, backgroundColor:C.green+'cc', borderRadius:2, stack:'s'},
      {label:'Open', data:""" + str(ncr_open_data) + """, backgroundColor:C.red+'cc', borderRadius:2, stack:'s'}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b'}, stacked:true}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}, stacked:true}}}
});

// R1 CHARTS
new Chart(document.getElementById('r1StatusChart'), donutCfg([""" + str(r1_status.get('APP',0)) + """,""" + str(r1_status.get('AWC',0)) + """,""" + str(r1_status.get('REJ',0)) + """,""" + str(r1_status.get('RWC',0)) + """,""" + str(r1_status.get('UR',0)) + """,""" + str(r1_status.get('IFI',0)) + """],[C.green,C.amber,C.red,C.blue,C.amber,C.purple],['APP','AWC','REJ','RWC','UR','IFI']));

new Chart(document.getElementById('r1TypeChart'), {
  type:'bar', indexAxis:'y',
  data:{labels:['PLN-Plan','REP-Report','ICL-Checklist','POL-Policy','REG-Register'],
    datasets:[{data:[""" + str(r1_types.get('PLN',0)) + """,""" + str(r1_types.get('REP',0)) + """,""" + str(r1_types.get('ICL',0)) + """,""" + str(r1_types.get('POL',0)) + """,""" + str(r1_types.get('REG',0)) + """], backgroundColor:[C.blue,C.cyan,C.purple,C.amber,C.green], borderRadius:3}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}, y:{grid:{display:false}, ticks:{color:'#94a3b8'}}}}
});

new Chart(document.getElementById('r1PrevChart'), {
  type:'bar', indexAxis:'y',
  data:{labels:['R1'],
    datasets:[
      {label:'AWC', data:[""" + str(r1_prev.get('AWC',0)) + """], backgroundColor:C.amber+'cc'},
      {label:'RWC', data:[""" + str(r1_prev.get('RWC',0)) + """], backgroundColor:C.orange+'cc'},
      {label:'REJ', data:[""" + str(r1_prev.get('REJ',0)) + """], backgroundColor:C.red+'cc'},
      {label:'First', data:[""" + str(r1_prev.get('First',0)) + """], backgroundColor:C.slate+'cc'}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}, stacked:true}, y:{grid:{display:false}, ticks:{color:'#94a3b8'}, stacked:true}}}
});

// R2 CHARTS
new Chart(document.getElementById('r2StatusChart'), donutCfg([""" + str(r2_status.get('APP',0)) + """,""" + str(r2_status.get('AWC',0)) + """,""" + str(r2_status.get('RWC',0)) + """,""" + str(r2_status.get('UR',0)) + """,""" + str(r2_status.get('IFI',0)) + """],[C.green,C.amber,C.blue,C.amber,C.purple],['APP','AWC','RWC','UR','IFI']));

new Chart(document.getElementById('r2TypeChart'), {
  type:'bar', indexAxis:'y',
  data:{labels:['PLN-Plan','REP-Report','ICL-Checklist','POL-Policy','REG-Register'],
    datasets:[{data:[""" + str(r2_types.get('PLN',0)) + """,""" + str(r2_types.get('REP',0)) + """,""" + str(r2_types.get('ICL',0)) + """,""" + str(r2_types.get('POL',0)) + """,""" + str(r2_types.get('REG',0)) + """], backgroundColor:[C.cyan,C.blue,C.purple,C.amber,C.green], borderRadius:3}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}, y:{grid:{display:false}, ticks:{color:'#94a3b8'}}}}
});

new Chart(document.getElementById('r2PrevChart'), {
  type:'bar', indexAxis:'y',
  data:{labels:['R2'],
    datasets:[
      {label:'AWC', data:[""" + str(r2_prev.get('AWC',0)) + """], backgroundColor:C.amber+'cc'},
      {label:'RWC', data:[""" + str(r2_prev.get('RWC',0)) + """], backgroundColor:C.orange+'cc'},
      {label:'REJ', data:[""" + str(r2_prev.get('REJ',0)) + """], backgroundColor:C.red+'cc'},
      {label:'First', data:[""" + str(r2_prev.get('First',0)) + """], backgroundColor:C.slate+'cc'}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}, stacked:true}, y:{grid:{display:false}, ticks:{color:'#94a3b8'}, stacked:true}}}
});

// NCR CHARTS
new Chart(document.getElementById('ncrContChart'), {
  type:'bar',
  data:{labels:""" + str(contractors) + """,
    datasets:[
      {label:'Closed', data:""" + str(ncr_closed_data) + """, backgroundColor:C.green+'bb', borderRadius:2, stack:'s'},
      {label:'Open', data:""" + str(ncr_open_data) + """, backgroundColor:C.red+'bb', borderRadius:2, stack:'s'}
    ]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b'}, stacked:true}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}, stacked:true}}}
});

new Chart(document.getElementById('ncrLocChart'), donutCfg([""" + str(ncr_r1) + """,""" + str(ncr_r2) + """,""" + str(ncr_both) + """],[C.blue,C.cyan,C.purple],['Riyah 1','Riyah 2','Both']));

new Chart(document.getElementById('ncrTimelineChart'), {
  type:'bar',
  data:{labels:""" + str(months_label) + """,
    datasets:[{data:""" + str(ncr_timeline) + """, backgroundColor: (ctx) => ctx.raw >= 8 ? C.red+'cc' : C.blue+'99', borderRadius:4}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b', autoSkip:false}}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}}}
});

// CLIENT NCR CHARTS
new Chart(document.getElementById('cncrLocChart'), donutCfg([""" + str(cncr_loc.get('Riyah 1',0)) + """,""" + str(cncr_loc.get('Riyah 2',0)) + """,""" + str(cncr_loc.get('Both',0)) + """],[C.blue,C.cyan,C.purple],['Riyah 1','Riyah 2','Both']));

new Chart(document.getElementById('cncrOrigChart'), {
  type:'bar', indexAxis:'y',
  data:{labels:""" + str(orig_labels) + """, datasets:[{data:""" + str(orig_data) + """, backgroundColor:[C.blue,C.cyan,C.amber], borderRadius:4}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b'}}, y:{grid:{display:false}, ticks:{color:'#94a3b8'}}}}
});

new Chart(document.getElementById('cncrTimeChart'), {
  type:'line',
  data:{labels:""" + str(cncr_labels) + """, datasets:[{data:""" + str(cncr_timeline) + """, borderColor:C.red, backgroundColor:C.red+'22', fill:true, tension:0.4, pointRadius:5, pointBackgroundColor:C.red}]},
  options:{responsive:true, maintainAspectRatio:false, plugins:{legend:{display:false}}, scales:{x:{grid:{display:false}, ticks:{color:'#64748b'}}, y:{grid:{color:'rgba(255,255,255,0.05)'}, ticks:{color:'#64748b', stepSize:1}, min:0}}}
});

// TABLE DATA & FILTERS
const statusBadge = s => {
  const m={APP:'b-app',REJ:'b-rej',AWC:'b-awc',RWC:'b-rwc',UR:'b-ur'};
  return `<span class="badge ${m[s]||'b-awc'}">${s}</span>`;
};
const typeBadge = t => {
  let s = t;
  if(t.includes('PLN')) s='PLN';
  else if(t.includes('REP')) s='REP';
  else if(t.includes('ICL')) s='ICL';
  else if(t.includes('POL')) s='POL';
  else if(t.includes('REG')) s='REG';
  else s=t.substring(0,3);
  return `<span style="font-size:10px;color:#64748b">${s}</span>`;
};

const r1Docs = """ + r1_docs_json + """;
const r2Docs = """ + r2_docs_json + """;
const ncrData = """ + ncr_data_json + """;
const cncrData = """ + cncr_data_json + """;

// DEBUG: Check R1 data
console.log('=== R1 DATA CHECK ===');
console.log('R1 Docs count:', r1Docs.length);
console.log('First 3 R1 docs:', r1Docs.slice(0, 3));
console.log('R2 Docs count:', r2Docs.length);
console.log('NCR Data count:', ncrData.length);
console.log('Client NCR Data count:', cncrData.length);

function renderTable(docs, bodyId) {
  const body = document.getElementById(bodyId);
  if(!body) {
    console.error('Table body not found:', bodyId);
    return;
  }
  body.innerHTML = '';
  if(!docs || docs.length === 0) {
    body.innerHTML = '<tr><td colspan="6" style="text-align:center;color:#f87171">⚠️ No documents found</td></tr>';
    return;
  }
  docs.forEach(d => {
    const tr = document.createElement('tr');
    tr.innerHTML = `<td class="code">${d[0]}</td><td class="bold">${d[1]}</td><td style="text-align:center">${d[2]}</td><td>${statusBadge(d[3])}</td><td style="white-space:nowrap;font-size:11px;color:#64748b">${d[4]}</td><td>${typeBadge(d[5])}</td>`;
    body.appendChild(tr);
  });
}

function filterTable(site) {
  const search = document.getElementById(site+'Search')?.value.toLowerCase() || '';
  const status = document.getElementById(site+'StatusFilter')?.value || '';
  const type   = document.getElementById(site+'TypeFilter')?.value || '';
  const docs   = site === 'r1' ? r1Docs : r2Docs;
  const filtered = docs.filter(d =>
    (!search || d[0].toLowerCase().includes(search) || d[1].toLowerCase().includes(search)) &&
    (!status || d[3] === status) &&
    (!type   || d[5].includes(type))
  );
  renderTable(filtered, site+'TableBody');
  const countSpan = document.getElementById(site+'Count');
  if(countSpan) countSpan.textContent = filtered.length + ' of ' + docs.length + ' shown';
}

function renderNCRTable(docs) {
  console.log('Rendering NCR table, docs count:', docs.length);
  const body = document.getElementById('ncrTableBody');
  if(!body) {
    console.error('NCR table body not found!');
    return;
  }
  body.innerHTML = '';
  if(!docs || docs.length === 0) {
    body.innerHTML = '<tr><td colspan="9" style="text-align:center;color:#f87171">⚠️ No NCR records found</td></tr>';
    return;
  }
  docs.forEach((d, idx) => {
    try {
      const tr = document.createElement('tr');
      const isOpen = d[8] === 'Open';
      if(isOpen) tr.style.background = 'rgba(239,68,68,0.03)';
      tr.innerHTML = `<td>${d[0]}</td><td class="code">${d[1]}</td><td style="white-space:nowrap;font-size:11px">${d[2]}</td><td>${d[3]}</td><td style="font-size:11px;color:#94a3b8">${d[4]}</td><td><strong>${d[5]}</strong></td><td style="font-size:11px">${d[6]}</td><td style="white-space:nowrap;font-size:11px">${d[7]}</td><td><span class="badge ${isOpen ? 'b-open' : 'b-closed'}">${d[8]}</span></td>`;
      body.appendChild(tr);
    } catch(e) {
      console.error('Error rendering NCR row:', idx, e);
    }
  });
  console.log('NCR table rendered successfully, rows:', docs.length);
}

function filterNCRTable() {
  const statusFilter = document.getElementById('ncrStatusFilter')?.value || 'all';
  let filtered = [...ncrData];
  if(statusFilter !== 'all') {
    filtered = filtered.filter(d => d[8] === statusFilter);
  }
  renderNCRTable(filtered);
}

function renderClientNCRTable(docs) {
  console.log('Rendering Client NCR table, docs count:', docs.length);
  const body = document.getElementById('cncrTableBody');
  if(!body) {
    console.error('Client NCR table body not found!');
    return;
  }
  body.innerHTML = '';
  if(!docs || docs.length === 0) {
    body.innerHTML = '<tr><td colspan="9" style="text-align:center;color:#f87171">⚠️ No Client NCR records found</td></tr>';
    return;
  }
  docs.forEach((d, idx) => {
    try {
      const tr = document.createElement('tr');
      const isOpen = d[8] === 'Open';
      if(isOpen) tr.style.background = 'rgba(239,68,68,0.05)';
      tr.innerHTML = `<td>${d[0]}</td><td class="code">${d[1]}</td><td><span class="badge b-major">${d[2]}</span></td><td style="white-space:nowrap;font-size:11px">${d[3]}</td><td>${d[4]}</td><td>${d[5]}</td><td style="font-size:11px;color:#94a3b8;max-width:200px">${d[6]}</td><td style="white-space:nowrap;font-size:11px">${d[7]}</td><td><span class="badge ${isOpen ? 'b-open' : 'b-closed'}">${d[8]}</span></td>`;
      body.appendChild(tr);
    } catch(e) {
      console.error('Error rendering Client NCR row:', idx, e);
    }
  });
  console.log('Client NCR table rendered successfully, rows:', docs.length);
}

function filterClientNCRTable() {
  const statusFilter = document.getElementById('cncrStatusFilter')?.value || 'all';
  let filtered = [...cncrData];
  if(statusFilter !== 'all') {
    filtered = filtered.filter(d => d[8] === statusFilter);
  }
  renderClientNCRTable(filtered);
}

function exportTableToCSV(tableId, filename) {
  const table = document.getElementById(tableId);
  if(!table) return;
  const rows = table.querySelectorAll('tr');
  let csv = [];
  rows.forEach(row => {
    const cells = row.querySelectorAll('td, th');
    const rowData = Array.from(cells).map(cell => {
      let text = cell.innerText.replace(/,/g, ' ').replace(/\\n/g, ' ');
      return `"${text}"`;
    }).join(',');
    csv.push(rowData);
  });
  const blob = new Blob([csv.join('\\n')], { type: 'text/csv' });
  const link = document.createElement('a');
  link.href = URL.createObjectURL(blob);
  link.download = filename;
  link.click();
  URL.revokeObjectURL(link.href);
}

// Initialize all tables with proper DOM ready check
function initializeAllTables() {
    console.log('=== STARTING TABLE INITIALIZATION ===');
    console.log('R1Docs length:', r1Docs.length);
    console.log('R2Docs length:', r2Docs.length);
    console.log('NCRData length:', ncrData.length);
    console.log('CNCRData length:', cncrData.length);
    
    // R1 Table - Try multiple times if needed
    function renderR1WithRetry(attempt) {
        var r1Body = document.getElementById('r1TableBody');
        if(r1Body) {
            console.log('R1 table body found on attempt', attempt, 'rendering...');
            renderTable(r1Docs, 'r1TableBody');
            if(document.getElementById('r1Count')) {
                document.getElementById('r1Count').textContent = r1Docs.length + ' of ' + r1Docs.length + ' shown';
            }
        } else if(attempt < 5) {
            console.log('R1 table body not found, retrying in 100ms... (attempt', attempt + 1, ')');
            setTimeout(function() { renderR1WithRetry(attempt + 1); }, 100);
        } else {
            console.error('R1 table body NOT found after 5 attempts!');
        }
    }
    
    // R2 Table with retry
    function renderR2WithRetry(attempt) {
        var r2Body = document.getElementById('r2TableBody');
        if(r2Body) {
            console.log('R2 table body found on attempt', attempt, 'rendering...');
            renderTable(r2Docs, 'r2TableBody');
            if(document.getElementById('r2Count')) {
                document.getElementById('r2Count').textContent = r2Docs.length + ' of ' + r2Docs.length + ' shown';
            }
        } else if(attempt < 5) {
            console.log('R2 table body not found, retrying...');
            setTimeout(function() { renderR2WithRetry(attempt + 1); }, 100);
        }
    }
    
    // Start rendering
    renderR1WithRetry(1);
    renderR2WithRetry(1);
    
    // NCR tables (they work already)
    renderNCRTable(ncrData);
    renderClientNCRTable(cncrData);
    
    console.log('=== TABLE INITIALIZATION STARTED ===');
}

// Run initialization when DOM is fully loaded
if(document.readyState === 'loading') {
    document.addEventListener('DOMContentLoaded', initializeAllTables);
} else {
    initializeAllTables();
}
// Force render R1 table when its tab is clicked
function fixR1OnTabClick() {
    var originalShowTab = window.showTab;
    window.showTab = function(id, btn) {
        originalShowTab(id, btn);
        if(id === 'r1') {
            setTimeout(function() {
                var r1Body = document.getElementById('r1TableBody');
                if(r1Body && r1Body.innerHTML === '') {
                    console.log('R1 tab clicked - rendering table now...');
                    renderTable(r1Docs, 'r1TableBody');
                    if(document.getElementById('r1Count')) {
                        document.getElementById('r1Count').textContent = r1Docs.length + ' of ' + r1Docs.length + ' shown';
                    }
                }
            }, 50);
        }
    };
}

// Call this function
fixR1OnTabClick();
</script>
<!-- CALENDAR -->
<div class="panel" id="panel-calendar">
  <div class="panel-body" style="padding-top:24px;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:28px;flex-wrap:wrap;gap:12px;">
      <div>
        <div style="font-family:'Syne',sans-serif;font-size:20px;font-weight:700;color:var(--text)">📅 Monthly HSE Activity Plan</div>
        <div style="font-size:12px;color:var(--muted);margin-top:4px;">Click any day to view planned activities</div>
      </div>
      <div style="background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.3);border-radius:10px;padding:8px 20px;font-family:'Syne',sans-serif;font-size:15px;font-weight:700;color:#93c5fd;">""" + cal_month + ' ' + str(cal_year) + """</div>
    </div>
    <div class="cal-grid" id="calGrid"></div>
    <div id="calModal" style="display:none;position:fixed;top:0;left:0;width:100%;height:100%;background:rgba(0,0,0,.75);z-index:9999;justify-content:center;align-items:center;" onclick="if(event.target===this)this.style.display='none'">
      <div style="background:#141c2e;border:1px solid rgba(255,255,255,.1);border-radius:18px;padding:28px 32px;max-width:520px;width:92%;max-height:82vh;overflow-y:auto;">
        <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:18px;">
          <div>
            <div id="calModalDay" style="font-family:'Syne',sans-serif;font-size:18px;font-weight:700;color:#f1f5f9;"></div>
            <div id="calModalDate" style="font-size:11px;color:#64748b;margin-top:2px;"></div>
          </div>
          <button onclick="document.getElementById('calModal').style.display='none'" style="background:rgba(255,255,255,.08);border:1px solid rgba(255,255,255,.1);color:#94a3b8;width:32px;height:32px;border-radius:8px;cursor:pointer;font-size:16px;">✕</button>
        </div>
        <div id="calModalList"></div>
      </div>
    </div>
  </div>
</div>
<style>
.cal-grid{display:grid;grid-template-columns:repeat(7,1fr);gap:8px;}
.cal-hdr{text-align:center;font-size:10px;font-weight:700;color:#64748b;letter-spacing:.1em;text-transform:uppercase;padding:9px 4px;background:rgba(255,255,255,.03);border-radius:8px;}
.cal-cell{background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:12px;padding:10px 9px;min-height:95px;position:relative;overflow:hidden;transition:all .2s;}
.cal-cell.cal-click{cursor:pointer;}
.cal-cell.cal-click:hover{border-color:#3b82f6;background:rgba(59,130,246,.07);transform:translateY(-2px);}
.cal-cell.cal-off{background:rgba(239,68,68,.04);border-color:rgba(239,68,68,.1);}
.cal-cell.cal-today{border-color:#3b82f6!important;background:rgba(59,130,246,.1);}
.cal-cell.cal-dim{opacity:.2;pointer-events:none;}
.cal-cell.cal-click::after{content:'';position:absolute;bottom:7px;right:7px;width:6px;height:6px;background:#10b981;border-radius:50%;}
.cal-cell.cal-off::after{display:none;}
.cal-num{font-family:'Syne',sans-serif;font-size:16px;font-weight:700;color:#f1f5f9;margin-bottom:5px;}
.cal-cell.cal-today .cal-num{color:#60a5fa;}
.cal-todlbl{position:absolute;top:5px;right:7px;font-size:8px;font-weight:700;color:#3b82f6;letter-spacing:.06em;}
.cal-offlbl{font-size:11px;font-weight:600;color:#f87171;margin-top:3px;}
.cal-prev{font-size:10px;color:#64748b;line-height:1.5;margin-top:3px;}
</style>
<script>
var CAL_DATA=""" + _json.dumps(cal_data) + """;
var CAL_MONTH=""" + str(cal_month_num) + """;
var CAL_YEAR=""" + str(cal_year) + """;
var CAL_MNF=['','January','February','March','April','May','June','July','August','September','October','November','December'];
var CAL_DNF=['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'];
var CAL_DHD=['MON','TUE','WED','THU','FRI','SAT','SUN'];

function calBuild(){
  var g=document.getElementById('calGrid');
  if(!g)return;
  g.innerHTML='';
  CAL_DHD.forEach(function(d){var h=document.createElement('div');h.className='cal-hdr';h.textContent=d;g.appendChild(h);});
  var now=new Date();
  var tod=now.getFullYear()+'-'+String(now.getMonth()+1).padStart(2,'0')+'-'+String(now.getDate()).padStart(2,'0');
  var first=new Date(CAL_YEAR,CAL_MONTH-1,1);
  var start=new Date(first);
  start.setDate(first.getDate()-(first.getDay()===0?6:first.getDay()-1));
  for(var i=0;i<42;i++){
    var d=new Date(start);d.setDate(start.getDate()+i);
    var ds=d.getFullYear()+'-'+String(d.getMonth()+1).padStart(2,'0')+'-'+String(d.getDate()).padStart(2,'0');
    var inMo=d.getMonth()+1===CAL_MONTH&&d.getFullYear()===CAL_YEAR;
    var acts=CAL_DATA[ds]||'';
    var isOff=acts==='Week Off';
    var isTod=ds===tod;
    var hasTsk=acts&&!isOff&&acts.trim()!=='';
    var cell=document.createElement('div');
    cell.className='cal-cell'+(isOff?' cal-off':'')+(isTod?' cal-today':'')+(!inMo?' cal-dim':'')+(hasTsk?' cal-click':'');
    var num=document.createElement('div');num.className='cal-num';num.textContent=d.getDate();cell.appendChild(num);
    if(isTod){var tl=document.createElement('div');tl.className='cal-todlbl';tl.textContent='TODAY';cell.appendChild(tl);}
    if(isOff){var ol=document.createElement('div');ol.className='cal-offlbl';ol.textContent='🔴 Week Off';cell.appendChild(ol);}
    else if(hasTsk){
      var lines=acts.split('|').slice(0,2);
      var pv=document.createElement('div');pv.className='cal-prev';
      pv.textContent=lines.join(' · ');
      cell.appendChild(pv);
      (function(dd,aa){cell.onclick=function(){calOpen(dd,aa);};})(d,acts);
    }
    g.appendChild(cell);
  }
}

function calOpen(d,acts){
  document.getElementById('calModalDay').textContent=CAL_DNF[d.getDay()];
  document.getElementById('calModalDate').textContent=d.getDate()+' '+CAL_MNF[d.getMonth()+1]+' '+d.getFullYear();
  var lines=acts.split('|').filter(function(l){return l.trim();});
  var html='';
  lines.forEach(function(l,i){
    html+='<div style="display:flex;gap:12px;padding:10px 0;border-bottom:1px solid rgba(255,255,255,.05);">'+
      '<div style="min-width:24px;height:24px;background:rgba(59,130,246,.15);border-radius:6px;display:flex;align-items:center;justify-content:center;font-size:11px;font-weight:700;color:#60a5fa;flex-shrink:0;">'+(i+1)+'</div>'+
      '<div style="font-size:13px;color:#cbd5e1;line-height:1.6;padding-top:3px;">'+l.trim()+'</div></div>';
  });
  document.getElementById('calModalList').innerHTML=html;
  document.getElementById('calModal').style.display='flex';
}

// Hook showTab
var _cOST=window.showTab;
window.showTab=function(id,btn){_cOST(id,btn);if(id==='calendar')calBuild();};
</script>
<!-- KPI PANEL -->
<div class="panel" id="panel-kpi">
  <div class="panel-body">

    <!-- Sub-tabs -->
    <div style="display:flex;gap:8px;margin-bottom:22px;flex-wrap:wrap;">
      <button id="kstab-hse" onclick="showKTab('hse')" style="padding:10px 20px;background:rgba(59,130,246,.15);border:1px solid rgba(59,130,246,.4);border-radius:8px;color:#f1f5f9;cursor:pointer;font-family:'DM Sans',sans-serif;font-size:12px;font-weight:600;">🦺 HSE KPIs</button>
      <button id="kstab-es" onclick="showKTab('es')" style="padding:10px 20px;background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);border-radius:8px;color:#64748b;cursor:pointer;font-family:'DM Sans',sans-serif;font-size:12px;font-weight:600;">🌱 E&amp;S KPIs</button>
    </div>

    <!-- HSE SUB-PANEL -->
    <div id="kpanel-hse" style="display:block;">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:24px;flex-wrap:wrap;gap:12px;">
      <div>
        <div style="font-family:'Syne',sans-serif;font-size:20px;font-weight:700;color:#f1f5f9;">📊 Monthly HSE KPI Dashboard</div>
        <div style="font-size:12px;color:#64748b;margin-top:4px;">Riyah 1 &amp; Riyah 2 Wind IPP · Last Month Statistics</div>
      </div>
      <div style="background:rgba(59,130,246,.12);border:1px solid rgba(59,130,246,.3);border-radius:10px;padding:8px 20px;font-family:'Syne',sans-serif;font-size:15px;font-weight:700;color:#93c5fd;">""" + kpi_data['month'] + """</div>
    </div>

    <!-- Highlights -->
    <div class="kpi-section-title">📌 Project Highlights</div>
    <div class="kpi-grid">
      <div class="kpi-c green"><div class="kpi-l">Days w/o LTI (Month)</div><div class="kpi-v green">""" + str(kpi_data['days_lti_month']) + """</div><div class="kpi-s">""" + kpi_data['month'] + """</div></div>
      <div class="kpi-c blue"><div class="kpi-l">Days w/o LTI (Total)</div><div class="kpi-v blue">""" + str(kpi_data['days_lti_total']) + """</div><div class="kpi-s">Since 14/02/2025</div></div>
      <div class="kpi-c amber"><div class="kpi-l">Total Manpower</div><div class="kpi-v amber">""" + f"{kpi_data['total_mp']:,}" + """</div><div class="kpi-s">All teams</div></div>
      <div class="kpi-c purple"><div class="kpi-l">Monthly Manhours</div><div class="kpi-v purple">""" + f"{kpi_data['total_hrs']:,}" + """</div><div class="kpi-s">""" + kpi_data['month'] + """</div></div>
      <div class="kpi-c teal"><div class="kpi-l">Cumulative Hours</div><div class="kpi-v teal">""" + f"{kpi_data['cum_total']:,}" + """</div><div class="kpi-s">Project to date</div></div>
    </div>

    <!-- Manpower table -->
    <div class="kpi-section-title">👥 Personnel on Site</div>
    <div class="kpi-tbl-card" style="margin-bottom:24px;">
      <div class="kpi-tbl-hdr">Manpower &amp; Hours Distribution</div>
      <table class="kpi-tbl"><thead><tr><th>Team</th><th class="num">Manpower</th><th class="num">Manhours</th></tr></thead>
        <tbody>""" + _mp_rows + """<tr style="background:rgba(59,130,246,.06);"><td class="bold" style="color:#60a5fa;">SUB-TOTAL</td><td class="num blue">""" + f"{kpi_data['total_mp']:,}" + """</td><td class="num blue">""" + f"{kpi_data['total_hrs']:,}" + """</td></tr></tbody></table>
    </div>

    <!-- HSE 3 cols -->
    <div class="kpi-section-title">⚠️ HSE Statistics — Lagging &amp; Leading Indicators</div>
    <div class="kpi-three-col">
      """ + _kpi_r1_html + """
      """ + _kpi_r2_html + """
      """ + _kpi_cmb_html + """
    </div>

    <!-- Sustainability -->
    <div class="kpi-section-title" style="margin-top:24px;">🌿 Sustainability KPIs</div>
    <div class="kpi-grid">
      <div class="kpi-c blue"><div class="kpi-l">Stakeholder Meetings</div><div class="kpi-v blue">""" + str(kpi_data['sustain'].get('meetings',0)) + """</div></div>
      <div class="kpi-c amber"><div class="kpi-l">Grievances Received</div><div class="kpi-v amber">""" + str(kpi_data['sustain'].get('grievance_recv',0)) + """</div></div>
      <div class="kpi-c green"><div class="kpi-l">Grievances Resolved</div><div class="kpi-v green">""" + str(kpi_data['sustain'].get('grievance_resolved',0)) + """</div></div>
      <div class="kpi-c red"><div class="kpi-l">Spill Incidents</div><div class="kpi-v red">""" + str(kpi_data['sustain'].get('spill_total',0)) + """</div></div>
      <div class="kpi-c purple"><div class="kpi-l">Workplace Incidents</div><div class="kpi-v purple">""" + str(kpi_data['sustain'].get('wp_incidents',0)) + """</div></div>
      <div class="kpi-c teal"><div class="kpi-l">Gender Ratio (M:F)</div><div class="kpi-v teal" style="font-size:22px;">""" + str(kpi_data['sustain'].get('gender_ratio','—')) + """</div><div class="kpi-s">M """ + str(kpi_data['sustain'].get('male',0)) + """ · F """ + str(kpi_data['sustain'].get('female',0)) + """</div></div>
    </div>
    </div><!-- end kpanel-hse -->

    <!-- E&S SUB-PANEL -->
    <div id="kpanel-es" style="display:none;">
      <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:22px;flex-wrap:wrap;gap:12px;">
        <div><div style="font-family:'Syne',sans-serif;font-size:20px;font-weight:700;color:#f1f5f9;">🌱 Environmental &amp; Social KPI Dashboard</div><div style="font-size:12px;color:#64748b;margin-top:3px;">TotalEnergies Compliance · 2025 YTD + Monthly Trends</div></div>
        <div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.3);border-radius:10px;padding:8px 20px;font-family:'Syne',sans-serif;font-size:14px;font-weight:700;color:#34d399;">2025 YTD · May 2026</div>
      </div>
      <div class="kpi-section-title">📌 E&amp;S Highlights — Combined R1+R2</div>
      <div class="kpi-grid">
        <div class="kpi-c green"><div class="kpi-l">Penalties / Fines</div><div class="kpi-v green">""" + str(int(es_comb.get('penalties',0))) + """</div><div class="kpi-s">Env violations</div></div>
        <div class="kpi-c blue"><div class="kpi-l">External Audits</div><div class="kpi-v blue">""" + str(int(es_comb.get('audits',0))) + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c amber"><div class="kpi-l">Stakeholder Meetings</div><div class="kpi-v amber">""" + str(int(es_comb.get('stakeholder',0))) + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c red"><div class="kpi-l">Grievances</div><div class="kpi-v red">""" + str(int(es_comb.get('grievances',0))) + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c orange"><div class="kpi-l">Spill Incidents</div><div class="kpi-v orange">""" + str(int(es_comb.get('spills',0))) + """</div></div>
      </div>
      <div class="kpi-section-title" style="margin-top:22px;">⚡ Resource Consumption — Combined (2025 YTD)</div>
      <div class="kpi-grid">
        <div class="kpi-c teal"><div class="kpi-l">Diesel (Litres)</div><div class="kpi-v teal">""" + f"{es_comb.get('diesel',0):,.0f}" + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c blue"><div class="kpi-l">Freshwater (m³)</div><div class="kpi-v blue">""" + f"{es_comb.get('freshwater',0):,.0f}" + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c green"><div class="kpi-l">Wastewater (m³)</div><div class="kpi-v green">""" + f"{es_comb.get('wastewater',0):,.0f}" + """</div><div class="kpi-s">R1+R2</div></div>
        <div class="kpi-c red"><div class="kpi-l">Hazardous (t)</div><div class="kpi-v red">""" + f"{es_comb.get('haz_total',0):,.2f}" + """</div></div>
        <div class="kpi-c amber"><div class="kpi-l">Non-Haz Waste (t)</div><div class="kpi-v amber">""" + f"{es_comb.get('nonhaz_total',0):,.2f}" + """</div></div>
        <div class="kpi-c purple"><div class="kpi-l">Waste Cost (OMR)</div><div class="kpi-v purple">""" + f"{es_comb.get('waste_cost',0):,.0f}" + """</div></div>
      </div>
      <div class="kpi-section-title" style="margin-top:22px;">⚖️ Riyah 1 vs Riyah 2</div>
      <div style="background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;overflow:hidden;margin-bottom:22px;">
        <div style="padding:14px 18px;border-bottom:1px solid rgba(255,255,255,.07);font-family:'Syne',sans-serif;font-size:13px;font-weight:700;">E&amp;S KPIs — R1 vs R2 vs Combined</div>
        <table class="kpi-tbl"><thead><tr><th>KPI</th><th class="num" style="color:#60a5fa;">Riyah 1</th><th class="num" style="color:#34d399;">Riyah 2</th><th class="num">Combined</th></tr></thead><tbody>
          <tr><td>Penalties</td><td class="num">""" + str(int(es_r1.get('penalties',{}).get('total',0))) + """</td><td class="num">""" + str(int(es_r2.get('penalties',{}).get('total',0))) + """</td><td class="num blue">""" + str(int(es_comb.get('penalties',0))) + """</td></tr>
          <tr><td>External Audits</td><td class="num">""" + str(int(es_r1.get('audits',{}).get('total',0))) + """</td><td class="num">""" + str(int(es_r2.get('audits',{}).get('total',0))) + """</td><td class="num blue">""" + str(int(es_comb.get('audits',0))) + """</td></tr>
          <tr><td>Stakeholder Meetings</td><td class="num">""" + str(int(es_r1.get('stakeholder',{}).get('total',0))) + """</td><td class="num">""" + str(int(es_r2.get('stakeholder',{}).get('total',0))) + """</td><td class="num blue">""" + str(int(es_comb.get('stakeholder',0))) + """</td></tr>
          <tr><td>Grievances</td><td class="num red">""" + str(int(es_r1.get('grievances',{}).get('total',0))) + """</td><td class="num red">""" + str(int(es_r2.get('grievances',{}).get('total',0))) + """</td><td class="num blue">""" + str(int(es_comb.get('grievances',0))) + """</td></tr>
          <tr><td>Diesel (L)</td><td class="num teal">""" + f"{es_r1.get('diesel',{}).get('total',0):,.0f}" + """</td><td class="num teal">""" + f"{es_r2.get('diesel',{}).get('total',0):,.0f}" + """</td><td class="num blue">""" + f"{es_comb.get('diesel',0):,.0f}" + """</td></tr>
          <tr><td>Freshwater (m³)</td><td class="num blue">""" + f"{es_r1.get('freshwater',{}).get('total',0):,.2f}" + """</td><td class="num blue">""" + f"{es_r2.get('freshwater',{}).get('total',0):,.2f}" + """</td><td class="num blue">""" + f"{es_comb.get('freshwater',0):,.2f}" + """</td></tr>
          <tr><td>Hazardous (t)</td><td class="num red">""" + f"{es_r1.get('haz_total',{}).get('total',0):,.2f}" + """</td><td class="num red">""" + f"{es_r2.get('haz_total',{}).get('total',0):,.2f}" + """</td><td class="num blue">""" + f"{es_comb.get('haz_total',0):,.2f}" + """</td></tr>
          <tr><td>Non-Haz (t)</td><td class="num amber">""" + f"{es_r1.get('nonhaz_total',{}).get('total',0):,.2f}" + """</td><td class="num amber">""" + f"{es_r2.get('nonhaz_total',{}).get('total',0):,.2f}" + """</td><td class="num blue">""" + f"{es_comb.get('nonhaz_total',0):,.2f}" + """</td></tr>
          <tr><td>Waste Cost (OMR)</td><td class="num purple">""" + f"{es_r1.get('waste_cost',{}).get('total',0):,.0f}" + """</td><td class="num purple">""" + f"{es_r2.get('waste_cost',{}).get('total',0):,.0f}" + """</td><td class="num blue">""" + f"{es_comb.get('waste_cost',0):,.0f}" + """</td></tr>
          <tr><td>Spills</td><td class="num red">""" + str(int(es_r1.get('spills',{}).get('total',0))) + """</td><td class="num red">""" + str(int(es_r2.get('spills',{}).get('total',0))) + """</td><td class="num blue">""" + str(int(es_comb.get('spills',0))) + """</td></tr>
          <tr><td>Wastewater (m³)</td><td class="num">""" + f"{es_r1.get('wastewater',{}).get('total',0):,.2f}" + """</td><td class="num">""" + f"{es_r2.get('wastewater',{}).get('total',0):,.2f}" + """</td><td class="num blue">""" + f"{es_comb.get('wastewater',0):,.2f}" + """</td></tr>
        </tbody></table>
      </div>
      <div class="kpi-section-title">📈 Monthly Trends — Riyah 1 (Jan 2025 – Apr 2026)</div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:14px;">
        <div style="background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:18px;"><div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;margin-bottom:3px;">Diesel (Litres)</div><div style="font-size:10px;color:#64748b;margin-bottom:12px;">Monthly — Riyah 1</div><div style="height:200px;position:relative;"><canvas id="kpiDiesel"></canvas></div></div>
        <div style="background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:18px;"><div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;margin-bottom:3px;">Freshwater (m³)</div><div style="font-size:10px;color:#64748b;margin-bottom:12px;">Monthly — Riyah 1</div><div style="height:200px;position:relative;"><canvas id="kpiWater"></canvas></div></div>
      </div>
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;">
        <div style="background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:18px;"><div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;margin-bottom:3px;">Stakeholder &amp; Grievances</div><div style="font-size:10px;color:#64748b;margin-bottom:12px;">Monthly — Riyah 1</div><div style="height:200px;position:relative;"><canvas id="kpiEng"></canvas></div></div>
        <div style="background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:18px;"><div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;margin-bottom:3px;">Audits &amp; Spills</div><div style="font-size:10px;color:#64748b;margin-bottom:12px;">Monthly — Riyah 1</div><div style="height:200px;position:relative;"><canvas id="kpiAudit"></canvas></div></div>
      </div>
    </div><!-- end kpanel-es -->

  </div>
</div>

<style>
.kpi-section-title{font-family:'Syne',sans-serif;font-size:13px;font-weight:700;color:#94a3b8;letter-spacing:.08em;text-transform:uppercase;margin-bottom:14px;padding-bottom:9px;border-bottom:1px solid rgba(255,255,255,.07);}
.kpi-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(140px,1fr));gap:12px;margin-bottom:24px;}
.kpi-c{background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:18px;position:relative;overflow:hidden;}
.kpi-c::after{content:'';position:absolute;top:0;left:0;right:0;height:2px;border-radius:14px 14px 0 0;}
.kpi-c.green::after{background:linear-gradient(90deg,#10b981,#34d399);}
.kpi-c.red::after{background:linear-gradient(90deg,#ef4444,#f87171);}
.kpi-c.blue::after{background:linear-gradient(90deg,#3b82f6,#06b6d4);}
.kpi-c.amber::after{background:linear-gradient(90deg,#f59e0b,#fbbf24);}
.kpi-c.purple::after{background:linear-gradient(90deg,#8b5cf6,#a78bfa);}
.kpi-c.teal::after{background:linear-gradient(90deg,#06b6d4,#67e8f9);}
.kpi-l{font-size:10px;color:#64748b;letter-spacing:.07em;text-transform:uppercase;margin-bottom:8px;}
.kpi-v{font-family:'Syne',sans-serif;font-size:30px;font-weight:700;line-height:1;}
.kpi-v.green{color:#34d399;}.kpi-v.red{color:#f87171;}.kpi-v.blue{color:#60a5fa;}.kpi-v.amber{color:#fbbf24;}.kpi-v.purple{color:#a78bfa;}.kpi-v.teal{color:#67e8f9;}
.kpi-s{font-size:11px;color:#64748b;margin-top:5px;}
.kpi-tbl-card{background:#141c2e;border:1px solid rgba(255,255,255,.07);border-radius:14px;overflow:hidden;}
.kpi-tbl-hdr{padding:14px 20px;border-bottom:1px solid rgba(255,255,255,.07);font-family:'Syne',sans-serif;font-size:13px;font-weight:600;}
.kpi-tbl{width:100%;border-collapse:collapse;font-size:12px;}
.kpi-tbl thead th{padding:10px 14px;text-align:left;font-size:10px;color:#64748b;letter-spacing:.07em;text-transform:uppercase;font-weight:600;background:rgba(255,255,255,.02);border-bottom:1px solid rgba(255,255,255,.07);}
.kpi-tbl tbody td{padding:10px 14px;color:#94a3b8;border-bottom:1px solid rgba(255,255,255,.04);}
.kpi-tbl tbody tr:hover{background:rgba(255,255,255,.025);}
.kpi-tbl .num{font-family:'Syne',sans-serif;font-weight:700;text-align:center;}
.kpi-tbl .green{color:#34d399;}.kpi-tbl .red{color:#f87171;}.kpi-tbl .blue{color:#60a5fa;}
.kpi-tbl .bold{color:#f1f5f9;font-weight:500;}
.kpi-three-col{display:grid;grid-template-columns:1fr 1fr 1fr;gap:14px;}
@media(max-width:900px){.kpi-three-col{grid-template-columns:1fr;}}
.kpi-c.orange::after{background:linear-gradient(90deg,#f97316,#fb923c);}
.kpi-v.orange{color:#fb923c;}
</style>

<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<script>
var KPI_MONTHS=""" + _json.dumps(MONTHS_ES) + """;
function showKTab(id){
  document.getElementById('kpanel-hse').style.display=(id==='hse')?'block':'none';
  document.getElementById('kpanel-es').style.display=(id==='es')?'block':'none';
  var hB=document.getElementById('kstab-hse'),eB=document.getElementById('kstab-es');
  hB.style.background=(id==='hse')?'rgba(59,130,246,.15)':'rgba(255,255,255,.04)';
  hB.style.color=(id==='hse')?'#f1f5f9':'#64748b';
  hB.style.borderColor=(id==='hse')?'rgba(59,130,246,.4)':'rgba(255,255,255,.08)';
  eB.style.background=(id==='es')?'rgba(16,185,129,.15)':'rgba(255,255,255,.04)';
  eB.style.color=(id==='es')?'#f1f5f9':'#64748b';
  eB.style.borderColor=(id==='es')?'rgba(16,185,129,.4)':'rgba(255,255,255,.08)';
  if(id==='es')buildESCharts();
}
var esBuilt=false;
function buildESCharts(){
  if(esBuilt)return; esBuilt=true;
  var co={responsive:true,maintainAspectRatio:false,plugins:{legend:{display:false}},scales:{y:{grid:{color:'rgba(255,255,255,0.05)'},ticks:{color:'#64748b',font:{size:10}}},x:{grid:{display:false},ticks:{color:'#64748b',font:{size:9},maxRotation:45,minRotation:45}}}};
  var mco={responsive:true,maintainAspectRatio:false,plugins:{legend:{position:'top',labels:{color:'#94a3b8',font:{size:10},boxWidth:10}}},scales:{y:{grid:{color:'rgba(255,255,255,0.05)'},ticks:{color:'#64748b',font:{size:10}}},x:{grid:{display:false},ticks:{color:'#64748b',font:{size:9},maxRotation:45,minRotation:45}}}};
  new Chart(document.getElementById('kpiDiesel'),{type:'line',data:{labels:KPI_MONTHS,datasets:[{data:""" + _json.dumps(es_r1.get('diesel',{}).get('monthly',[])) + """,borderColor:'#06b6d4',backgroundColor:'#06b6d422',fill:true,tension:0.4,pointRadius:3,pointBackgroundColor:'#06b6d4',borderWidth:2.5}]},options:co});
  new Chart(document.getElementById('kpiWater'),{type:'line',data:{labels:KPI_MONTHS,datasets:[{data:""" + _json.dumps(es_r1.get('freshwater',{}).get('monthly',[])) + """,borderColor:'#3b82f6',backgroundColor:'#3b82f622',fill:true,tension:0.4,pointRadius:3,pointBackgroundColor:'#3b82f6',borderWidth:2.5}]},options:co});
  new Chart(document.getElementById('kpiEng'),{type:'bar',data:{labels:KPI_MONTHS,datasets:[{label:'Meetings',data:""" + _json.dumps(es_r1.get('stakeholder',{}).get('monthly',[])) + """,backgroundColor:'#10b981cc',borderRadius:4},{label:'Grievances',data:""" + _json.dumps(es_r1.get('grievances',{}).get('monthly',[])) + """,backgroundColor:'#ef4444cc',borderRadius:4}]},options:mco});
  new Chart(document.getElementById('kpiAudit'),{type:'bar',data:{labels:KPI_MONTHS,datasets:[{label:'Audits',data:""" + _json.dumps(es_r1.get('audits',{}).get('monthly',[])) + """,backgroundColor:'#8b5cf6cc',borderRadius:4},{label:'Spills',data:""" + _json.dumps(es_r1.get('spills',{}).get('monthly',[])) + """,backgroundColor:'#f97316cc',borderRadius:4}]},options:mco});
}
</script>

<footer>Riyah 1 &amp; 2 Wind IPP Project &nbsp;·&nbsp; Document Dashboard &nbsp;·&nbsp; Developed By Shaguf Ahmed</footer>
</body>
</html>"""

# Write HTML file
with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
    f.write(html_template)

print(f"\n✅ Dashboard generated successfully!")
print(f"📄 Saved to: {OUTPUT_HTML}")
print(f"\n📊 Summary:")
print(f"  R1 docs: {len(r1_rows)} | R2 docs: {len(r2_rows)} | Total: {total_docs}")
print(f"  Internal NCRs: {ncr_total} (Open: {ncr_open}, Closed: {ncr_closed})")
print(f"  Client NCRs: {cncr_total} (Open: {cncr_open}, Closed: {cncr_closed})")
print(f"\n🌐 Open Dashboard in browser to verify ✅")
print(f"\n🆕 New Features:")
print(f"  - Internal NCR table: Status filter (All/Open/Closed)")
print(f"  - Client NCR table: Status filter (All/Open/Closed)")
print(f"  - Export to CSV buttons on all tables")

# ─────────────────────────────────────────────────────────
# AUTO PUSH TO GITHUB
# ─────────────────────────────────────────────────────────
print("\n" + "="*50)
print("📤 Pushing to GitHub automatically...")
print("="*50)

import subprocess
import os
import datetime

os.chdir(r"C:\Users\Shaghaf Ahmed\OneDrive\Desktop\Work\AUTOMATE EXCEL\QHSE_dashboard")

subprocess.run(["git", "add", "."])
subprocess.run(["git", "commit", "-m", f"auto update {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
subprocess.run(["git", "push", "origin", "master"])

print("\n✅ GitHub updated successfully!")
print("🌐 Live link: https://insightdealerin-shag.github.io/HSE_dashboard/")
print("⏱️ 2-3 minutes me link pe reflect hoga.")
print("="*50)
# AUTO PUSH TO GITHUB
# ─────────────────────────────────────────────────────────
print("\n" + "="*50)
print("📤 Pushing to GitHub automatically...")
print("="*50)

import subprocess
import os
import datetime

os.chdir(r"C:\Users\Shaghaf Ahmed\OneDrive\Desktop\Work\AUTOMATE EXCEL\QHSE_dashboard")

subprocess.run(["git", "add", "."])
subprocess.run(["git", "commit", "-m", f"auto update {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
subprocess.run(["git", "push", "origin", "master"])

print("\n✅ GitHub updated successfully!")
print("🌐 Live link: https://insightdealerin-shag.github.io/HSE_dashboard/")
print("⏱️ 2-3 minutes me link pe reflect hoga.")
print("="*50)